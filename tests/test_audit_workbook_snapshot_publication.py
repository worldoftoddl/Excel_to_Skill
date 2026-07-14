"""Durable asset and domain foundations for workbook snapshot publication."""
from __future__ import annotations

import hashlib
import json
import stat
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from excel_to_skill.audit.workbook_snapshot_publication import (
    AcquiredWorkbook,
    LocalImmutableWorkbookAssetStore,
    MAX_PUBLISHED_WORKBOOK_BYTES,
    SnapshotPublicationBasis,
    StoredWorkbookAsset,
    WorkbookSnapshotPublicationError,
    build_snapshot_publication,
    reacquire_and_store_workbook,
    validate_acquired_workbook,
    validate_saved_workbook_manifest,
)


BASE_BYTES = b"synthetic workbook base"
NEW_BYTES = b"synthetic workbook revision with approved edit"
BASE_SHA = hashlib.sha256(BASE_BYTES).hexdigest()
MANIFEST_SHA = "3" * 64


def _basis(**changes) -> SnapshotPublicationBasis:
    values = {
        "bundle_id": "bundle-a",
        "execution_id": "edit-execution-a",
        "manifest_ref": "edit-manifest:" + MANIFEST_SHA,
        "manifest_sha256": MANIFEST_SHA,
        "base_snapshot_id": "1" * 64,
        "base_workbook_sha256": BASE_SHA,
        "base_revision_id": "revision-a",
        "sheet": "C",
        "worksheet_id": "worksheet-a",
        "workbook_instance_id": "private-instance-a",
    }
    values.update(changes)
    return SnapshotPublicationBasis(**values)


def _acquired(**changes) -> AcquiredWorkbook:
    values = {
        "provider_revision_id": "revision-b",
        "predecessor_revision_id": "revision-a",
        "worksheet_id": "worksheet-a",
        "workbook_instance_id": "private-instance-a",
        "content": NEW_BYTES,
    }
    values.update(changes)
    return AcquiredWorkbook(**values)


def _asset_path(root: Path, asset: StoredWorkbookAsset) -> Path:
    return root / "objects" / f"{asset.workbook_sha256}.xlsx"


def _assert_code(code: str, call) -> WorkbookSnapshotPublicationError:
    with pytest.raises(WorkbookSnapshotPublicationError) as caught:
        call()
    assert caught.value.code == code
    return caught.value


def test_local_store_is_content_addressed_private_idempotent_and_restart_safe(
    tmp_path: Path,
) -> None:
    root = tmp_path / "private-assets"
    first_store = LocalImmutableWorkbookAssetStore(root)
    first = first_store.put_if_absent(NEW_BYTES)
    again = first_store.put_if_absent(NEW_BYTES)

    assert first == again
    assert first.asset_ref == "workbook-asset:" + hashlib.sha256(NEW_BYTES).hexdigest()
    assert first_store.read_verified(first) == NEW_BYTES
    assert len(list((root / "objects").glob("*.xlsx"))) == 1

    restarted = LocalImmutableWorkbookAssetStore(root)
    assert restarted.put_if_absent(NEW_BYTES) == first
    assert restarted.read_verified(first) == NEW_BYTES
    assert stat.S_IMODE(root.stat().st_mode) == 0o700
    assert stat.S_IMODE((root / "objects").stat().st_mode) == 0o700
    assert stat.S_IMODE(_asset_path(root, first).stat().st_mode) == 0o600


def test_concurrent_identical_writers_publish_one_immutable_object(tmp_path: Path) -> None:
    root = tmp_path / "assets"
    stores = [LocalImmutableWorkbookAssetStore(root) for _ in range(4)]

    with ThreadPoolExecutor(max_workers=8) as pool:
        results = list(
            pool.map(
                lambda index: stores[index % len(stores)].put_if_absent(NEW_BYTES),
                range(16),
            )
        )

    assert all(result == results[0] for result in results)
    assert len(list((root / "objects").glob("*.xlsx"))) == 1
    assert stores[0].read_verified(results[0]) == NEW_BYTES


def test_local_store_rejects_empty_oversized_and_non_bytes(tmp_path: Path) -> None:
    store = LocalImmutableWorkbookAssetStore(tmp_path / "assets")

    _assert_code("INVALID_ACQUIRED_WORKBOOK", lambda: store.put_if_absent(b""))
    _assert_code(
        "SOURCE_LIMIT_EXCEEDED",
        lambda: store.put_if_absent(b"x" * (MAX_PUBLISHED_WORKBOOK_BYTES + 1)),
    )
    _assert_code("INVALID_ACQUIRED_WORKBOOK", lambda: store.put_if_absent(bytearray(b"x")))


def test_existing_asset_tamper_is_detected_and_never_replaced(tmp_path: Path) -> None:
    root = tmp_path / "assets"
    store = LocalImmutableWorkbookAssetStore(root)
    asset = store.put_if_absent(NEW_BYTES)
    target = _asset_path(root, asset)
    tampered = b"x" * len(NEW_BYTES)
    target.write_bytes(tampered)

    _assert_code("ASSET_INTEGRITY_MISMATCH", lambda: store.read_verified(asset))
    _assert_code("ASSET_INTEGRITY_MISMATCH", lambda: store.put_if_absent(NEW_BYTES))
    assert target.read_bytes() == tampered


def test_asset_with_broadened_permissions_fails_closed(tmp_path: Path) -> None:
    root = tmp_path / "assets"
    store = LocalImmutableWorkbookAssetStore(root)
    asset = store.put_if_absent(NEW_BYTES)
    target = _asset_path(root, asset)
    target.chmod(0o644)

    _assert_code("ASSET_INTEGRITY_MISMATCH", lambda: store.read_verified(asset))
    _assert_code("ASSET_INTEGRITY_MISMATCH", lambda: store.put_if_absent(NEW_BYTES))


def test_asset_symlink_and_symlinked_store_components_fail_closed(tmp_path: Path) -> None:
    root = tmp_path / "assets"
    store = LocalImmutableWorkbookAssetStore(root)
    digest = hashlib.sha256(NEW_BYTES).hexdigest()
    target = root / "objects" / f"{digest}.xlsx"
    outside = tmp_path / "outside.xlsx"
    outside.write_bytes(NEW_BYTES)
    target.symlink_to(outside)

    asset = StoredWorkbookAsset(
        asset_ref="workbook-asset:" + digest,
        workbook_sha256=digest,
        size_bytes=len(NEW_BYTES),
    )
    _assert_code("ASSET_STORE_UNAVAILABLE", lambda: store.read_verified(asset))
    _assert_code("ASSET_STORE_UNAVAILABLE", lambda: store.put_if_absent(NEW_BYTES))
    assert outside.read_bytes() == NEW_BYTES

    linked_root = tmp_path / "linked-root"
    linked_root.symlink_to(root, target_is_directory=True)
    _assert_code(
        "ASSET_STORE_UNAVAILABLE",
        lambda: LocalImmutableWorkbookAssetStore(linked_root),
    )

    root_two = tmp_path / "assets-two"
    root_two.mkdir()
    outside_dir = tmp_path / "outside-dir"
    outside_dir.mkdir()
    (root_two / "objects").symlink_to(outside_dir, target_is_directory=True)
    _assert_code(
        "ASSET_STORE_UNAVAILABLE",
        lambda: LocalImmutableWorkbookAssetStore(root_two),
    )


def test_reacquirer_receives_only_exact_private_binding_and_store_limit(tmp_path: Path) -> None:
    basis = _basis()

    class Reacquirer:
        def __init__(self) -> None:
            self.kwargs = None

        def reacquire_saved_workbook(self, **kwargs):
            self.kwargs = kwargs
            return _acquired()

    reacquirer = Reacquirer()
    store = LocalImmutableWorkbookAssetStore(tmp_path / "assets")
    acquired, stored = reacquire_and_store_workbook(
        basis=basis,
        reacquirer=reacquirer,
        assets=store,
    )

    assert reacquirer.kwargs == {
        "expected_workbook_instance_id": "private-instance-a",
        "base_revision_id": "revision-a",
        "expected_sheet": "C",
        "expected_worksheet_id": "worksheet-a",
        "max_bytes": MAX_PUBLISHED_WORKBOOK_BYTES,
    }
    assert store.read_verified(stored) == acquired.content


def test_reacquirer_failure_is_sanitized_without_private_provider_details(tmp_path: Path) -> None:
    class BrokenReacquirer:
        def reacquire_saved_workbook(self, **kwargs):
            del kwargs
            raise RuntimeError("/private/client.xlsx token=secret")

    with pytest.raises(WorkbookSnapshotPublicationError) as caught:
        reacquire_and_store_workbook(
            basis=_basis(),
            reacquirer=BrokenReacquirer(),
            assets=LocalImmutableWorkbookAssetStore(tmp_path / "assets"),
        )
    assert caught.value.code == "ASSET_STORE_UNAVAILABLE"
    assert "private" not in str(caught.value).lower()
    assert "secret" not in str(caught.value).lower()


@pytest.mark.parametrize(
    ("acquired", "code"),
    [
        (_acquired(workbook_instance_id="copy-instance"), "WORKBOOK_INSTANCE_MISMATCH"),
        (_acquired(worksheet_id="worksheet-copy"), "WORKSHEET_MISMATCH"),
        (
            _acquired(predecessor_revision_id="unrelated-revision"),
            "REVISION_CHAIN_MISMATCH",
        ),
        (_acquired(provider_revision_id="revision-a"), "REVISION_NOT_ADVANCED"),
        (_acquired(content=BASE_BYTES), "WORKBOOK_NOT_CHANGED"),
    ],
)
def test_acquired_workbook_must_advance_exact_instance_revision_and_bytes(
    acquired: AcquiredWorkbook,
    code: str,
) -> None:
    _assert_code(code, lambda: validate_acquired_workbook(_basis(), acquired))


def test_publication_is_deterministic_exact_and_contains_no_private_locator(
    tmp_path: Path,
) -> None:
    basis = _basis()
    acquired = _acquired()
    store = LocalImmutableWorkbookAssetStore(tmp_path / "private-client-path")
    asset = store.put_if_absent(acquired.content)

    first = build_snapshot_publication(basis=basis, acquired=acquired, stored=asset)
    second = build_snapshot_publication(basis=basis, acquired=acquired, stored=asset)

    assert first == second
    assert set(first) == {
        "schema_version",
        "bundle_id",
        "execution_id",
        "manifest_ref",
        "manifest_sha256",
        "base_snapshot_id",
        "base_revision_id",
        "snapshot_id",
        "workbook_sha256",
        "revision_id",
        "asset_persisted",
        "prepared_bundle_created",
    }
    assert first["schema_version"] == "audit_workbook_snapshot_publication.v1"
    assert first["snapshot_id"] != basis.base_snapshot_id
    assert first["workbook_sha256"] != basis.base_workbook_sha256
    assert first["revision_id"] != basis.base_revision_id
    assert first["asset_persisted"] is True
    assert first["prepared_bundle_created"] is False

    serialized = json.dumps(first, ensure_ascii=False, sort_keys=True)
    assert "private-instance-a" not in serialized
    assert "private-client-path" not in serialized
    assert asset.asset_ref not in serialized
    assert "path" not in first
    assert "provider" not in first


def test_publication_rejects_stored_asset_not_bound_to_exact_acquired_bytes() -> None:
    acquired = _acquired()
    other = b"another stored workbook"
    other_sha = hashlib.sha256(other).hexdigest()
    mismatched = StoredWorkbookAsset(
        asset_ref="workbook-asset:" + other_sha,
        workbook_sha256=other_sha,
        size_bytes=len(other),
    )

    _assert_code(
        "ASSET_INTEGRITY_MISMATCH",
        lambda: build_snapshot_publication(
            basis=_basis(),
            acquired=acquired,
            stored=mismatched,
        ),
    )


def test_publication_rejects_same_derived_snapshot_identity(monkeypatch, tmp_path: Path) -> None:
    import excel_to_skill.audit.workbook_snapshot_publication as publication_module

    basis = _basis()
    acquired = _acquired()
    asset = LocalImmutableWorkbookAssetStore(tmp_path / "assets").put_if_absent(
        acquired.content
    )
    monkeypatch.setattr(
        publication_module,
        "_snapshot_id",
        lambda *_args, **_kwargs: basis.base_snapshot_id,
    )

    _assert_code(
        "SNAPSHOT_NOT_ADVANCED",
        lambda: build_snapshot_publication(
            basis=basis,
            acquired=acquired,
            stored=asset,
        ),
    )


def test_strict_domain_values_and_errors_do_not_leak_private_values(tmp_path: Path) -> None:
    with pytest.raises(WorkbookSnapshotPublicationError) as caught:
        _basis(workbook_instance_id="/private/client.xlsx")
    assert caught.value.code == "INVALID_BASIS"
    assert "/private" not in str(caught.value)

    with pytest.raises(WorkbookSnapshotPublicationError) as caught:
        LocalImmutableWorkbookAssetStore(Path("relative/private-assets"))
    assert caught.value.code == "ASSET_STORE_UNAVAILABLE"
    assert "relative" not in str(caught.value)

    basis = _basis(workbook_instance_id="private-secret-instance")
    acquired = _acquired(workbook_instance_id="private-secret-instance")
    assert "private-secret-instance" not in repr(basis)
    assert "private-secret-instance" not in repr(acquired)
    assert NEW_BYTES.decode() not in repr(acquired)


def test_stored_asset_contract_is_strict() -> None:
    digest = hashlib.sha256(NEW_BYTES).hexdigest()
    _assert_code(
        "INVALID_STORED_ASSET",
        lambda: StoredWorkbookAsset(
            asset_ref="workbook-asset:" + "0" * 64,
            workbook_sha256=digest,
            size_bytes=len(NEW_BYTES),
        ),
    )
    _assert_code(
        "INVALID_STORED_ASSET",
        lambda: StoredWorkbookAsset(
            asset_ref="workbook-asset:" + digest,
            workbook_sha256=digest,
            size_bytes=0,
        ),
    )


def test_saved_xlsx_must_contain_every_verified_authored_cell_and_number_format() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "C"
    worksheet["A1"] = "승인됨"
    worksheet["B2"] = "=1+1"
    worksheet["B2"].number_format = "0.00"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    acquired = AcquiredWorkbook(
        provider_revision_id="revision-b",
        predecessor_revision_id="revision-a",
        worksheet_id="worksheet-a",
        workbook_instance_id="private-instance-a",
        content=buffer.getvalue(),
    )
    manifest = {
        "office_binding": {"sheet": "C", "worksheet_id": "worksheet-a"},
        "expected_after": [
            {
                "cell": "A1",
                "authored": {"kind": "value", "value": "승인됨"},
                "number_format": "General",
            },
            {
                "cell": "B2",
                "authored": {"kind": "formula", "formula": "=1+1"},
                "number_format": "0.00",
            },
        ],
    }

    validate_saved_workbook_manifest(acquired, manifest)

    manifest["expected_after"][0]["authored"]["value"] = "다른 값"
    _assert_code(
        "SAVED_WORKBOOK_MISMATCH",
        lambda: validate_saved_workbook_manifest(acquired, manifest),
    )
