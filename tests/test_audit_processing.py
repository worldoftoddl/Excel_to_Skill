from __future__ import annotations

import json
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from excel_to_skill.audit.processing import (
    ProcessingService,
    ProcessingServiceError,
)
from excel_to_skill.audit.processing_sqlite import SQLiteProcessingRepository
from excel_to_skill.audit.processing_store import LocalPreparedBundleStore
from excel_to_skill.audit.service import (
    AuditConversationService,
    AuditConversationServiceError,
    BundleSnapshotNotFoundError,
    ConversationTurnCommand,
    InMemoryConversationArtifactRepository,
    ServicePrincipal,
)
from excel_to_skill.audit.workbook_asset_service import (
    WorkbookAssetService,
    WorkbookAssetServiceError,
)
from excel_to_skill.audit.workbook_asset_sqlite import SQLiteWorkbookAssetRepository
from excel_to_skill.audit.workbook_snapshot_publication import (
    LocalImmutableWorkbookAssetStore,
)

from test_audit_aggregate import SelectionClient
from test_audit_prepare import PipelineClient, StubRetriever, _DESCRIPTOR


PRINCIPAL = ServicePrincipal(tenant_id="tenant-a", subject_id="auditor-a")
OTHER = ServicePrincipal(tenant_id="tenant-b", subject_id="auditor-b")


def _xlsx() -> bytes:
    workbook = openpyxl.Workbook()
    alpha = workbook.active
    alpha.title = "Alpha"
    alpha.append(["계정", "위험", "절차"])
    alpha.append(["매출채권", "가공 계상", "외부조회"])
    beta = workbook.create_sheet("Beta")
    beta.append(["계정", "위험", "절차"])
    beta.append(["재고자산", "진부화", "평가 검토"])
    workbook.create_sheet("Empty")
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


class _Factories:
    def __init__(
        self,
        *,
        fail_sheet: str | None = None,
        delay_once_seconds: float = 0,
    ) -> None:
        self.prepare_clients: list[PipelineClient] = []
        self.retrievers: list[StubRetriever] = []
        self.aggregate_clients: list[SelectionClient] = []
        self.fail_sheet = fail_sheet
        self.delay_once_seconds = delay_once_seconds
        self.delay_started = threading.Event()
        self._delay_lock = threading.Lock()
        self._delay_used = False

    def client(self):
        owner = self
        fail_sheet = owner.fail_sheet

        class Client(PipelineClient):
            def __call__(self, *, system: str, user: str, schema: dict):
                should_delay = False
                with owner._delay_lock:
                    if owner.delay_once_seconds and not owner._delay_used:
                        owner._delay_used = True
                        should_delay = True
                if should_delay:
                    owner.delay_started.set()
                    time.sleep(owner.delay_once_seconds)
                properties = schema.get("properties", {})
                if (
                    fail_sheet is not None
                    and "readiness" in properties
                    and f'"observed_sheets":["{fail_sheet}"]' in user
                ):
                    raise RuntimeError("injected brief failure")
                result = super().__call__(system=system, user=user, schema=schema)
                if "region_id" in properties:
                    payload = json.loads(user)
                    result["facts"][0]["sources"][0]["ref"] = (
                        f"{payload['sheet']}!{payload['cells'][0]['cell']}"
                    )
                return result

        client = Client()
        self.prepare_clients.append(client)
        return client

    def retriever(self):
        retriever = StubRetriever()
        self.retrievers.append(retriever)
        return retriever

    def aggregate(self):
        client = SelectionClient()
        self.aggregate_clients.append(client)
        return client

    @property
    def prepare_model_calls(self) -> int:
        return sum(len(client.calls) for client in self.prepare_clients)

    @property
    def standards_calls(self) -> int:
        return sum(len(retriever.calls) for retriever in self.retrievers)

    @property
    def aggregate_calls(self) -> int:
        return sum(client.calls for client in self.aggregate_clients)


def _stack(
    tmp_path: Path,
    factories: _Factories,
    *,
    repository_kwargs: dict[str, object] | None = None,
    processing_kwargs: dict[str, object] | None = None,
):
    root = tmp_path / "server"
    raw_assets = LocalImmutableWorkbookAssetStore(root / "raw-assets")
    raw_repository = SQLiteWorkbookAssetRepository(root / "raw.sqlite3")
    workbook_assets = WorkbookAssetService(raw_repository, raw_assets)
    processing_repository = SQLiteProcessingRepository(
        root / "processing.sqlite3", **(repository_kwargs or {})
    )
    bundle_store = LocalPreparedBundleStore(root / "prepared-bundles")
    processing = ProcessingService(
        repository=processing_repository,
        workbook_assets=workbook_assets,
        bundle_store=bundle_store,
        workspace_root=root / "processing-workspace",
        model="stub-model",
        client_factory=factories.client,
        standards_retriever_factory=factories.retriever,
        retriever_descriptor=_DESCRIPTOR,
        aggregate_client_factory=factories.aggregate,
        max_prepare_workers=2,
        **(processing_kwargs or {}),
    )
    return root, workbook_assets, processing_repository, bundle_store, processing


def _upload(workbook_assets: WorkbookAssetService, *, key: str = "upload-a"):
    return workbook_assets.upload(PRINCIPAL, _xlsx(), key).snapshot


def _start(processing: ProcessingService, snapshot, *, key: str = "process-a"):
    return processing.start(
        principal=PRINCIPAL,
        workbook_id=snapshot.workbook_id,
        raw_snapshot_id=snapshot.raw_snapshot_id,
        idempotency_key=key,
    )


def _analyzable(job) -> list[dict]:
    return [item for item in job.scope_plan["sheets"] if item["analyzable"]]


def test_start_is_model_free_and_workbook_mode_publishes_commit_gated_bundle(
    tmp_path: Path,
) -> None:
    factories = _Factories()
    _, workbook_assets, _, _, processing = _stack(tmp_path, factories)
    snapshot = _upload(workbook_assets)

    started = _start(processing, snapshot)

    assert started.replayed is False
    assert started.job.status == "awaiting_scope"
    assert started.job.scope_plan_sha256
    assert [item["scope"]["sheet"] for item in started.job.scope_plan["sheets"]] == [
        "Alpha",
        "Beta",
        "Empty",
    ]
    assert factories.prepare_clients == []
    assert factories.retrievers == []
    assert factories.aggregate_clients == []

    published = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="workbook",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[],
        idempotency_key="scope-workbook",
    )

    assert published.job.status == "published"
    result = published.job.result
    assert result is not None
    assert result.sheet is None and result.aggregate_id is None
    assert result.included_sheets == ("Alpha", "Beta", "Empty")
    assert factories.prepare_model_calls > 0
    assert factories.standards_calls > 0
    assert factories.aggregate_calls == 0
    bundle = processing.resolve(principal=PRINCIPAL, bundle_id=result.bundle_id)
    assert bundle.snapshot_id == result.snapshot_id
    assert bundle.package_path.is_dir()
    assert bundle.workbook_source_provider is not None


def test_selected_single_sheet_uses_sheet_chat_binding_and_never_prepares_sibling(
    tmp_path: Path,
) -> None:
    factories = _Factories()
    _, workbook_assets, _, _, processing = _stack(tmp_path, factories)
    started = _start(processing, _upload(workbook_assets))
    alpha = _analyzable(started.job)[0]

    published = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="selected_sheets",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[alpha["scope"]["id"]],
        idempotency_key="scope-alpha",
    )

    result = published.job.result
    assert result is not None
    assert result.sheet == "Alpha"
    assert result.aggregate_id is None
    assert result.included_sheets == ("Alpha",)
    assert len(factories.prepare_clients) == 1
    assert factories.aggregate_calls == 0
    serialized_model_inputs = "\n".join(
        client.brief_users[0] for client in factories.prepare_clients
    )
    assert '"observed_sheets":["Alpha"]' in serialized_model_inputs
    assert '"observed_sheets":["Beta"]' not in serialized_model_inputs


def test_all_sheets_skips_empty_aggregates_exact_scopes_and_enters_real_chat(
    tmp_path: Path,
) -> None:
    langgraph = pytest.importorskip("langgraph")
    del langgraph
    from langgraph.checkpoint.memory import InMemorySaver
    from test_audit_conversation_aggregate import RootSelectionClient

    factories = _Factories()
    _, workbook_assets, _, _, processing = _stack(tmp_path, factories)
    started = _start(processing, _upload(workbook_assets))

    published = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="all_sheets",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[],
        idempotency_key="scope-all",
    )

    result = published.job.result
    assert result is not None and result.aggregate_id is not None
    assert result.sheet is None
    assert result.included_sheets == ("Alpha", "Beta")
    assert result.skipped_empty_sheets == ("Empty",)
    assert len(factories.prepare_clients) == 2
    assert factories.aggregate_calls == 1

    main_client = RootSelectionClient()
    conversations = AuditConversationService(
        bundles=processing,
        artifacts=InMemoryConversationArtifactRepository(),
        model="main-model",
        client=main_client,
        checkpointer=InMemorySaver(),
    )
    turn = conversations.submit_turn(
        principal=PRINCIPAL,
        command=ConversationTurnCommand(
            bundle_id=result.bundle_id,
            question="계정별 핵심 사항은?",
        ),
        idempotency_key="conversation-a",
    )

    assert turn.receipt.bundle_id == result.bundle_id
    assert turn.receipt.snapshot_id == result.snapshot_id
    evidence = turn.receipt.result["response"]["evidence"]["records"]
    assert {item["scope"]["sheet"] for item in evidence} == {"Alpha", "Beta"}
    serialized = json.dumps(turn.receipt.to_dict(), ensure_ascii=False)
    for forbidden in (
        "package_path",
        "runtime_root",
        "asset_ref",
        str(tmp_path),
    ):
        assert forbidden not in serialized

    with pytest.raises(AuditConversationServiceError) as rebound:
        conversations.submit_turn(
            principal=PRINCIPAL,
            command=ConversationTurnCommand(
                bundle_id=result.bundle_id,
                sheet="Alpha",
                question="다른 scope로 바꿔줘",
            ),
            idempotency_key="conversation-scope-pivot",
        )
    assert rebound.value.code == "BUNDLE_SCOPE_CONFLICT"
    assert rebound.value.status_code == 409


def test_one_failed_sheet_never_publishes_aggregate_or_bundle(tmp_path: Path) -> None:
    factories = _Factories(fail_sheet="Beta")
    _, workbook_assets, repository, _, processing = _stack(tmp_path, factories)
    started = _start(processing, _upload(workbook_assets))

    with pytest.raises(ProcessingServiceError) as raised:
        processing.select_scope(
            principal=PRINCIPAL,
            job_id=started.job.job_id,
            mode="all_sheets",
            scope_plan_sha256=started.job.scope_plan_sha256,
            scope_ids=[],
            idempotency_key="scope-fails",
        )

    assert raised.value.code == "PREPARATION_FAILED"
    failed = processing.get_job(principal=PRINCIPAL, job_id=started.job.job_id)
    assert failed.status == "failed"
    assert failed.failure is not None and failed.failure.stage == "preparing"
    assert failed.result is None
    assert factories.aggregate_calls == 0
    assert repository.bundle_count(principal=PRINCIPAL) == 0

    replay = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="all_sheets",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[],
        idempotency_key="scope-fails",
    )
    assert replay.replayed is True and replay.job == failed

    with pytest.raises(ProcessingServiceError) as rebound:
        processing.select_scope(
            principal=PRINCIPAL,
            job_id=started.job.job_id,
            mode="selected_sheets",
            scope_plan_sha256=started.job.scope_plan_sha256,
            scope_ids=[_analyzable(started.job)[0]["scope"]["id"]],
            idempotency_key="scope-fails-another-command",
        )
    assert rebound.value.code == "SCOPE_CONFLICT"


def test_replay_restart_principal_isolation_and_digest_pinning(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    factories = _Factories()
    root, workbook_assets, _, _, processing = _stack(tmp_path, factories)
    content = _xlsx()
    snapshot = workbook_assets.upload(PRINCIPAL, content, "upload-restart").snapshot
    started = _start(processing, snapshot, key="process-restart")
    alpha = _analyzable(started.job)[0]
    published = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="selected_sheets",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[alpha["scope"]["id"]],
        idempotency_key="scope-restart",
    )
    replay = processing.select_scope(
        principal=PRINCIPAL,
        job_id=started.job.job_id,
        mode="selected_sheets",
        scope_plan_sha256=started.job.scope_plan_sha256,
        scope_ids=[alpha["scope"]["id"]],
        idempotency_key="scope-restart",
    )
    assert replay.replayed is True and replay.job == published.job

    restarted_assets = WorkbookAssetService(
        SQLiteWorkbookAssetRepository(root / "raw.sqlite3"),
        LocalImmutableWorkbookAssetStore(root / "raw-assets"),
    )
    restarted = ProcessingService(
        repository=SQLiteProcessingRepository(root / "processing.sqlite3"),
        workbook_assets=restarted_assets,
        bundle_store=LocalPreparedBundleStore(root / "prepared-bundles"),
        workspace_root=root / "processing-workspace",
        model="stub-model",
        client_factory=factories.client,
        standards_retriever_factory=factories.retriever,
        retriever_descriptor={
            **_DESCRIPTOR,
            "retrieved_at": "2026-07-15T00:00:00Z",
        },
        aggregate_client_factory=factories.aggregate,
    )

    def expired_raw(*args, **kwargs):
        del args, kwargs
        raise WorkbookAssetServiceError("WORKBOOK_NOT_FOUND")

    monkeypatch.setattr(restarted_assets, "get_snapshot", expired_raw)
    monkeypatch.setattr(restarted_assets, "bind_source_provider", expired_raw)
    replayed_start = restarted.start(
        principal=PRINCIPAL,
        workbook_id=snapshot.workbook_id,
        raw_snapshot_id=snapshot.raw_snapshot_id,
        idempotency_key="process-restart",
    )
    assert replayed_start.replayed is True
    result = replayed_start.job.result
    assert result is not None
    resolved = restarted.resolve(principal=PRINCIPAL, bundle_id=result.bundle_id)
    assert resolved.snapshot_id == result.snapshot_id
    assert resolved.workbook_source_provider is None
    wrong_store = ProcessingService(
        repository=SQLiteProcessingRepository(root / "processing.sqlite3"),
        workbook_assets=restarted_assets,
        bundle_store=LocalPreparedBundleStore(root / "another-prepared-store"),
        workspace_root=root / "processing-workspace",
        model="stub-model",
        client_factory=factories.client,
        standards_retriever_factory=factories.retriever,
        retriever_descriptor=_DESCRIPTOR,
        aggregate_client_factory=factories.aggregate,
    )
    with pytest.raises(ProcessingServiceError) as profile_changed:
        wrong_store.start(
            principal=PRINCIPAL,
            workbook_id=snapshot.workbook_id,
            raw_snapshot_id=snapshot.raw_snapshot_id,
            idempotency_key="process-restart",
        )
    assert profile_changed.value.code == "PROCESSING_PROFILE_CHANGED"
    with pytest.raises(ProcessingServiceError) as hidden_job:
        restarted.get_job(principal=OTHER, job_id=started.job.job_id)
    assert hidden_job.value.code == "JOB_NOT_FOUND"
    with pytest.raises(BundleSnapshotNotFoundError):
        restarted.resolve(principal=OTHER, bundle_id=result.bundle_id)


def test_plan_digest_and_unknown_scope_reject_before_any_model(tmp_path: Path) -> None:
    factories = _Factories()
    _, workbook_assets, _, _, processing = _stack(tmp_path, factories)
    started = _start(processing, _upload(workbook_assets))

    with pytest.raises(ProcessingServiceError) as stale:
        processing.select_scope(
            principal=PRINCIPAL,
            job_id=started.job.job_id,
            mode="all_sheets",
            scope_plan_sha256="0" * 64,
            scope_ids=[],
            idempotency_key="scope-stale",
        )
    assert stale.value.code == "PLAN_CHANGED"

    with pytest.raises(ProcessingServiceError) as unknown:
        processing.select_scope(
            principal=PRINCIPAL,
            job_id=started.job.job_id,
            mode="selected_sheets",
            scope_plan_sha256=started.job.scope_plan_sha256,
            scope_ids=["f" * 64],
            idempotency_key="scope-unknown",
        )
    assert unknown.value.code == "INVALID_SCOPE_SELECTION"
    assert factories.prepare_clients == []
    assert factories.retrievers == []


def test_execution_heartbeat_keeps_a_slow_model_call_inside_one_fence(
    tmp_path: Path,
) -> None:
    factories = _Factories(delay_once_seconds=1.4)
    _, workbook_assets, _, _, processing = _stack(
        tmp_path,
        factories,
        repository_kwargs={"lease_ttl_seconds": 1},
        processing_kwargs={"lease_heartbeat_seconds": 0.1},
    )
    started = _start(processing, _upload(workbook_assets), key="process-heartbeat")
    command = {
        "principal": PRINCIPAL,
        "job_id": started.job.job_id,
        "mode": "all_sheets",
        "scope_plan_sha256": started.job.scope_plan_sha256,
        "scope_ids": [],
        "idempotency_key": "scope-heartbeat",
    }
    with ThreadPoolExecutor(max_workers=2) as executor:
        future = executor.submit(processing.select_scope, **command)
        assert factories.delay_started.wait(timeout=5)
        time.sleep(1.1)
        with pytest.raises(ProcessingServiceError) as pending:
            processing.select_scope(**command)
        assert pending.value.code == "JOB_IN_PROGRESS"
        published = future.result(timeout=15)
    assert published.job.status == "published"


def test_post_aggregate_gate_failure_records_terminal_aggregate_failure(
    tmp_path: Path,
) -> None:
    from excel_to_skill.verify import verify_package

    calls = 0

    def verifier(package: Path, source: Path | None):
        nonlocal calls
        calls += 1
        if calls == 3:
            raise RuntimeError("injected final gate failure")
        return verify_package(package, source)

    factories = _Factories()
    _, workbook_assets, _, _, processing = _stack(
        tmp_path,
        factories,
        processing_kwargs={"verifier": verifier},
    )
    started = _start(processing, _upload(workbook_assets), key="process-final-gate")
    with pytest.raises(ProcessingServiceError) as raised:
        processing.select_scope(
            principal=PRINCIPAL,
            job_id=started.job.job_id,
            mode="all_sheets",
            scope_plan_sha256=started.job.scope_plan_sha256,
            scope_ids=[],
            idempotency_key="scope-final-gate",
        )
    assert raised.value.code == "AGGREGATION_FAILED"
    failed = processing.get_job(principal=PRINCIPAL, job_id=started.job.job_id)
    assert failed.status == "failed"
    assert failed.failure is not None
    assert (failed.failure.code, failed.failure.stage) == (
        "AGGREGATION_FAILED",
        "aggregating",
    )
