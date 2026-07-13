from __future__ import annotations

import copy
import hashlib
import importlib.util
import json
import traceback
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import jsonschema
import openpyxl
import pytest

from excel_to_skill.audit.scope import AuditScope
from excel_to_skill.audit.workbook_inspection import (
    MAX_LEDGER_BYTES,
    WorkbookInspectionError,
    inspection_records,
    inspection_summary,
    run_workbook_inspection,
    validate_workbook_inspection_result,
    validate_workbook_inspection_summary,
)
from excel_to_skill.audit.workbook_source import BoundWorkbookSourceProvider


def _xlsx_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    rows = [
        ["id", "category", "amount", "double"],
        [1, "A", 10, "=C2*2"],
        [2, "A", 11, "=INDIRECT(\"C3\")"],
        [2, "A", 100, None],
        [3, "B", 12, None],
        [4, "B", 13, None],
        [5, "C", 14, None],
        [6, "C", 15, None],
    ]
    for row in rows:
        sheet.append(row)
    other = workbook.create_sheet("Other")
    other["A1"] = "=Data!C2"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _cell(
    sheet: str,
    cell: str,
    row: int,
    col: int,
    value: object,
    *,
    formula: str | None = None,
    cached_value: object = None,
) -> dict:
    return {
        "sheet": sheet,
        "cell": cell,
        "row": row,
        "col": col,
        "value": value,
        "formula": formula,
        "cached_value": cached_value,
        "data_type": "f" if formula else ("s" if isinstance(value, str) else "n"),
        "number_format": "General",
        "merged_range": None,
        "bold": False,
        "border": False,
        "fill": None,
    }


def _package(tmp_path: Path) -> tuple[Path, bytes]:
    source = _xlsx_bytes()
    pkg = tmp_path / "pkg"
    data = pkg / "data"
    data.mkdir(parents=True)
    meta = {
        "source": {
            "filename": "opaque.xlsx",
            "sha256": hashlib.sha256(source).hexdigest(),
            "size_bytes": len(source),
            "format": "xlsx",
        },
        "sheets": [
            {"name": "Data", "dimensions": "A1:D8", "max_row": 8, "max_col": 4},
            {"name": "Other", "dimensions": "A1", "max_row": 1, "max_col": 1},
        ],
    }
    (pkg / "meta.json").write_text(json.dumps(meta), encoding="utf-8")
    records = [
        _cell("Data", "A1", 1, 1, "id"),
        _cell("Data", "B1", 1, 2, "category"),
        _cell("Data", "C1", 1, 3, "amount"),
        _cell("Data", "D1", 1, 4, "double"),
    ]
    for row, values in enumerate(
        [
            [1, "A", 10],
            [2, "A", 11],
            [2, "A", 100],
            [3, "B", 12],
            [4, "B", 13],
            [5, "C", 14],
            [6, "C", 15],
        ],
        2,
    ):
        for col, value in enumerate(values, 1):
            records.append(_cell("Data", f"{chr(64 + col)}{row}", row, col, value))
    records.extend([
        _cell("Data", "D2", 2, 4, None, formula="C2*2", cached_value=20),
        _cell("Data", "D3", 3, 4, None, formula='INDIRECT("C3")'),
        _cell("Other", "A1", 1, 1, None, formula="Data!C2"),
    ])
    records.sort(key=lambda item: (item["sheet"], item["row"], item["col"]))
    ledger = "".join(
        json.dumps(item, ensure_ascii=False, separators=(",", ":")) + "\n"
        for item in records
    )
    (data / "cells.jsonl").write_text(ledger, encoding="utf-8")
    references = {
        "edges": [
            {
                "from": "Data!D2",
                "to": "Data!C2",
                "formula": "C2*2",
                "ref_type": "cell",
            },
            {
                "from": "Other!A1",
                "to": "Data!C2",
                "formula": "Data!C2",
                "ref_type": "cell",
            },
        ],
        "impacts": {"Data!C2": ["Data!D2", "Other!A1"]},
        "external_refs": [],
        "unresolved": [
            {"cell": "Data!D3", "formula": 'INDIRECT("C3")', "reason": "indirect"}
        ],
        "observability": {"workbook": "full", "note": None},
    }
    (data / "references.json").write_text(json.dumps(references), encoding="utf-8")
    return pkg, source


def _request(operation: str, cell_range: str, parameters: dict) -> dict:
    return {
        "operation": operation,
        "sheet": "Data",
        "range": cell_range,
        "parameters": parameters,
    }


def test_inspect_range_is_ledger_first_canonical_and_digest_bound(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    request = _request("inspect_range", "$A$2", {"source": "ledger", "limit": 20})

    first = run_workbook_inspection(pkg, request)
    second = run_workbook_inspection(pkg, request)

    assert first == second
    assert first["inspection_ref"].startswith("inspection:")
    assert first["status"] == first["evidence_status"] == "computed"
    assert first["review_status"] == "unreviewed"
    assert first["documentation_status"] == "not_documented"
    assert first["turn_scoped"] is True
    assert first["outside_prepared_bundle"] is True
    assert first["source"]["kind"] == "package_ledger"
    assert first["input"]["range"] == "A2"
    assert first["result"]["range_area"] == 1
    assert first["result"]["cells"][0]["value"] == 1
    assert inspection_records(first) == {first["inspection_ref"]: first}


def test_result_schema_and_digests_reject_tampering(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    result = run_workbook_inspection(
        pkg, _request("inspect_range", "A1:B2", {"source": "ledger", "limit": 20})
    )

    tampered = copy.deepcopy(result)
    tampered["result"]["cells"][0]["value"] = "forged"
    with pytest.raises(WorkbookInspectionError, match="digest") as error:
        validate_workbook_inspection_result(tampered)
    assert error.value.code == "CONTRACT_MISMATCH"

    unexpected = copy.deepcopy(result)
    unexpected["performed"] = True
    with pytest.raises(WorkbookInspectionError) as error:
        validate_workbook_inspection_result(unexpected)
    assert error.value.code == "CONTRACT_MISMATCH"


def test_raw_provider_is_digest_bound_and_does_not_expose_asset_or_path(tmp_path: Path) -> None:
    pkg, source = _package(tmp_path)
    calls: list[tuple[str, int]] = []

    def reader(asset_id: str, max_bytes: int) -> bytes:
        calls.append((asset_id, max_bytes))
        return source

    provider = BoundWorkbookSourceProvider("tenant/secret/source.xlsx", reader)
    result = run_workbook_inspection(
        pkg,
        _request("inspect_range", "C2:D3", {"source": "raw", "limit": 20}),
        source_provider=provider,
    )

    assert calls and calls[0][0] == "tenant/secret/source.xlsx"
    assert result["source"]["kind"] == "raw_workbook"
    serialized = json.dumps(result, ensure_ascii=False)
    assert "tenant/secret" not in serialized
    assert "source.xlsx" not in serialized

    mismatch = BoundWorkbookSourceProvider("asset", lambda _asset, _limit: b"wrong")
    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg,
            _request("inspect_range", "A1", {"source": "raw", "limit": 1}),
            source_provider=mismatch,
        )
    assert error.value.code == "SOURCE_DIGEST_MISMATCH"

    def leaking_reader(_asset: str, _limit: int) -> bytes:
        raise FileNotFoundError("/srv/private/clients/acme.xlsx")

    leaking = BoundWorkbookSourceProvider("asset", leaking_reader)
    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg,
            _request("inspect_range", "A1", {"source": "raw", "limit": 1}),
            source_provider=leaking,
        )
    assert error.value.code == "SOURCE_UNAVAILABLE"
    assert "/srv/private" not in str(error.value)
    assert "acme.xlsx" not in str(error.value)
    formatted = "".join(traceback.format_exception(error.value))
    assert "/srv/private" not in formatted
    assert "acme.xlsx" not in formatted


def test_raw_parser_preflights_zip_expansion_and_opens_value_view_only_for_formulas(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pkg, source = _package(tmp_path)
    provider = BoundWorkbookSourceProvider("asset", lambda _asset, _limit: source)
    original_load = openpyxl.load_workbook
    calls: list[bool] = []

    def counted_load(*args, **kwargs):
        calls.append(bool(kwargs.get("data_only")))
        return original_load(*args, **kwargs)

    monkeypatch.setattr(
        "excel_to_skill.audit.workbook_inspection.openpyxl.load_workbook",
        counted_load,
    )
    run_workbook_inspection(
        pkg,
        _request("inspect_range", "A1:B1", {"source": "raw", "limit": 10}),
        source_provider=provider,
    )
    assert calls == [False]

    calls.clear()
    run_workbook_inspection(
        pkg,
        _request("inspect_range", "D2", {"source": "raw", "limit": 10}),
        source_provider=provider,
    )
    assert calls == [False, True]

    bomb = BytesIO()
    with ZipFile(bomb, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr("xl/worksheets/sheet1.xml", b"A" * 1_000_000)
    bomb_bytes = bomb.getvalue()
    meta_path = pkg / "meta.json"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    meta["source"]["sha256"] = hashlib.sha256(bomb_bytes).hexdigest()
    meta["source"]["size_bytes"] = len(bomb_bytes)
    meta_path.write_text(json.dumps(meta), encoding="utf-8")
    bomb_provider = BoundWorkbookSourceProvider(
        "bomb", lambda _asset, _limit: bomb_bytes
    )

    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg,
            _request("inspect_range", "A1", {"source": "raw", "limit": 1}),
            source_provider=bomb_provider,
        )
    assert error.value.code == "SOURCE_LIMIT_EXCEEDED"


def test_inspection_and_dependencies_do_not_import_pandas(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    pkg, _ = _package(tmp_path)

    def forbidden(_name: str):
        raise AssertionError("pandas import must remain lazy")

    monkeypatch.setattr(
        "excel_to_skill.audit.workbook_inspection.importlib.import_module", forbidden
    )
    run_workbook_inspection(
        pkg, _request("inspect_range", "A1", {"source": "ledger", "limit": 1})
    )
    run_workbook_inspection(
        pkg,
        _request(
            "inspect_formula_dependencies",
            "D2",
            {"direction": "precedents", "limit": 10},
        ),
    )


def test_missing_pandas_has_fixed_dependency_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    pkg, _ = _package(tmp_path)

    def missing(name: str):
        assert name == "pandas"
        raise ImportError("machine-specific detail")

    monkeypatch.setattr(
        "excel_to_skill.audit.workbook_inspection.importlib.import_module", missing
    )
    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg, _request("profile_table", "A1:C8", {"header": True})
        )
    assert error.value.code == "DEPENDENCY_UNAVAILABLE"
    assert "machine-specific" not in str(error.value)


def test_formula_dependencies_respect_exact_direction_and_range(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)

    precedents = run_workbook_inspection(
        pkg,
        _request(
            "inspect_formula_dependencies",
            "D2",
            {"direction": "precedents", "limit": 10},
        ),
    )["result"]
    assert precedents["dependencies"] == [
        {
            "direction": "precedent",
            "from": "Data!D2",
            "to": "Data!C2",
            "formula": "C2*2",
            "ref_type": "cell",
        }
    ]
    assert precedents["unresolved"] == []
    assert precedents["total_unresolved"] == 0
    assert precedents["unresolved_truncated"] is False
    assert precedents["total_external_refs"] == 0
    assert precedents["external_refs_truncated"] is False

    dependents = run_workbook_inspection(
        pkg,
        _request(
            "inspect_formula_dependencies",
            "C2",
            {"direction": "dependents", "limit": 10},
        ),
    )["result"]
    assert [(item["from"], item["direction"]) for item in dependents["dependencies"]] == [
        ("Data!D2", "dependent"),
        ("Other!A1", "dependent"),
    ]

    unresolved = run_workbook_inspection(
        pkg,
        _request(
            "inspect_formula_dependencies",
            "D3",
            {"direction": "precedents", "limit": 10},
        ),
    )["result"]
    assert unresolved["dependencies"] == []
    assert unresolved["unresolved"][0]["reason"] == "indirect"


def test_dependency_auxiliary_coverage_reports_truncation(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    path = pkg / "data" / "references.json"
    references = json.loads(path.read_text(encoding="utf-8"))
    references["unresolved"] = [
        {
            "cell": f"Data!D{row}",
            "formula": f'INDIRECT("C{row}")',
            "reason": "indirect",
        }
        for row in range(2, 23)
    ]
    references["external_refs"] = [
        {
            "cell": f"Data!D{row}",
            "raw": f"[external.xlsx]Sheet1!A{row}",
            "target": f"[external.xlsx]Sheet1!A{row}",
        }
        for row in range(2, 23)
    ]
    path.write_text(json.dumps(references), encoding="utf-8")

    result = run_workbook_inspection(
        pkg,
        _request(
            "inspect_formula_dependencies",
            "D2:D22",
            {"direction": "precedents", "limit": 10},
        ),
    )["result"]

    assert len(result["unresolved"]) == 20
    assert result["total_unresolved"] == 21
    assert result["unresolved_truncated"] is True
    assert len(result["external_refs"]) == 20
    assert result["total_external_refs"] == 21
    assert result["external_refs_truncated"] is True


def test_exact_range_and_ledger_byte_boundaries(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    pkg, _ = _package(tmp_path)
    accepted = run_workbook_inspection(
        pkg, _request("inspect_range", "A1:CV100", {"source": "ledger", "limit": 1})
    )
    assert accepted["result"]["range_area"] == 10_000

    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg,
            _request("inspect_range", "A1:CV101", {"source": "ledger", "limit": 1}),
        )
    assert error.value.code == "LIMIT_EXCEEDED"

    for invalid_range in ("XFE1", "A1048577", "B2:A1", "A0"):
        with pytest.raises(WorkbookInspectionError) as error:
            run_workbook_inspection(
                pkg,
                _request(
                    "inspect_range",
                    invalid_range,
                    {"source": "ledger", "limit": 1},
                ),
            )
        assert error.value.code == "INVALID_REQUEST"

    ledger_size = (pkg / "data/cells.jsonl").stat().st_size
    monkeypatch.setattr(
        "excel_to_skill.audit.workbook_inspection.MAX_LEDGER_BYTES", ledger_size
    )
    run_workbook_inspection(
        pkg, _request("inspect_range", "A1", {"source": "ledger", "limit": 1})
    )
    monkeypatch.setattr(
        "excel_to_skill.audit.workbook_inspection.MAX_LEDGER_BYTES", ledger_size - 1
    )
    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg, _request("inspect_range", "A1", {"source": "ledger", "limit": 1})
        )
    assert error.value.code == "LIMIT_EXCEEDED"


def test_scope_is_exact_and_direct_function_does_not_claim_commit_gate(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    result = run_workbook_inspection(
        pkg,
        _request("inspect_range", "A1", {"source": "ledger", "limit": 1}),
        scope=AuditScope.for_sheet("Data"),
        scope_id="a" * 64,
    )
    assert result["scope"] == AuditScope.for_sheet("Data").identity()
    assert result["scope_id"] == "a" * 64
    assert "commit" not in result

    with pytest.raises(WorkbookInspectionError) as error:
        run_workbook_inspection(
            pkg,
            _request("inspect_range", "A1", {"source": "ledger", "limit": 1}),
            scope=AuditScope.for_sheet("Other"),
        )
    assert error.value.code == "INVALID_REQUEST"


def test_summary_requires_exact_current_turn_observation(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    result = run_workbook_inspection(
        pkg, _request("inspect_range", "A1", {"source": "ledger", "limit": 1})
    )
    observations = [{"tool": "workbook_inspection", "result": result}]
    assert inspection_summary(observations, selected_refs=[]) is None

    summary = inspection_summary(observations, selected_refs=[result["inspection_ref"]])
    assert summary is not None
    assert summary["inspections"] == [result]
    assert validate_workbook_inspection_summary(summary, observations=observations) == summary

    forged = copy.deepcopy(summary)
    forged["documentation_status"] = "documented"
    with pytest.raises(WorkbookInspectionError) as error:
        validate_workbook_inspection_summary(forged, observations=observations)
    assert error.value.code == "CONTRACT_MISMATCH"


@pytest.mark.skipif(
    importlib.util.find_spec("pandas") is None,
    reason="inspection extra is not installed",
)
def test_pandas_operations_are_deterministic_and_bounded(tmp_path: Path) -> None:
    pkg, _ = _package(tmp_path)
    profile = run_workbook_inspection(
        pkg, _request("profile_table", "A1:C8", {"header": True})
    )
    assert profile["result"]["columns"][2] == {
        "column": "C",
        "header": "amount",
        "inferred_type": "numeric",
        "non_null_count": 7,
        "null_count": 0,
        "distinct_count": 7,
        "numeric_count": 7,
        "minimum": 10,
        "maximum": 100,
        "mean": 25,
    }

    duplicates = run_workbook_inspection(
        pkg,
        _request(
            "find_duplicates",
            "A1:C8",
            {"header": True, "columns": ["A", "B"], "limit": 10},
        ),
    )
    assert duplicates["result"]["total_groups"] == 1
    assert duplicates["result"]["groups"][0]["rows"] == [3, 4]

    outliers = run_workbook_inspection(
        pkg,
        _request(
            "find_outliers",
            "A1:C8",
            {"header": True, "column": "C", "limit": 10},
        ),
    )
    assert outliers["result"]["total_outliers"] == 1
    assert outliers["result"]["outliers"] == [
        {"row": 4, "cell": "C4", "value": 100, "side": "upper"}
    ]


def test_schema_is_strict_draft7() -> None:
    path = Path(__file__).parents[1] / "schemas" / "audit_workbook_inspection.schema.json"
    schema = json.loads(path.read_text(encoding="utf-8"))
    jsonschema.Draft7Validator.check_schema(schema)
    assert MAX_LEDGER_BYTES == 64 * 1024 * 1024
