"""V9 픽스처 3종을 openpyxl로 생성한다(바이너리 xlsx의 의도 추적용).

바이너리 xlsx는 이 스크립트가 만든 결과를 git에 커밋한다. 재생성하려면:

    python tests/fixtures/make_fixtures.py

각 파일이 노리는 진단 포인트:

- fx1_merge_formula.xlsx — 병합 anchor가 cells에 남고 자식은 빠지는지 / 셀·범위
  수식 edge가 기대대로 나오는지.
- fx2_refs.xlsx — 시트간 참조·범위 참조 edge / INDIRECT가 unresolved에 잡히는지.
- fx3_slots_hidden.xlsx — 테두리만 있는 빈 입력 슬롯이 cells에 남고, 그 빈 칸을
  참조하는 수식이 diagnostics.blank_source_formulas에 잡히는지 / 숨김 시트·행·열이
  diagnostics.hidden에 잡히는지.
- fx4_defined_names.xlsx — 정의된 이름 이원 집계(전역/시트 스코프), #REF!·레거시
  경로 플래그, 이메일 P7 마스킹, --full-names 전량 덤프(defined_names_full.json)와
  full_dump_present 연동.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

HERE = Path(__file__).parent
_THIN = Side(style="thin")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FILL = PatternFill(fill_type="solid", fgColor="FFFF00")


def fx1_merge_formula() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.merge_cells("A1:B1")  # 병합 anchor=A1, 자식 B1
    ws["A1"] = "매출 요약"
    ws["A2"], ws["B2"] = "1분기", 100
    ws["A3"], ws["B3"] = "2분기", 200
    ws["A4"], ws["B4"] = "증가", "=B3-B2"       # cell edge ×2
    ws["A5"], ws["B5"] = "합계", "=SUM(B2:B3)"  # range edge
    wb.save(HERE / "fx1_merge_formula.xlsx")


def fx2_refs() -> None:
    wb = Workbook()
    s1 = wb.active
    s1.title = "S1"
    s1["A1"], s1["A2"], s1["A3"] = 10, 20, 30
    s2 = wb.create_sheet("S2")
    s2["B1"] = "=S1!A1"               # 시트간 cell edge
    s2["B2"] = "=SUM(S1!A1:A3)"       # 시트간 range edge
    s2["B3"] = '=INDIRECT("S1!A1")'   # unresolved: indirect
    wb.save(HERE / "fx2_refs.xlsx")


def fx3_slots_hidden() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "이름"
    ws["B1"].border = BORDER          # 값 없는 입력 슬롯(테두리)
    ws["A2"] = "메모"
    ws["B2"].fill = FILL              # 값 없는 입력 슬롯(배경)
    ws["A3"] = "확인"
    ws["B3"] = "=Main!B1"             # 빈 칸 참조 → blank_source_formulas
    ws.row_dimensions[5].hidden = True       # 숨김 행
    ws.column_dimensions["D"].hidden = True  # 숨김 열
    hidden = wb.create_sheet("숨김시트")
    hidden["A1"] = "비밀"
    hidden.sheet_state = "hidden"            # 숨김 시트
    wb.save(HERE / "fx3_slots_hidden.xlsx")


def fx4_defined_names() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["A2"], ws["A3"] = 1, 2, 3
    ws["B1"] = 9
    # 전역 정의 이름: 정상 범위·#REF! 손상·레거시 경로·이메일(P7 마스킹 대상)
    wb.defined_names.add(DefinedName("GlobalRange", attr_text="Data!$A$1:$A$3"))
    wb.defined_names.add(DefinedName("BrokenName", attr_text="#REF!$A$1"))
    wb.defined_names.add(
        DefinedName("LegacyPath", attr_text="'C:\\old\\book.xlsx'!$A$1")
    )
    wb.defined_names.add(DefinedName("Contact", attr_text='"user@example.com"'))
    # 시트 스코프 정의 이름
    ws.defined_names.add(DefinedName("LocalCell", attr_text="Data!$B$1"))
    wb.save(HERE / "fx4_defined_names.xlsx")


def main() -> None:
    fx1_merge_formula()
    fx2_refs()
    fx3_slots_hidden()
    fx4_defined_names()
    for p in sorted(HERE.glob("fx*.xlsx")):
        print("WROTE", p.name, p.stat().st_size, "bytes")


if __name__ == "__main__":
    main()
