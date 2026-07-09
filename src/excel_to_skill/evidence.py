"""§8.1 V2 — evidence 실재성 검증(형식별 주소 파서 플러그인).

semantics.json의 모든 주소 주장이 (a) 형식 유효 (b) 실존 대상 (c) 범위 내인지 본다.
범위 기준은 `meta.sheets[].dimensions`(§5 D-01 재계산 used range) — 파일 dimension
레코드가 아니라 콘텐츠 실재 기준으로 재계산된 used range다. used range 밖을 가리키는
주소는 '실존하지 않는 근거'로 간주해 실패로 본다.

검증 대상(전부 주소를 주장하는 필드):
  - workbook_claims[].evidence[]          — 시트!셀 또는 시트!범위(절대)
  - sheets[].evidence[]                   — 시트!셀 또는 시트!범위(절대)
  - sheets[].sections[].evidence[]        — 시트!셀 또는 시트!범위(절대)
  - sheets[].sections[].fields[].label_cell / value_cell
        — 해당 section이 속한 sheet 기준 상대 단일 셀(A1). null이면 검증 생략,
          문자열이면 반드시 검증(스키마가 빈 문자열은 이미 거른다).

주소 문법은 **형식별 플러그인**으로 분리한다. M3는 스프레드시트 구현만 등록하고,
docx는 `get_address_plugin`에서 NotImplementedError로 자리만 연다(M4에서 붙임).
evidence 문자열은 수식이 아니므로 전열(`A:C`)·전행(`1:5`) 참조는 형식 무효다.
"""
from __future__ import annotations

import re

from openpyxl.utils import range_boundaries

# A1 단일 셀 / A1:B2 범위(대소문자 허용, 좌표는 검증 전 대문자화). 전열·전행은 불허.
_CELL_RE = re.compile(r"^[A-Za-z]{1,3}[0-9]{1,7}$")
_RANGE_RE = re.compile(r"^[A-Za-z]{1,3}[0-9]{1,7}:[A-Za-z]{1,3}[0-9]{1,7}$")


class SpreadsheetAddress:
    """스프레드시트(xlsx·xls) 주소 파서. meta.sheets(name→dimensions)로 초기화."""

    def __init__(self, sheets: list[dict]) -> None:
        # 시트명 → (min_col, min_row, max_col, max_row). dimensions 파싱 실패 시 None.
        self._bounds: dict[str, tuple[int, int, int, int] | None] = {}
        for s in sheets:
            name = s.get("name")
            if name is None:
                continue
            try:
                self._bounds[name] = range_boundaries(str(s["dimensions"]))
            except (ValueError, KeyError, TypeError):
                self._bounds[name] = None

    def check_evidence(self, addr: str) -> str | None:
        """절대 주소 `시트!셀|범위` 검증. 통과면 None, 아니면 사유 문자열."""
        if not isinstance(addr, str) or "!" not in addr:
            return f"형식 무효(시트! 없음): {addr!r}"
        # 시트명은 '!'를 포함하지 않으므로 오른쪽에서 1회만 분리한다.
        sheet, coord = addr.rsplit("!", 1)
        if sheet not in self._bounds:
            return f"실존하지 않는 시트: {addr!r}"
        if not (_CELL_RE.match(coord) or _RANGE_RE.match(coord)):
            return f"형식 무효(셀/범위 아님): {addr!r}"
        return self._contains(sheet, coord.upper(), addr)

    def check_field_cell(self, sheet: str, coord: str, label: str) -> str | None:
        """section 소속 sheet 기준 상대 단일 셀 `A1` 검증. 통과면 None."""
        if sheet not in self._bounds:
            return f"실존하지 않는 시트: {label} (시트 {sheet!r})"
        if not isinstance(coord, str) or not _CELL_RE.match(coord):
            return f"형식 무효(단일 셀 아님): {label}={coord!r}"
        return self._contains(sheet, coord.upper(), label)

    def check_sheet_name(self, name, label: str) -> str | None:
        """sheets[].name이 meta에 실존하는 시트인지. 통과면 None."""
        if not isinstance(name, str) or name not in self._bounds:
            return f"실존하지 않는 시트: {label}={name!r}"
        return None

    def check_relative_ref(self, sheet, ref: str, label: str) -> str | None:
        """section 소속 sheet 기준 상대 셀/범위 `A1`·`A1:B2` 검증. 통과면 None."""
        if not isinstance(sheet, str) or sheet not in self._bounds:
            return f"실존하지 않는 시트: {label} (시트 {sheet!r})"
        if not isinstance(ref, str) or not (_CELL_RE.match(ref) or _RANGE_RE.match(ref)):
            return f"형식 무효(셀/범위 아님): {label}={ref!r}"
        return self._contains(sheet, ref.upper(), label)

    def _contains(self, sheet: str, coord: str, label: str) -> str | None:
        dims = self._bounds[sheet]
        if dims is None:
            return f"시트 dimensions 파싱 불가: {label}"
        dmin_c, dmin_r, dmax_c, dmax_r = dims
        tmin_c, tmin_r, tmax_c, tmax_r = range_boundaries(coord)
        if tmin_c >= dmin_c and tmin_r >= dmin_r and tmax_c <= dmax_c and tmax_r <= dmax_r:
            return None
        return f"used range({sheet} {dims_str(dims)}) 밖: {label}"


def dims_str(b: tuple[int, int, int, int]) -> str:
    """(min_col,min_row,max_col,max_row) → 'A1:B5' (사람이 읽는 사유용)."""
    from openpyxl.utils import get_column_letter

    c0, r0, c1, r1 = b
    return f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"


def get_address_plugin(fmt: str, sheets: list[dict]):
    """meta.source.format으로 주소 플러그인을 고른다. docx는 M4 자리만 연다."""
    if fmt in ("xlsx", "xls"):
        return SpreadsheetAddress(sheets)
    if fmt == "docx":
        raise NotImplementedError("docx 주소 플러그인은 M4에서 구현")
    raise ValueError(f"알 수 없는 format: {fmt!r}")


def collect_evidence_problems(semantics: dict, meta: dict) -> list[str]:
    """semantics의 모든 주소 필드를 meta 기준으로 검증해 문제 목록을 돌려준다.

    빈 리스트 = 전부 실재(V2 통과). NotImplementedError(docx 등)는 상위(verify)가
    생략 처리하도록 그대로 전파한다.
    """
    plugin = get_address_plugin(
        meta.get("source", {}).get("format", ""), meta.get("sheets", [])
    )
    problems: list[str] = []

    def _abs(addrs, where: str) -> None:
        for addr in addrs or []:
            p = plugin.check_evidence(addr)
            if p:
                problems.append(f"{where}: {p}")

    # isinstance 방어: verify_package가 스키마 통과 문서만 넘기지만, 이 함수를
    # 단독 호출해도(검증 안 된 dict) 크래시 없이 문제로 보고하도록 이중 방어한다.
    for i, wc in enumerate(_as_list(semantics.get("workbook_claims"))):
        if isinstance(wc, dict):
            _abs(wc.get("evidence"), f"workbook_claims[{i}].evidence")

    for si, sh in enumerate(_as_list(semantics.get("sheets"))):
        if not isinstance(sh, dict):
            continue
        sheet_name = sh.get("name")
        # sheets[].name은 meta에 실존하는 시트여야 한다.
        p = plugin.check_sheet_name(sheet_name, f"sheets[{si}].name")
        if p:
            problems.append(p)
        _abs(sh.get("evidence"), f"sheets[{si}].evidence")
        for ci, sec in enumerate(_as_list(sh.get("sections"))):
            if not isinstance(sec, dict):
                continue
            # sections[].range는 소속 sheet 기준 상대 셀/범위로 used range 안이어야.
            p = plugin.check_relative_ref(
                sheet_name, sec.get("range"), f"sheets[{si}].sections[{ci}].range"
            )
            if p:
                problems.append(p)
            _abs(sec.get("evidence"), f"sheets[{si}].sections[{ci}].evidence")
            for fi, fld in enumerate(_as_list(sec.get("fields"))):
                if not isinstance(fld, dict):
                    continue
                for key in ("label_cell", "value_cell"):
                    val = fld.get(key)
                    if isinstance(val, str):  # null/누락은 검증 생략, 문자열이면 필수
                        p = plugin.check_field_cell(
                            sheet_name,
                            val,
                            f"sheets[{si}].sections[{ci}].fields[{fi}].{key}",
                        )
                        if p:
                            problems.append(p)
    return problems


def _as_list(v) -> list:
    """리스트가 아니면 빈 리스트로(스키마 밖 문서의 순회 크래시 방어)."""
    return v if isinstance(v, list) else []
