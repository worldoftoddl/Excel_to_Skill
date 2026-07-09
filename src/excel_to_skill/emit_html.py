"""§4.3 layout/{시트}.html 방출기 — SheetIR → 단일 테이블 HTML (결정론 계층).

한 시트 = `<table data-sheet="{시트명}">` 하나. 규칙(§4.3, 오너 확정 기준):

- used range(A1:{max}) 전체 격자를 순회해 **빈 칸도 빈 `<td>`로** 그린다.
  병합에 먹힌 자식 칸은 `<td>`를 만들지 않는다(스팬 점유 맵). 병합 anchor는
  `colspan`/`rowspan`.
- 모든 `<td>`에 `data-cell="B4"` 도장(⑧ 값 계약: 원장 대응 칸은 cells.jsonl의
  cell 주소와 문자 단위 일치). 빈 칸도 격자 유지를 위해 도장을 찍는다.
- 수식 칸: `data-formula`에 원문, 표시 텍스트는 계산값 있으면 값, 없으면
  `[수식: =…]`.
- 스타일 최소주의: 굵게 `class="b"`, 테두리 `class="bd"`, `#RRGGBB` 배경만
  `style="background:…"`. theme/indexed 배경·그 외 서식은 버린다(근거는 cells.jsonl).
- **절단(--max-rows)**: 원장은 절대 자르지 않고 layout HTML에만 적용한다. 총 행이
  `max_rows + TAIL`을 넘을 때만 첫 max_rows행 + 말미 TAIL행을 렌더하고 중간은 생략
  `<tr>`로 표시한다(생략 행수라는 사실만 적음). 절단 사실은 diagnostics.truncations에.

판단·요약 문장은 넣지 않는다(P3). 개행 LF, 끝 개행(결정론).
"""
from __future__ import annotations

import html
from pathlib import Path

from openpyxl.utils import get_column_letter, range_boundaries

from .cache import slugify
from .emit_cells import _json_value, _merge_maps
from .extractor import CellIR, SheetIR, WorkbookIR

DEFAULT_MAX_ROWS = 5000  # §3 기본 표 행 상한
_TAIL = 5  # 절단 시 말미 보존 행수(§3)

_STYLE = (
    "table{border-collapse:collapse}"
    "td{border:1px solid #ddd;padding:2px 6px;vertical-align:top}"
    ".b{font-weight:bold}"
    ".bd{outline:1px solid #333}"
    ".truncated td{text-align:center;color:#888;font-style:italic}"
)


def assign_layout_filenames(sheets: list[SheetIR]) -> dict[str, str]:
    """시트명 → 파일명({slug}.html). slug 충돌 시 `_2`, `_3` … 접미(§4.0)."""
    used: set[str] = set()
    result: dict[str, str] = {}
    for sh in sheets:
        base = slugify(sh.name)
        name = base
        n = 2
        while name in used:
            name = f"{base}_{n}"
            n += 1
        used.add(name)
        result[sh.name] = f"{name}.html"
    return result


def _display_text(cell: CellIR | None) -> str:
    """표시 텍스트: 값 있으면 값, 없고 수식이면 계산값 또는 `[수식: =…]`."""
    if cell is None:
        return ""
    if cell.value is not None:
        return str(_json_value(cell.value))
    if cell.formula is not None:
        if cell.cached_value is not None:
            return str(_json_value(cell.cached_value))
        return f"[수식: ={cell.formula}]"
    return ""


def _cell_td(coord: str, cell: CellIR | None, *, colspan: int, rowspan: int) -> str:
    """한 칸의 `<td>` 문자열. 속성 순서 고정(결정론)."""
    attrs = [f'data-cell="{coord}"']
    if colspan > 1:
        attrs.append(f'colspan="{colspan}"')
    if rowspan > 1:
        attrs.append(f'rowspan="{rowspan}"')
    if cell is not None:
        classes = []
        if cell.bold:
            classes.append("b")
        if cell.border:
            classes.append("bd")
        if classes:
            attrs.append(f'class="{" ".join(classes)}"')
        if cell.fill and cell.fill.startswith("#"):
            attrs.append(f'style="background:{cell.fill}"')
        if cell.formula is not None:
            attrs.append(f'data-formula="{html.escape(cell.formula, quote=True)}"')
    text = html.escape(_display_text(cell))
    return f"<td {' '.join(attrs)}>{text}</td>"


def _rendered_rows(max_row: int, max_rows: int) -> tuple[list[int], int | None]:
    """렌더할 행 번호(1-based)와 (절단 시) 중간 생략 행수를 돌려준다.

    절단은 total > max_rows + TAIL일 때만(중간이 실제로 생략될 때) 발생한다.
    """
    if max_row <= max_rows + _TAIL:
        return list(range(1, max_row + 1)), None
    head = list(range(1, max_rows + 1))
    tail = list(range(max_row - _TAIL + 1, max_row + 1))
    omitted = (max_row - _TAIL) - max_rows  # head 끝 다음 ~ tail 시작 전
    return head + tail, omitted


def render_sheet_html(
    sheet: SheetIR, *, max_rows: int = DEFAULT_MAX_ROWS
) -> tuple[str, dict | None]:
    """시트 하나를 단일 테이블 HTML로 렌더. (html, truncation|None) 반환."""
    max_row = max(sheet.max_row, 1)
    max_col = max(sheet.max_col, 1)
    anchors, children = _merge_maps(sheet)
    span_bounds = {}  # anchor 좌표 → (병합 최대행, 병합 최대열)
    for (r, c), ref in anchors.items():
        min_col, min_row, mx_col, mx_row = range_boundaries(ref)
        span_bounds[(r, c)] = (mx_row, mx_col)

    rows, omitted = _rendered_rows(max_row, max_rows)
    rendered_set = set(rows)

    lines = [
        "<!DOCTYPE html>",
        '<html lang="ko">',
        '<head><meta charset="utf-8">',
        f"<style>{_STYLE}</style>",
        "</head>",
        "<body>",
        f'<table data-sheet="{html.escape(sheet.name, quote=True)}">',
    ]
    prev = None
    for r in rows:
        if prev is not None and omitted is not None and r != prev + 1:
            lines.append(
                f'<tr class="truncated"><td colspan="{max_col}">'
                f"({omitted}행 생략)</td></tr>"
            )
        prev = r
        tds = []
        for c in range(1, max_col + 1):
            if (r, c) in children:
                continue  # 병합 자식 — td 없음
            coord = f"{get_column_letter(c)}{r}"
            cell = sheet.cells.get((r, c))
            colspan = rowspan = 1
            if (r, c) in span_bounds:
                mx_row, mx_col = span_bounds[(r, c)]
                colspan = mx_col - c + 1
                # 절단으로 렌더 안 되는 행은 rowspan에서 제외(HTML 안 깨짐)
                rowspan = sum(1 for rr in range(r, mx_row + 1) if rr in rendered_set)
            tds.append(_cell_td(coord, cell, colspan=colspan, rowspan=rowspan))
        lines.append(f"<tr>{''.join(tds)}</tr>")
    lines += ["</table>", "</body>", "</html>", ""]
    html_text = "\n".join(lines)

    trunc = None
    if omitted is not None:
        trunc = {
            "sheet": sheet.name,
            "kept_head": max_rows,
            "kept_tail": _TAIL,
            "total_rows": max_row,
            "target": "layout",
        }
    return html_text, trunc


def write_layout(
    ir: WorkbookIR, layout_dir: Path, *, max_rows: int = DEFAULT_MAX_ROWS
) -> tuple[dict[str, str], list[dict]]:
    """layout/{시트}.html 전부를 쓰고 (시트명→파일명, truncations) 반환."""
    layout_dir.mkdir(parents=True, exist_ok=True)
    filenames = assign_layout_filenames(ir.sheets)
    truncations: list[dict] = []
    for sh in ir.sheets:
        html_text, trunc = render_sheet_html(sh, max_rows=max_rows)
        path = layout_dir / filenames[sh.name]
        with path.open("w", encoding="utf-8", newline="\n") as f:
            f.write(html_text)
        if trunc is not None:
            truncations.append(trunc)
    return filenames, truncations
