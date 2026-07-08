"""내부 Workbook IR — 로더 사다리로 연 통합문서를 통일된 자료구조로 옮긴다.

IR은 파이썬 자료구조로만 존재하며 파일로 내보내지 않는다(§9 확정 결정).
하류(emit_*)는 이 IR만 바라본다 — xlsx/xls/read_only 차이를 몰라도 된다.

관찰 불가(P6)의 표현: 형식 한계로 취득할 수 없는 항목은 None이 아니라
필드별 관례로 구분한다 — CellIR의 서식 플래그 None = 관찰 불가(False = 없음
관찰됨), SheetIR.hidden_rows/hidden_cols None = 관찰 불가(빈 목록 = 없음),
WorkbookIR.format_limitations에 사유를 기록한다.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import xlrd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .loader import detect_format, merges_from_xml, open_xls, open_xlsx_pair


@dataclass
class CellIR:
    row: int
    col: int
    value: object = None
    formula: str | None = None
    cached_value: object = None
    data_type: str = "n"
    number_format: str | None = None
    bold: bool | None = None      # None = 관찰 불가
    border: bool | None = None    # None = 관찰 불가
    fill: str | None = None       # "#RRGGBB" | "theme:N" | "indexed:N" | None(없음/불가)

    @property
    def coord(self) -> str:
        return f"{get_column_letter(self.col)}{self.row}"

    @property
    def has_content(self) -> bool:
        return self.value is not None or self.formula is not None

    @property
    def has_marks(self) -> bool:
        return bool(self.bold) or bool(self.border) or self.fill is not None


@dataclass
class SheetIR:
    name: str
    index: int
    cells: dict[tuple[int, int], CellIR] = field(default_factory=dict)
    merged_ranges: list[str] = field(default_factory=list)
    max_row: int = 0
    max_col: int = 0
    state: str = "visible"  # visible | hidden | veryHidden
    hidden_rows: list[int] | None = None       # None = 관찰 불가
    hidden_cols: list[str] | None = None       # None = 관찰 불가

    @property
    def dimensions(self) -> str:
        if not self.max_row or not self.max_col:
            return "A1:A1"
        return f"A1:{get_column_letter(self.max_col)}{self.max_row}"


@dataclass
class DefinedNameIR:
    name: str
    value: str | None
    scope: str | None = None  # None = 통합문서 전역, 아니면 시트명


@dataclass
class WorkbookIR:
    source_path: Path
    format: str            # xlsx | xls
    loader_path: str       # openpyxl_normal | openpyxl_read_only | openpyxl_read_only+xml_merge | xlrd
    sheets: list[SheetIR] = field(default_factory=list)
    defined_names: list[DefinedNameIR] = field(default_factory=list)
    external_links: list[str] | None = None    # None = 관찰 불가
    format_limitations: str | None = None


def extract_workbook(path: Path) -> WorkbookIR:
    """진입점. 미지원 형식이면 UnsupportedFormatError."""
    fmt = detect_format(path)
    if fmt == "xlsx":
        return _extract_xlsx(path)
    return _extract_xls(path)


# ── xlsx (openpyxl) ──────────────────────────────────────────────


def _formula_text(value) -> str:
    if isinstance(value, str):
        return value
    text = getattr(value, "text", None)  # ArrayFormula / DataTableFormula
    if text is not None:
        return text
    return str(value)


def _fill_repr(fill) -> str | None:
    try:
        if fill is None or fill.patternType in (None, "none"):
            return None
        color = fill.fgColor
        if color is None:
            return None
        if color.type == "rgb" and isinstance(color.rgb, str):
            rgb = color.rgb[-6:]
            if rgb.upper() == "000000" and fill.patternType != "solid":
                return None
            return f"#{rgb.upper()}"
        if color.type == "theme":
            return f"theme:{color.theme}"
        if color.type == "indexed":
            return f"indexed:{color.indexed}"
    except Exception:
        return None
    return None


def _border_repr(border) -> bool:
    if border is None:
        return False
    return any(
        side is not None and side.style is not None
        for side in (border.left, border.right, border.top, border.bottom)
    )


def _cell_style_flags(cell) -> tuple[bool | None, bool | None, str | None]:
    """(bold, border, fill). 스타일 접근 불가 셀은 (None, None, None)."""
    try:
        font = cell.font
    except (AttributeError, TypeError):
        return None, None, None
    bold = bool(font.bold) if font is not None else False
    return bold, _border_repr(cell.border), _fill_repr(cell.fill)


def _extract_xlsx(path: Path) -> WorkbookIR:
    wb_f, wb_v, mode = open_xlsx_pair(path)
    try:
        ir = WorkbookIR(source_path=path, format="xlsx", loader_path=f"openpyxl_{mode}")

        # 캐시값 통합문서를 먼저 훑어 (시트, 행, 열) → 마지막 저장 계산값 맵 구성
        cached: dict[tuple[str, int, int], object] = {}
        for ws in wb_v.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if getattr(cell, "value", None) is not None and cell.row is not None:
                        cached[(ws.title, cell.row, cell.column)] = cell.value

        xml_merges: dict[str, list[str]] | None = None
        for idx, ws in enumerate(wb_f.worksheets):
            sheet = SheetIR(name=ws.title, index=idx, state=ws.sheet_state)

            for row in ws.iter_rows():
                for cell in row:
                    if getattr(cell, "row", None) is None:  # read_only EmptyCell
                        continue
                    bold, border, fill = _cell_style_flags(cell)
                    raw = cell.value
                    has_style_marks = bool(bold) or bool(border) or fill is not None
                    if raw is None and not has_style_marks:
                        continue
                    c = CellIR(
                        row=cell.row,
                        col=cell.column,
                        data_type=cell.data_type,
                        number_format=cell.number_format,
                        bold=bold,
                        border=border,
                        fill=fill,
                    )
                    if cell.data_type == "f":
                        c.formula = _formula_text(raw)
                        c.cached_value = cached.get((ws.title, cell.row, cell.column))
                    else:
                        c.value = raw
                    sheet.cells[(cell.row, cell.column)] = c

            if sheet.cells:
                sheet.max_row = max(r for r, _ in sheet.cells)
                sheet.max_col = max(c for _, c in sheet.cells)

            # 병합 범위: 2차(read_only)에서 미취득 시 3차 XML 직파싱
            try:
                sheet.merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            except (AttributeError, TypeError):
                if xml_merges is None:
                    xml_merges = merges_from_xml(path)
                    ir.loader_path = "openpyxl_read_only+xml_merge"
                sheet.merged_ranges = xml_merges.get(ws.title, [])

            # 숨김 행·열: read_only 모드에서는 관찰 불가
            if isinstance(ws, Worksheet):
                sheet.hidden_rows = sorted(
                    r for r, dim in ws.row_dimensions.items() if dim.hidden
                )
                sheet.hidden_cols = sorted(
                    letter for letter, dim in ws.column_dimensions.items() if dim.hidden
                )

            ir.sheets.append(sheet)

        for name in wb_f.defined_names:
            dn = wb_f.defined_names[name]
            ir.defined_names.append(DefinedNameIR(name=dn.name, value=dn.value))
        for ws in wb_f.worksheets:
            local = getattr(ws, "defined_names", None)
            if local:
                for name in local:
                    dn = local[name]
                    ir.defined_names.append(
                        DefinedNameIR(name=dn.name, value=dn.value, scope=ws.title)
                    )

        links = getattr(wb_f, "_external_links", None) or []
        ir.external_links = [
            link.file_link.Target for link in links if link.file_link is not None
        ]

        if mode == "read_only":
            ir.format_limitations = (
                "openpyxl read_only 모드: 숨김 행·열 정보 관찰 불가"
            )
        return ir
    finally:
        wb_f.close()
        wb_v.close()


# ── xls (xlrd) ───────────────────────────────────────────────────

_XLS_TYPE_MAP = {
    xlrd.XL_CELL_TEXT: "s",
    xlrd.XL_CELL_NUMBER: "n",
    xlrd.XL_CELL_DATE: "d",
    xlrd.XL_CELL_BOOLEAN: "b",
    xlrd.XL_CELL_ERROR: "e",
}

_XLS_STATE_MAP = {0: "visible", 1: "hidden", 2: "veryHidden"}


def _xls_fill(book: xlrd.book.Book, xf) -> str | None:
    try:
        if xf.background.fill_pattern == 0:
            return None
        rgb = book.colour_map.get(xf.background.pattern_colour_index)
        if rgb is None:
            return f"indexed:{xf.background.pattern_colour_index}"
        return "#%02X%02X%02X" % rgb
    except Exception:
        return None


def _extract_xls(path: Path) -> WorkbookIR:
    book = open_xls(path)
    ir = WorkbookIR(
        source_path=path,
        format="xls",
        loader_path="xlrd",
        format_limitations=(
            "xls(xlrd): 수식 원문 접근 불가 — 참조 그래프 관찰 불가(P6). "
            "외부 링크 목록 관찰 불가."
        ),
    )

    for idx in range(book.nsheets):
        sh = book.sheet_by_index(idx)
        sheet = SheetIR(
            name=sh.name,
            index=idx,
            state=_XLS_STATE_MAP.get(sh.visibility, "visible"),
        )

        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ctype = sh.cell_type(r, c)
                try:
                    xf = book.xf_list[sh.cell_xf_index(r, c)]
                except (IndexError, xlrd.XLRDError):
                    xf = None

                bold = border = None
                fill = None
                number_format = None
                if xf is not None:
                    font = book.font_list[xf.font_index] if xf.font_index < len(book.font_list) else None
                    bold = bool(font and (font.bold or font.weight >= 700))
                    border = any(
                        getattr(xf.border, s) != 0
                        for s in ("top_line_style", "bottom_line_style",
                                  "left_line_style", "right_line_style")
                    )
                    fill = _xls_fill(book, xf)
                    fmt_obj = book.format_map.get(xf.format_key)
                    number_format = fmt_obj.format_str if fmt_obj else None

                value = None
                if ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    value = sh.cell_value(r, c)
                    if ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate_as_datetime(value, book.datemode)
                        except Exception:
                            pass
                    elif ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    elif ctype == xlrd.XL_CELL_NUMBER and isinstance(value, float) and value.is_integer():
                        value = int(value)

                has_marks = bool(bold) or bool(border) or fill is not None
                if value is None and not has_marks:
                    continue
                sheet.cells[(r + 1, c + 1)] = CellIR(
                    row=r + 1,
                    col=c + 1,
                    value=value,
                    data_type=_XLS_TYPE_MAP.get(ctype, "n"),
                    number_format=number_format,
                    bold=bold,
                    border=border,
                    fill=fill,
                )

        if sheet.cells:
            sheet.max_row = max(r for r, _ in sheet.cells)
            sheet.max_col = max(c for _, c in sheet.cells)

        sheet.merged_ranges = [
            f"{get_column_letter(clo + 1)}{rlo + 1}:{get_column_letter(chi)}{rhi}"
            for rlo, rhi, clo, chi in sh.merged_cells
        ]
        sheet.hidden_rows = sorted(
            r + 1 for r, info in sh.rowinfo_map.items() if info.hidden
        )
        sheet.hidden_cols = sorted(
            get_column_letter(c + 1)
            for c, info in sh.colinfo_map.items()
            if info.hidden
        )
        ir.sheets.append(sheet)

    for name_obj in book.name_obj_list:
        scope = None
        if name_obj.scope != -1 and 0 <= name_obj.scope < book.nsheets:
            scope = book.sheet_by_index(name_obj.scope).name
        try:
            value = name_obj.formula_text()
        except Exception:
            value = None
        ir.defined_names.append(
            DefinedNameIR(name=name_obj.name, value=value, scope=scope)
        )

    return ir
