"""Bounded, deterministic workbook inspection for ``audit-chat``.

The converted package ledger is the default and authoritative inspection source.  A caller may
explicitly request one raw range re-read, but only through an opaque provider already bound to the
package source digest.  Analytical results are computed observations, never documented workbook
facts, and remain current-turn material outside prepared audit artifacts.
"""
from __future__ import annotations

import copy
import hashlib
import importlib
import json
import math
import re
from collections.abc import Mapping
from io import BytesIO
from pathlib import Path
from pathlib import PurePosixPath
from zipfile import BadZipFile, ZIP_DEFLATED, ZIP_STORED, ZipFile

import jsonschema
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from ..emit_cells import _json_value
from ..extractor import _formula_text
from ..resources import SCHEMA_DIR
from .model import json_sha256
from .scope import AuditScope
from .workbook_source import (
    WorkbookSourceError,
    WorkbookSourceProvider,
    read_verified_workbook_source,
)


INSPECTION_VERSION = "0.1.0"
INSPECTION_RESULT_SCHEMA = "audit_workbook_inspection.v1"
INSPECTION_SUMMARY_SCHEMA = "audit_workbook_inspection_summary.v1"
INSPECTION_SCHEMA_FILE = "audit_workbook_inspection.schema.json"

MAX_RANGE_CELLS = 10_000
MAX_RANGE_ROWS = 2_000
MAX_RANGE_COLUMNS = 100
MAX_EXCEL_ROWS = 1_048_576
MAX_EXCEL_COLUMNS = 16_384
MAX_INSPECT_CELLS = 200
MAX_DEPENDENCIES = 100
MAX_ANALYTIC_RESULTS = 50
MAX_DUPLICATE_COLUMNS = 10
MAX_SELECTED_INSPECTIONS = 5
MAX_LEDGER_BYTES = 64 * 1024 * 1024
MAX_REFERENCES_BYTES = 32 * 1024 * 1024
MAX_RESULT_BYTES = 300_000
MAX_XLSX_MEMBERS = 10_000
MAX_XLSX_MEMBER_BYTES = 64 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200

_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")
_COLUMN_RE = re.compile(r"^[A-Z]{1,3}$")
_OPERATIONS = {
    "inspect_range",
    "inspect_formula_dependencies",
    "profile_table",
    "find_duplicates",
    "find_outliers",
}
_LIMITATIONS = (
    "계산 결과는 조서에 문서화된 감사증거가 아닙니다.",
    "결과는 현재 turn과 지정 범위에만 유효하며 사람이 검토해야 합니다.",
)
_LEDGER_LIMITATION = (
    "package ledger는 유의 셀만 포함하므로 완전한 빈 셀과 일부 서식은 관찰되지 않습니다."
)
_RAW_LIMITATION = (
    "raw 재조회는 package snapshot과 동일 digest의 업로드 asset에 한정됩니다."
)

_CELL_FIELDS = (
    "sheet",
    "cell",
    "row",
    "col",
    "value",
    "formula",
    "cached_value",
    "data_type",
    "number_format",
)
_UNSET = object()


class WorkbookInspectionError(RuntimeError):
    """A bounded inspection failed without creating workbook or audit authority."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        super().__init__(message)


def _fail_source(error: WorkbookSourceError) -> WorkbookInspectionError:
    return WorkbookInspectionError(error.code, str(error))


def _read_json_object(path: Path, *, label: str, max_bytes: int | None = None) -> dict:
    try:
        if max_bytes is not None and path.stat().st_size > max_bytes:
            raise WorkbookInspectionError(
                "LIMIT_EXCEEDED", f"{label} byte 상한을 초과했습니다."
            )
        value = json.loads(path.read_text(encoding="utf-8"))
    except WorkbookInspectionError:
        raise
    except (FileNotFoundError, OSError, UnicodeDecodeError, json.JSONDecodeError) as e:
        raise WorkbookInspectionError(
            "PACKAGE_INVALID", f"{label}을 검증할 수 없습니다."
        ) from e
    if not isinstance(value, dict):
        raise WorkbookInspectionError("PACKAGE_INVALID", f"{label}은 JSON 객체여야 합니다.")
    return value


def _file_sha256(path: Path, *, label: str, max_bytes: int) -> str:
    try:
        if path.stat().st_size > max_bytes:
            raise WorkbookInspectionError(
                "LIMIT_EXCEEDED", f"{label} byte 상한을 초과했습니다."
            )
        digest = hashlib.sha256()
        with path.open("rb") as file:
            for chunk in iter(lambda: file.read(1 << 20), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except WorkbookInspectionError:
        raise
    except (FileNotFoundError, OSError) as e:
        raise WorkbookInspectionError(
            "PACKAGE_INVALID", f"{label}을 검증할 수 없습니다."
        ) from e


def _package_binding(pkg: Path) -> tuple[dict, str, str, set[str]]:
    meta = _read_json_object(pkg / "meta.json", label="meta.json")
    source = meta.get("source")
    sheets = meta.get("sheets")
    if not isinstance(source, Mapping) or not isinstance(sheets, list):
        raise WorkbookInspectionError("PACKAGE_INVALID", "meta source/sheets 계약이 없습니다.")
    workbook_sha = source.get("sha256")
    if not isinstance(workbook_sha, str) or _SHA256_RE.fullmatch(workbook_sha) is None:
        raise WorkbookInspectionError("PACKAGE_INVALID", "meta source digest가 유효하지 않습니다.")
    names: list[str] = []
    for item in sheets:
        if not isinstance(item, Mapping) or not isinstance(item.get("name"), str):
            raise WorkbookInspectionError("PACKAGE_INVALID", "meta sheet 계약이 유효하지 않습니다.")
        names.append(item["name"])
    if len(names) != len(set(names)):
        raise WorkbookInspectionError("PACKAGE_INVALID", "meta sheet 이름이 중복되었습니다.")
    ledger_sha = _file_sha256(
        pkg / "data" / "cells.jsonl",
        label="data/cells.jsonl",
        max_bytes=MAX_LEDGER_BYTES,
    )
    return dict(source), workbook_sha, ledger_sha, set(names)


def _scope_identity(value: object) -> dict:
    if value is None:
        return {"kind": "workbook"}
    if isinstance(value, AuditScope):
        return value.identity()
    if not isinstance(value, Mapping):
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection scope가 유효하지 않습니다.")
    scope = dict(value)
    if scope == {"kind": "workbook"}:
        return scope
    if set(scope) == {"kind", "sheet", "id"} and scope.get("kind") == "sheet":
        sheet = scope.get("sheet")
        if (
            isinstance(sheet, str)
            and bool(sheet)
            and scope.get("id") == hashlib.sha256(sheet.encode("utf-8")).hexdigest()
        ):
            return scope
    raise WorkbookInspectionError("INVALID_REQUEST", "inspection scope가 유효하지 않습니다.")


def _scope_id(value: object) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or _SHA256_RE.fullmatch(value) is None:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection scope_id가 유효하지 않습니다.")
    return value


def _canonical_range(value: object) -> tuple[str, tuple[int, int, int, int]]:
    if not isinstance(value, str) or not value or "!" in value:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection range가 유효하지 않습니다.")
    try:
        bounds = range_boundaries(value)
    except (TypeError, ValueError) as e:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection range가 유효하지 않습니다.") from e
    if any(item is None for item in bounds):
        raise WorkbookInspectionError("INVALID_REQUEST", "전행/전열 범위는 허용하지 않습니다.")
    min_col, min_row, max_col, max_row = bounds
    if (
        min_col < 1
        or min_row < 1
        or max_col < min_col
        or max_row < min_row
        or max_col > MAX_EXCEL_COLUMNS
        or max_row > MAX_EXCEL_ROWS
    ):
        raise WorkbookInspectionError(
            "INVALID_REQUEST", "inspection range가 Excel grid 범위를 벗어났습니다."
        )
    rows = max_row - min_row + 1
    columns = max_col - min_col + 1
    area = rows * columns
    if rows > MAX_RANGE_ROWS or columns > MAX_RANGE_COLUMNS or area > MAX_RANGE_CELLS:
        raise WorkbookInspectionError("LIMIT_EXCEEDED", "inspection range 상한을 초과했습니다.")
    first = f"{get_column_letter(min_col)}{min_row}"
    last = f"{get_column_letter(max_col)}{max_row}"
    return (first if first == last else f"{first}:{last}"), bounds


def _bounded_int(value: object, *, maximum: int, field: str) -> int:
    if (
        not isinstance(value, int)
        or isinstance(value, bool)
        or not 1 <= value <= maximum
    ):
        raise WorkbookInspectionError("LIMIT_EXCEEDED", f"{field} 상한이 유효하지 않습니다.")
    return value


def _column(value: object, *, bounds: tuple[int, int, int, int], field: str) -> str:
    if not isinstance(value, str):
        raise WorkbookInspectionError("INVALID_REQUEST", f"{field}이 유효하지 않습니다.")
    column = value.strip().upper().replace("$", "")
    if _COLUMN_RE.fullmatch(column) is None:
        raise WorkbookInspectionError("INVALID_REQUEST", f"{field}이 유효하지 않습니다.")
    try:
        index = column_index_from_string(column)
    except ValueError as e:
        raise WorkbookInspectionError("INVALID_REQUEST", f"{field}이 유효하지 않습니다.") from e
    if not bounds[0] <= index <= bounds[2]:
        raise WorkbookInspectionError("INVALID_REQUEST", f"{field}이 요청 범위 밖입니다.")
    return get_column_letter(index)


def _parameters(
    operation: str,
    value: object,
    *,
    bounds: tuple[int, int, int, int],
) -> dict:
    if not isinstance(value, Mapping):
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection parameters가 객체가 아닙니다.")
    parameters = dict(value)
    if operation == "inspect_range":
        if set(parameters) != {"source", "limit"} or parameters.get("source") not in {
            "ledger", "raw"
        }:
            raise WorkbookInspectionError("INVALID_REQUEST", "inspect_range parameters가 유효하지 않습니다.")
        return {
            "source": parameters["source"],
            "limit": _bounded_int(
                parameters.get("limit"), maximum=MAX_INSPECT_CELLS, field="inspect limit"
            ),
        }
    if operation == "inspect_formula_dependencies":
        if set(parameters) != {"direction", "limit"} or parameters.get("direction") not in {
            "precedents", "dependents", "both"
        }:
            raise WorkbookInspectionError(
                "INVALID_REQUEST", "formula dependency parameters가 유효하지 않습니다."
            )
        return {
            "direction": parameters["direction"],
            "limit": _bounded_int(
                parameters.get("limit"), maximum=MAX_DEPENDENCIES, field="dependency limit"
            ),
        }
    if operation == "profile_table":
        if set(parameters) != {"header"} or not isinstance(parameters.get("header"), bool):
            raise WorkbookInspectionError("INVALID_REQUEST", "profile parameters가 유효하지 않습니다.")
        return {"header": parameters["header"]}
    if operation == "find_duplicates":
        columns = parameters.get("columns")
        if (
            set(parameters) != {"header", "columns", "limit"}
            or not isinstance(parameters.get("header"), bool)
            or not isinstance(columns, list)
            or not 1 <= len(columns) <= MAX_DUPLICATE_COLUMNS
        ):
            raise WorkbookInspectionError("INVALID_REQUEST", "duplicate parameters가 유효하지 않습니다.")
        clean_columns = [
            _column(item, bounds=bounds, field="duplicate column") for item in columns
        ]
        if len(clean_columns) != len(set(clean_columns)):
            raise WorkbookInspectionError("INVALID_REQUEST", "duplicate column이 중복되었습니다.")
        return {
            "header": parameters["header"],
            "columns": clean_columns,
            "limit": _bounded_int(
                parameters.get("limit"), maximum=MAX_ANALYTIC_RESULTS, field="duplicate limit"
            ),
        }
    if operation == "find_outliers":
        if (
            set(parameters) != {"header", "column", "limit"}
            or not isinstance(parameters.get("header"), bool)
        ):
            raise WorkbookInspectionError("INVALID_REQUEST", "outlier parameters가 유효하지 않습니다.")
        return {
            "header": parameters["header"],
            "column": _column(
                parameters.get("column"), bounds=bounds, field="outlier column"
            ),
            "limit": _bounded_int(
                parameters.get("limit"), maximum=MAX_ANALYTIC_RESULTS, field="outlier limit"
            ),
        }
    raise WorkbookInspectionError("INVALID_REQUEST", "지원하지 않는 inspection operation입니다.")


def _clean_request(value: object) -> tuple[dict, tuple[int, int, int, int]]:
    if not isinstance(value, Mapping) or set(value) != {
        "operation", "sheet", "range", "parameters"
    }:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection request 필드가 유효하지 않습니다.")
    operation = value.get("operation")
    sheet = value.get("sheet")
    if operation not in _OPERATIONS or not isinstance(sheet, str) or not sheet:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection operation/sheet가 유효하지 않습니다.")
    cell_range, bounds = _canonical_range(value.get("range"))
    return {
        "operation": operation,
        "sheet": sheet,
        "range": cell_range,
        "parameters": _parameters(operation, value.get("parameters"), bounds=bounds),
    }, bounds


def _cell_record(value: object, *, lineno: int) -> dict:
    if not isinstance(value, Mapping):
        raise WorkbookInspectionError("PACKAGE_INVALID", f"cells.jsonl {lineno}행 계약이 유효하지 않습니다.")
    row = value.get("row")
    col = value.get("col")
    cell = value.get("cell")
    sheet = value.get("sheet")
    if (
        not isinstance(sheet, str)
        or not isinstance(cell, str)
        or not isinstance(row, int)
        or isinstance(row, bool)
        or not isinstance(col, int)
        or isinstance(col, bool)
        or row < 1
        or col < 1
    ):
        raise WorkbookInspectionError("PACKAGE_INVALID", f"cells.jsonl {lineno}행 계약이 유효하지 않습니다.")
    return {field: copy.deepcopy(value.get(field)) for field in _CELL_FIELDS}


def _ledger_cells(
    pkg: Path,
    *,
    sheet: str,
    bounds: tuple[int, int, int, int],
) -> list[dict]:
    min_col, min_row, max_col, max_row = bounds
    path = pkg / "data" / "cells.jsonl"
    selected: list[dict] = []
    try:
        with path.open(encoding="utf-8") as file:
            for lineno, line in enumerate(file, 1):
                if not line.strip():
                    continue
                try:
                    raw = json.loads(line)
                except json.JSONDecodeError as e:
                    raise WorkbookInspectionError(
                        "PACKAGE_INVALID", f"cells.jsonl {lineno}행을 검증할 수 없습니다."
                    ) from e
                cell = _cell_record(raw, lineno=lineno)
                if (
                    cell["sheet"] == sheet
                    and min_row <= cell["row"] <= max_row
                    and min_col <= cell["col"] <= max_col
                ):
                    selected.append(cell)
    except WorkbookInspectionError:
        raise
    except (FileNotFoundError, OSError, UnicodeDecodeError) as e:
        raise WorkbookInspectionError("PACKAGE_INVALID", "data/cells.jsonl을 읽을 수 없습니다.") from e
    selected.sort(key=lambda item: (item["row"], item["col"], item["cell"]))
    return selected


def _raw_cells(
    data: bytes,
    *,
    sheet: str,
    bounds: tuple[int, int, int, int],
) -> list[dict]:
    _validate_xlsx_archive(data)
    formula_book = value_book = None
    try:
        formula_book = openpyxl.load_workbook(
            BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
        if sheet not in formula_book.sheetnames:
            raise WorkbookInspectionError("INVALID_REQUEST", "원본 workbook에 요청 시트가 없습니다.")
        formula_sheet = formula_book[sheet]
        min_col, min_row, max_col, max_row = bounds
        formula_rows = formula_sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        )
        selected: list[dict] = []
        formula_positions: dict[str, int] = {}
        for formula_row in formula_rows:
            for formula_cell in formula_row:
                raw = formula_cell.value
                if raw is None:
                    continue
                is_formula = getattr(formula_cell, "data_type", None) == "f"
                if is_formula:
                    formula_positions[formula_cell.coordinate] = len(selected)
                selected.append({
                    "sheet": sheet,
                    "cell": formula_cell.coordinate,
                    "row": formula_cell.row,
                    "col": formula_cell.column,
                    "value": None if is_formula else _json_value(raw),
                    "formula": _formula_text(raw) if is_formula else None,
                    "cached_value": None,
                    "data_type": getattr(formula_cell, "data_type", None),
                    "number_format": getattr(formula_cell, "number_format", None),
                })
        if formula_positions:
            value_book = openpyxl.load_workbook(
                BytesIO(data), read_only=True, data_only=True, keep_links=False
            )
            if sheet not in value_book.sheetnames:
                raise WorkbookInspectionError(
                    "SOURCE_CONTRACT_MISMATCH", "원본 xlsx의 계산값 시트가 일치하지 않습니다."
                )
            value_sheet = value_book[sheet]
            for value_row in value_sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for value_cell in value_row:
                    position = formula_positions.get(value_cell.coordinate)
                    if position is not None:
                        selected[position]["cached_value"] = _json_value(value_cell.value)
        return selected
    except WorkbookInspectionError:
        raise
    except Exception as e:  # noqa: BLE001 - openpyxl parser boundary; sanitize workbook internals
        raise WorkbookInspectionError("SOURCE_CONTRACT_MISMATCH", "원본 xlsx를 열 수 없습니다.") from e
    finally:
        for book in (formula_book, value_book):
            if book is not None:
                try:
                    book.close()
                except Exception:
                    pass


def _validate_xlsx_archive(data: bytes) -> None:
    """Reject archives whose metadata exceeds a bounded XLSX parser envelope."""
    try:
        with ZipFile(BytesIO(data)) as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_XLSX_MEMBERS:
                raise WorkbookInspectionError(
                    "SOURCE_LIMIT_EXCEEDED", "원본 xlsx archive member 상한을 초과했습니다."
                )
            names: set[str] = set()
            total_uncompressed = 0
            for member in members:
                name = member.filename
                path = PurePosixPath(name)
                if (
                    not name
                    or len(name) > 512
                    or "\x00" in name
                    or "\\" in name
                    or path.is_absolute()
                    or ".." in path.parts
                    or name in names
                    or member.flag_bits & 0x1
                    or member.compress_type not in {ZIP_STORED, ZIP_DEFLATED}
                ):
                    raise WorkbookInspectionError(
                        "SOURCE_CONTRACT_MISMATCH", "원본 xlsx archive 계약이 유효하지 않습니다."
                    )
                names.add(name)
                if member.is_dir():
                    continue
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise WorkbookInspectionError(
                        "SOURCE_LIMIT_EXCEEDED", "원본 xlsx member byte 상한을 초과했습니다."
                    )
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise WorkbookInspectionError(
                        "SOURCE_LIMIT_EXCEEDED", "원본 xlsx 압축해제 byte 상한을 초과했습니다."
                    )
                if (
                    member.file_size > 0
                    and (
                        member.compress_size <= 0
                        or member.file_size
                        > member.compress_size * MAX_XLSX_COMPRESSION_RATIO
                    )
                ):
                    raise WorkbookInspectionError(
                        "SOURCE_LIMIT_EXCEEDED", "원본 xlsx 압축률 상한을 초과했습니다."
                    )
            if not {"[Content_Types].xml", "xl/workbook.xml"}.issubset(names):
                raise WorkbookInspectionError(
                    "SOURCE_CONTRACT_MISMATCH", "원본 xlsx 필수 archive member가 없습니다."
                )
    except WorkbookInspectionError:
        raise
    except (BadZipFile, OSError, RuntimeError, ValueError) as e:
        raise WorkbookInspectionError(
            "SOURCE_CONTRACT_MISMATCH", "원본 xlsx archive를 검증할 수 없습니다."
        ) from e


def _range_area(bounds: tuple[int, int, int, int]) -> int:
    return (bounds[2] - bounds[0] + 1) * (bounds[3] - bounds[1] + 1)


def _inspect_result(cells: list[dict], *, bounds: tuple[int, int, int, int], limit: int) -> dict:
    selected = cells[:limit]
    return {
        "kind": "inspect_range",
        "range_area": _range_area(bounds),
        "returned": len(selected),
        "total_records": len(cells),
        "truncated": len(cells) > len(selected),
        "cells": copy.deepcopy(selected),
    }


def _absolute_bounds(value: object) -> tuple[str, tuple[int, int, int, int]] | None:
    if not isinstance(value, str) or "!" not in value:
        return None
    sheet, coord = value.rsplit("!", 1)
    if not sheet:
        return None
    try:
        bounds = range_boundaries(coord)
    except (TypeError, ValueError):
        return None
    if any(item is None for item in bounds):
        return None
    return sheet, bounds


def _overlaps(
    absolute: object,
    *,
    sheet: str,
    bounds: tuple[int, int, int, int],
) -> bool:
    parsed = _absolute_bounds(absolute)
    if parsed is None or parsed[0] != sheet:
        return False
    other = parsed[1]
    return not (
        other[2] < bounds[0]
        or other[0] > bounds[2]
        or other[3] < bounds[1]
        or other[1] > bounds[3]
    )


def _dependency_result(
    pkg: Path,
    *,
    sheet: str,
    bounds: tuple[int, int, int, int],
    direction: str,
    limit: int,
) -> tuple[dict, dict]:
    doc = _read_json_object(
        pkg / "data" / "references.json",
        label="data/references.json",
        max_bytes=MAX_REFERENCES_BYTES,
    )
    edges = doc.get("edges")
    unresolved = doc.get("unresolved")
    external = doc.get("external_refs")
    observability = doc.get("observability")
    if (
        not isinstance(edges, list)
        or not isinstance(unresolved, list)
        or not isinstance(external, list)
        or not isinstance(observability, Mapping)
    ):
        raise WorkbookInspectionError("PACKAGE_INVALID", "references.json 계약이 유효하지 않습니다.")
    matched: list[dict] = []
    for edge in edges:
        if not isinstance(edge, Mapping) or set(edge) != {"from", "to", "formula", "ref_type"}:
            raise WorkbookInspectionError("PACKAGE_INVALID", "reference edge 계약이 유효하지 않습니다.")
        if direction in {"precedents", "both"} and _overlaps(
            edge.get("from"), sheet=sheet, bounds=bounds
        ):
            matched.append({"direction": "precedent", **copy.deepcopy(dict(edge))})
        if direction in {"dependents", "both"} and _overlaps(
            edge.get("to"), sheet=sheet, bounds=bounds
        ):
            matched.append({"direction": "dependent", **copy.deepcopy(dict(edge))})
    matched.sort(
        key=lambda item: (
            item["direction"], str(item.get("from")), str(item.get("to")), str(item.get("ref_type"))
        )
    )

    def auxiliary(values: list, *, key: str) -> tuple[list[dict], list[dict]]:
        selected: list[dict] = []
        for item in values:
            if not isinstance(item, Mapping):
                raise WorkbookInspectionError("PACKAGE_INVALID", f"reference {key} 계약이 유효하지 않습니다.")
            if _overlaps(item.get("cell"), sheet=sheet, bounds=bounds):
                selected.append(copy.deepcopy(dict(item)))
        selected.sort(key=lambda item: (str(item.get("cell")), json_sha256(item)))
        return selected[:20], selected

    unresolved_selected, unresolved_all = (
        auxiliary(unresolved, key="unresolved")
        if direction in {"precedents", "both"}
        else ([], [])
    )
    external_selected, external_all = (
        auxiliary(external, key="external")
        if direction in {"precedents", "both"}
        else ([], [])
    )
    returned = matched[:limit]
    result = {
        "kind": "inspect_formula_dependencies",
        "direction": direction,
        "returned": len(returned),
        "total_dependencies": len(matched),
        "truncated": len(matched) > len(returned),
        "dependencies": returned,
        "unresolved": unresolved_selected,
        "total_unresolved": len(unresolved_all),
        "unresolved_truncated": len(unresolved_all) > len(unresolved_selected),
        "external_refs": external_selected,
        "total_external_refs": len(external_all),
        "external_refs_truncated": len(external_all) > len(external_selected),
        "observability": {
            "workbook": observability.get("workbook"),
            "note": observability.get("note"),
        },
    }
    source_content = {
        "dependencies": matched,
        "unresolved": unresolved_all,
        "external_refs": external_all,
        "observability": result["observability"],
    }
    return result, source_content


def _import_pandas():
    try:
        return importlib.import_module("pandas")
    except ImportError as e:
        raise WorkbookInspectionError(
            "DEPENDENCY_UNAVAILABLE", "표 분석에는 inspection extra의 pandas가 필요합니다."
        ) from e


def _effective_value(cell: Mapping[str, object]) -> object:
    if cell.get("formula") is not None:
        return cell.get("cached_value")
    return cell.get("value")


def _matrix(
    cells: list[dict],
    *,
    bounds: tuple[int, int, int, int],
) -> tuple[list[str], list[int], list[list[object]]]:
    min_col, min_row, max_col, max_row = bounds
    columns = [get_column_letter(index) for index in range(min_col, max_col + 1)]
    rows = list(range(min_row, max_row + 1))
    values = {(item["row"], item["col"]): _effective_value(item) for item in cells}
    matrix = [
        [copy.deepcopy(values.get((row, col))) for col in range(min_col, max_col + 1)]
        for row in rows
    ]
    return columns, rows, matrix


def _json_scalar(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, str):
        return value[:2_000]
    try:
        if bool(value != value):  # pandas/NumPy NaN without importing NumPy here
            return None
    except Exception:
        pass
    if hasattr(value, "item"):
        try:
            return _json_scalar(value.item())
        except Exception:
            pass
    return str(value)[:2_000]


def _header_values(columns: list[str], matrix: list[list[object]], *, header: bool) -> list[object]:
    if not header or not matrix:
        return [None for _ in columns]
    return [_json_scalar(value) for value in matrix[0]]


def _data_rows(
    rows: list[int], matrix: list[list[object]], *, header: bool
) -> tuple[list[int], list[list[object]]]:
    if header and rows:
        return rows[1:], matrix[1:]
    return rows, matrix


def _frame(columns: list[str], matrix: list[list[object]]):
    pandas = _import_pandas()
    try:
        return pandas, pandas.DataFrame(matrix, columns=columns)
    except Exception as e:  # noqa: BLE001 - optional dependency boundary
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "표 분석 frame을 만들 수 없습니다.") from e


def _number(value: object) -> float | int | None:
    if value is None:
        return None
    result = float(value)
    if not math.isfinite(result):
        return None
    if result.is_integer():
        return int(result)
    return round(result, 12)


def _profile_result(
    cells: list[dict],
    *,
    bounds: tuple[int, int, int, int],
    header: bool,
) -> dict:
    columns, rows, matrix = _matrix(cells, bounds=bounds)
    headers = _header_values(columns, matrix, header=header)
    data_numbers, data = _data_rows(rows, matrix, header=header)
    pandas, frame = _frame(columns, data)
    profiles: list[dict] = []
    for index, column in enumerate(columns):
        series = frame[column]
        non_null = int(series.notna().sum())
        numeric = pandas.to_numeric(series, errors="coerce")
        numeric_count = int(numeric.notna().sum())
        if non_null == 0:
            inferred = "empty"
        elif numeric_count == non_null:
            inferred = "numeric"
        elif numeric_count:
            inferred = "mixed"
        else:
            inferred = "text"
        profiles.append({
            "column": column,
            "header": headers[index],
            "inferred_type": inferred,
            "non_null_count": non_null,
            "null_count": len(data_numbers) - non_null,
            "distinct_count": int(series.nunique(dropna=True)),
            "numeric_count": numeric_count,
            "minimum": _number(numeric.min()) if numeric_count else None,
            "maximum": _number(numeric.max()) if numeric_count else None,
            "mean": _number(numeric.mean()) if numeric_count else None,
        })
    return {
        "kind": "profile_table",
        "header": header,
        "range_area": _range_area(bounds),
        "row_count": len(rows),
        "data_row_count": len(data_numbers),
        "column_count": len(columns),
        "columns": profiles,
    }


def _duplicates_result(
    cells: list[dict],
    *,
    bounds: tuple[int, int, int, int],
    header: bool,
    selected_columns: list[str],
    limit: int,
) -> dict:
    columns, rows, matrix = _matrix(cells, bounds=bounds)
    data_numbers, data = _data_rows(rows, matrix, header=header)
    _, frame = _frame(columns, data)
    try:
        eligible = frame[selected_columns].notna().any(axis=1)
        duplicate_mask = frame.duplicated(subset=selected_columns, keep=False) & eligible
    except Exception as e:  # noqa: BLE001 - pandas boundary
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "중복 분석을 완료할 수 없습니다.") from e
    grouped: dict[str, dict] = {}
    for position in range(len(data_numbers)):
        if not bool(duplicate_mask.iloc[position]):
            continue
        key_values = [
            {"column": column, "value": _json_scalar(frame.iloc[position][column])}
            for column in selected_columns
        ]
        key = json_sha256(key_values)
        group = grouped.setdefault(key, {"key": key_values, "rows": [], "cells": []})
        row = data_numbers[position]
        group["rows"].append(row)
        group["cells"].extend(f"{column}{row}" for column in selected_columns)
    groups = sorted(grouped.values(), key=lambda item: (item["rows"][0], json_sha256(item["key"])))
    for group in groups:
        group["count"] = len(group["rows"])
    selected = groups[:limit]
    return {
        "kind": "find_duplicates",
        "header": header,
        "columns": list(selected_columns),
        "returned": len(selected),
        "total_groups": len(groups),
        "total_duplicate_rows": sum(item["count"] for item in groups),
        "truncated": len(groups) > len(selected),
        "groups": selected,
    }


def _outliers_result(
    cells: list[dict],
    *,
    bounds: tuple[int, int, int, int],
    header: bool,
    column: str,
    limit: int,
) -> dict:
    columns, rows, matrix = _matrix(cells, bounds=bounds)
    data_numbers, data = _data_rows(rows, matrix, header=header)
    pandas, frame = _frame(columns, data)
    try:
        numeric = pandas.to_numeric(frame[column], errors="coerce")
        finite_values = [
            (position, float(value))
            for position, value in enumerate(numeric)
            if value is not None and math.isfinite(float(value))
        ]
        clean = pandas.Series([value for _, value in finite_values], dtype="float64")
        if len(clean):
            q1 = float(clean.quantile(0.25, interpolation="linear"))
            q3 = float(clean.quantile(0.75, interpolation="linear"))
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
        else:
            q1 = q3 = iqr = lower = upper = None
    except Exception as e:  # noqa: BLE001 - pandas boundary
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "이상치 분석을 완료할 수 없습니다.") from e
    outliers: list[dict] = []
    if lower is not None and upper is not None:
        for position, value in finite_values:
            if value < lower or value > upper:
                row = data_numbers[position]
                outliers.append({
                    "row": row,
                    "cell": f"{column}{row}",
                    "value": _number(value),
                    "side": "lower" if value < lower else "upper",
                })
    selected = outliers[:limit]
    return {
        "kind": "find_outliers",
        "method": "iqr_1_5",
        "header": header,
        "column": column,
        "numeric_count": len(finite_values),
        "quartile_1": _number(q1),
        "quartile_3": _number(q3),
        "iqr": _number(iqr),
        "lower_bound": _number(lower),
        "upper_bound": _number(upper),
        "returned": len(selected),
        "total_outliers": len(outliers),
        "truncated": len(outliers) > len(selected),
        "outliers": selected,
    }


def _materialize(
    *,
    request: dict,
    result: dict,
    scope: dict,
    scope_id: str | None,
    source_kind: str,
    workbook_sha: str,
    ledger_sha: str,
    source_content: object,
) -> dict:
    source = {
        "kind": source_kind,
        "workbook_sha256": workbook_sha,
        "ledger_sha256": ledger_sha,
        "content_sha256": json_sha256(source_content),
    }
    source_sha = json_sha256(source)
    input_sha = json_sha256(request)
    result_sha = json_sha256(result)
    identity = {
        "schema_version": INSPECTION_RESULT_SCHEMA,
        "operation": request["operation"],
        "scope": scope,
        "scope_id": scope_id,
        "source_sha256": source_sha,
        "input_sha256": input_sha,
        "result_sha256": result_sha,
    }
    limitations = [*_LIMITATIONS, _RAW_LIMITATION if source_kind == "raw_workbook" else _LEDGER_LIMITATION]
    document = {
        "schema_version": INSPECTION_RESULT_SCHEMA,
        "inspection_ref": "inspection:" + json_sha256(identity),
        "status": "computed",
        "evidence_status": "computed",
        "review_status": "unreviewed",
        "documentation_status": "not_documented",
        "turn_scoped": True,
        "outside_prepared_bundle": True,
        "operation": request["operation"],
        "scope": copy.deepcopy(scope),
        "scope_id": scope_id,
        "source": source,
        "source_sha256": source_sha,
        "input": copy.deepcopy(request),
        "input_sha256": input_sha,
        "result": result,
        "result_sha256": result_sha,
        "limitations": limitations,
    }
    if len(json.dumps(document, ensure_ascii=False).encode("utf-8")) > MAX_RESULT_BYTES:
        raise WorkbookInspectionError("LIMIT_EXCEEDED", "inspection 결과 byte 상한을 초과했습니다.")
    return validate_workbook_inspection_result(
        document, expected_scope=scope, expected_scope_id=scope_id
    )


def run_workbook_inspection(
    pkg: Path | str,
    request: Mapping[str, object],
    *,
    scope: AuditScope | Mapping[str, object] | None = None,
    scope_id: str | None = None,
    source_provider: WorkbookSourceProvider | None = None,
) -> dict:
    """Run one bounded package-ledger or digest-bound raw inspection.

    This low-level function validates converted-package provenance but is not itself the prepared
    audit-bundle commit gate.  ``audit-chat`` and service callers must pass their existing
    committed-bundle gate before invoking it.
    """
    path = Path(pkg)
    clean, bounds = _clean_request(request)
    selected_scope = _scope_identity(scope)
    selected_scope_id = _scope_id(scope_id)
    source_meta, workbook_sha, ledger_sha, sheet_names = _package_binding(path)
    if clean["sheet"] not in sheet_names:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection sheet가 meta에 없습니다.")
    if selected_scope["kind"] == "sheet" and selected_scope.get("sheet") != clean["sheet"]:
        raise WorkbookInspectionError("INVALID_REQUEST", "inspection sheet가 고정 scope와 다릅니다.")
    operation = clean["operation"]
    parameters = clean["parameters"]
    source_kind = "package_ledger"

    if operation == "inspect_range" and parameters["source"] == "raw":
        if source_meta.get("format") != "xlsx":
            raise WorkbookInspectionError("RAW_FORMAT_UNSUPPORTED", "raw 재조회는 xlsx만 지원합니다.")
        try:
            data = read_verified_workbook_source(
                source_provider, expected_sha256=workbook_sha
            )
        except WorkbookSourceError as e:
            raise _fail_source(e) from e
        cells = _raw_cells(data, sheet=clean["sheet"], bounds=bounds)
        source_kind = "raw_workbook"
        result = _inspect_result(cells, bounds=bounds, limit=parameters["limit"])
        source_content: object = cells
    else:
        cells = _ledger_cells(path, sheet=clean["sheet"], bounds=bounds)
        if operation == "inspect_range":
            result = _inspect_result(cells, bounds=bounds, limit=parameters["limit"])
            source_content = cells
        elif operation == "inspect_formula_dependencies":
            result, source_content = _dependency_result(
                path,
                sheet=clean["sheet"],
                bounds=bounds,
                direction=parameters["direction"],
                limit=parameters["limit"],
            )
        elif operation == "profile_table":
            result = _profile_result(cells, bounds=bounds, header=parameters["header"])
            source_content = cells
        elif operation == "find_duplicates":
            result = _duplicates_result(
                cells,
                bounds=bounds,
                header=parameters["header"],
                selected_columns=parameters["columns"],
                limit=parameters["limit"],
            )
            source_content = cells
        elif operation == "find_outliers":
            result = _outliers_result(
                cells,
                bounds=bounds,
                header=parameters["header"],
                column=parameters["column"],
                limit=parameters["limit"],
            )
            source_content = cells
        else:  # pragma: no cover - _clean_request is exhaustive
            raise WorkbookInspectionError("INVALID_REQUEST", "지원하지 않는 inspection operation입니다.")
    return _materialize(
        request=clean,
        result=result,
        scope=selected_scope,
        scope_id=selected_scope_id,
        source_kind=source_kind,
        workbook_sha=workbook_sha,
        ledger_sha=ledger_sha,
        source_content=source_content,
    )


def _schema() -> dict:
    try:
        value = json.loads((SCHEMA_DIR / INSPECTION_SCHEMA_FILE).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection schema를 읽을 수 없습니다.") from e
    if not isinstance(value, dict):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection schema가 객체가 아닙니다.")
    return value


def _validate_result_semantics(
    result: Mapping[str, object],
    *,
    request: Mapping[str, object],
    bounds: tuple[int, int, int, int],
) -> None:
    """Validate count, range, ordering, and operation-specific result invariants."""
    operation = request["operation"]
    parameters = request["parameters"]
    if result.get("kind") != operation:
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection result kind가 다릅니다.")
    if operation == "inspect_range":
        cells = result.get("cells")
        returned = result.get("returned")
        total = result.get("total_records")
        if (
            not isinstance(cells, list)
            or returned != len(cells)
            or not isinstance(total, int)
            or total < len(cells)
            or total > _range_area(bounds)
            or result.get("range_area") != _range_area(bounds)
            or result.get("truncated") is not (total > len(cells))
            or len(cells) > parameters["limit"]
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspect_range count 계약이 다릅니다.")
        positions: list[tuple[int, int, str]] = []
        for cell in cells:
            if not isinstance(cell, Mapping):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspect cell 계약이 다릅니다.")
            row, col, coordinate = cell.get("row"), cell.get("col"), cell.get("cell")
            if (
                cell.get("sheet") != request["sheet"]
                or not isinstance(row, int)
                or not isinstance(col, int)
                or not isinstance(coordinate, str)
                or coordinate != f"{get_column_letter(col)}{row}"
                or not bounds[1] <= row <= bounds[3]
                or not bounds[0] <= col <= bounds[2]
            ):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspect cell 범위가 다릅니다.")
            positions.append((row, col, coordinate))
        if positions != sorted(positions) or len(positions) != len(set(positions)):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspect cell 순서가 다릅니다.")
        return
    if operation == "inspect_formula_dependencies":
        dependencies = result.get("dependencies")
        total = result.get("total_dependencies")
        if (
            not isinstance(dependencies, list)
            or result.get("returned") != len(dependencies)
            or not isinstance(total, int)
            or total < len(dependencies)
            or result.get("truncated") is not (total > len(dependencies))
            or len(dependencies) > parameters["limit"]
            or result.get("direction") != parameters["direction"]
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "dependency count 계약이 다릅니다.")
        allowed = {
            "precedents": {"precedent"},
            "dependents": {"dependent"},
            "both": {"precedent", "dependent"},
        }[parameters["direction"]]
        order: list[tuple[str, str, str, str]] = []
        for edge in dependencies:
            if not isinstance(edge, Mapping) or edge.get("direction") not in allowed:
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "dependency direction이 다릅니다.")
            locator = edge.get("from") if edge["direction"] == "precedent" else edge.get("to")
            if not _overlaps(locator, sheet=request["sheet"], bounds=bounds):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "dependency 범위가 다릅니다.")
            order.append((
                str(edge.get("direction")), str(edge.get("from")),
                str(edge.get("to")), str(edge.get("ref_type")),
            ))
        if order != sorted(order) or len(order) != len(set(order)):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "dependency 순서가 다릅니다.")
        if parameters["direction"] == "dependents" and (
            result.get("unresolved") or result.get("external_refs")
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "dependent 보조 참조가 허용되지 않습니다.")
        for field in ("unresolved", "external_refs"):
            values = result.get(field)
            if not isinstance(values, list) or any(
                not isinstance(item, Mapping)
                or not _overlaps(item.get("cell"), sheet=request["sheet"], bounds=bounds)
                for item in values
            ):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", f"dependency {field} 범위가 다릅니다.")
        if (
            result.get("total_unresolved") < len(result["unresolved"])
            or result.get("unresolved_truncated")
            is not (result["total_unresolved"] > len(result["unresolved"]))
            or result.get("total_external_refs") < len(result["external_refs"])
            or result.get("external_refs_truncated")
            is not (result["total_external_refs"] > len(result["external_refs"]))
        ):
            raise WorkbookInspectionError(
                "CONTRACT_MISMATCH", "dependency auxiliary coverage 계약이 다릅니다."
            )
        return
    rows = bounds[3] - bounds[1] + 1
    columns = [get_column_letter(index) for index in range(bounds[0], bounds[2] + 1)]
    data_rows = rows - 1 if parameters["header"] else rows
    if operation == "profile_table":
        profiles = result.get("columns")
        if (
            result.get("header") is not parameters["header"]
            or result.get("range_area") != _range_area(bounds)
            or result.get("row_count") != rows
            or result.get("data_row_count") != data_rows
            or result.get("column_count") != len(columns)
            or not isinstance(profiles, list)
            or [item.get("column") for item in profiles if isinstance(item, Mapping)] != columns
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "profile 범위 계약이 다릅니다.")
        for profile in profiles:
            if (
                not isinstance(profile, Mapping)
                or profile.get("non_null_count", 0) + profile.get("null_count", 0) != data_rows
                or profile.get("numeric_count", 0) > profile.get("non_null_count", 0)
                or profile.get("distinct_count", 0) > profile.get("non_null_count", 0)
            ):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "profile count 계약이 다릅니다.")
        return
    if operation == "find_duplicates":
        groups = result.get("groups")
        total = result.get("total_groups")
        if (
            result.get("header") is not parameters["header"]
            or result.get("columns") != parameters["columns"]
            or not isinstance(groups, list)
            or result.get("returned") != len(groups)
            or not isinstance(total, int)
            or total < len(groups)
            or result.get("truncated") is not (total > len(groups))
            or len(groups) > parameters["limit"]
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "duplicate count 계약이 다릅니다.")
        returned_rows = 0
        first_rows: list[int] = []
        for group in groups:
            if not isinstance(group, Mapping):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "duplicate group 계약이 다릅니다.")
            group_rows = group.get("rows")
            cells = group.get("cells")
            if (
                not isinstance(group_rows, list)
                or group_rows != sorted(group_rows)
                or len(group_rows) != len(set(group_rows))
                or group.get("count") != len(group_rows)
                or not isinstance(cells, list)
                or len(cells) != len(group_rows) * len(parameters["columns"])
                or [item.get("column") for item in group.get("key", [])] != parameters["columns"]
            ):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "duplicate group 내용이 다릅니다.")
            returned_rows += len(group_rows)
            first_rows.append(group_rows[0])
        if (
            first_rows != sorted(first_rows)
            or result.get("total_duplicate_rows", 0) < returned_rows
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "duplicate group 순서가 다릅니다.")
        return
    if operation == "find_outliers":
        outliers = result.get("outliers")
        total = result.get("total_outliers")
        if (
            result.get("header") is not parameters["header"]
            or result.get("column") != parameters["column"]
            or not isinstance(outliers, list)
            or result.get("returned") != len(outliers)
            or not isinstance(total, int)
            or total < len(outliers)
            or result.get("truncated") is not (total > len(outliers))
            or len(outliers) > parameters["limit"]
        ):
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "outlier count 계약이 다릅니다.")
        previous = -1
        for item in outliers:
            if (
                not isinstance(item, Mapping)
                or item.get("cell") != f"{parameters['column']}{item.get('row')}"
                or not bounds[1] <= item.get("row", 0) <= bounds[3]
                or item.get("row", 0) < previous
            ):
                raise WorkbookInspectionError("CONTRACT_MISMATCH", "outlier 범위가 다릅니다.")
            previous = item["row"]
        return
    raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection operation 계약이 다릅니다.")


def validate_workbook_inspection_result(
    value: object,
    *,
    expected_scope: AuditScope | Mapping[str, object] | None = None,
    expected_scope_id: str | None | object = _UNSET,
) -> dict:
    """Strictly validate schema, trust constants, digests, and exact scope binding."""
    if not isinstance(value, Mapping):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection 결과가 객체가 아닙니다.")
    document = copy.deepcopy(dict(value))
    try:
        jsonschema.validate(document, _schema())
    except jsonschema.ValidationError as e:
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection 결과 schema가 유효하지 않습니다.") from e
    scope = _scope_identity(document.get("scope"))
    scope_id = _scope_id(document.get("scope_id"))
    if expected_scope is not None and scope != _scope_identity(expected_scope):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection scope가 binding과 다릅니다.")
    if expected_scope_id is not _UNSET and scope_id != _scope_id(expected_scope_id):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection scope_id가 binding과 다릅니다.")
    source = document["source"]
    request = document["input"]
    result = document["result"]
    clean_request, bounds = _clean_request(request)
    if clean_request != request:
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection input이 canonical하지 않습니다.")
    _validate_result_semantics(result, request=clean_request, bounds=bounds)
    source_sha = json_sha256(source)
    input_sha = json_sha256(request)
    result_sha = json_sha256(result)
    identity = {
        "schema_version": INSPECTION_RESULT_SCHEMA,
        "operation": document["operation"],
        "scope": scope,
        "scope_id": scope_id,
        "source_sha256": source_sha,
        "input_sha256": input_sha,
        "result_sha256": result_sha,
    }
    expected_source_kind = (
        "raw_workbook"
        if request["operation"] == "inspect_range"
        and request["parameters"]["source"] == "raw"
        else "package_ledger"
    )
    if (
        document.get("status") != "computed"
        or document.get("evidence_status") != "computed"
        or document.get("review_status") != "unreviewed"
        or document.get("documentation_status") != "not_documented"
        or document.get("turn_scoped") is not True
        or document.get("outside_prepared_bundle") is not True
        or document.get("operation") != request.get("operation")
        or result.get("kind") != document.get("operation")
        or document.get("source_sha256") != source_sha
        or document.get("input_sha256") != input_sha
        or document.get("result_sha256") != result_sha
        or document.get("inspection_ref") != "inspection:" + json_sha256(identity)
        or source.get("kind") != expected_source_kind
        or scope.get("kind") == "sheet" and scope.get("sheet") != request.get("sheet")
        or document.get("limitations") != [
            *_LIMITATIONS,
            _RAW_LIMITATION if source.get("kind") == "raw_workbook" else _LEDGER_LIMITATION,
        ]
    ):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection digest/trust 계약이 다릅니다.")
    if len(json.dumps(document, ensure_ascii=False).encode("utf-8")) > MAX_RESULT_BYTES:
        raise WorkbookInspectionError("LIMIT_EXCEEDED", "inspection 결과 byte 상한을 초과했습니다.")
    return document


def inspection_records(value: object) -> dict[str, dict]:
    document = validate_workbook_inspection_result(value)
    return {document["inspection_ref"]: document}


def inspection_summary(
    observations: list[dict],
    *,
    selected_refs: list[str],
) -> dict | None:
    """Materialize selected current-turn inspection observations without granting later authority."""
    if (
        not isinstance(selected_refs, list)
        or len(selected_refs) > MAX_SELECTED_INSPECTIONS
        or len(selected_refs) != len(set(selected_refs))
        or any(
            not isinstance(ref, str)
            or re.fullmatch(r"inspection:[0-9a-f]{64}", ref) is None
            for ref in selected_refs
        )
    ):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "selected inspection refs가 유효하지 않습니다.")
    records: dict[str, dict] = {}
    for observation in observations:
        if not isinstance(observation, Mapping) or observation.get("tool") != "workbook_inspection":
            continue
        result = observation.get("result")
        if isinstance(result, Mapping) and result.get("schema_version") == INSPECTION_RESULT_SCHEMA:
            records.update(inspection_records(result))
    if not records:
        if selected_refs:
            raise WorkbookInspectionError("CONTRACT_MISMATCH", "선택된 inspection observation이 없습니다.")
        return None
    if any(ref not in records for ref in selected_refs):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "관찰되지 않은 inspection ref가 선택되었습니다.")
    if not selected_refs:
        return None
    selected = [copy.deepcopy(records[ref]) for ref in selected_refs]
    limitations: list[str] = []
    for document in selected:
        for limitation in document["limitations"]:
            if limitation not in limitations:
                limitations.append(limitation)
    return {
        "schema_version": INSPECTION_SUMMARY_SCHEMA,
        "status": "computed",
        "evidence_status": "computed",
        "review_status": "unreviewed",
        "documentation_status": "not_documented",
        "turn_scoped": True,
        "outside_prepared_bundle": True,
        "selected_refs": list(selected_refs),
        "inspections": selected,
        "limitations": limitations,
    }


def validate_workbook_inspection_summary(
    value: object,
    *,
    observations: list[dict],
) -> dict:
    """Exact-compare a public supplement with its current-turn private observations."""
    if not isinstance(value, Mapping):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection summary가 객체가 아닙니다.")
    selected = value.get("selected_refs")
    if not isinstance(selected, list):
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection summary refs가 없습니다.")
    expected = inspection_summary(observations, selected_refs=selected)
    if expected is None or dict(value) != expected:
        raise WorkbookInspectionError("CONTRACT_MISMATCH", "inspection summary가 observation과 다릅니다.")
    return copy.deepcopy(dict(value))


__all__ = [
    "INSPECTION_RESULT_SCHEMA",
    "INSPECTION_SUMMARY_SCHEMA",
    "INSPECTION_VERSION",
    "MAX_ANALYTIC_RESULTS",
    "MAX_DEPENDENCIES",
    "MAX_EXCEL_COLUMNS",
    "MAX_EXCEL_ROWS",
    "MAX_INSPECT_CELLS",
    "MAX_RANGE_CELLS",
    "MAX_SELECTED_INSPECTIONS",
    "WorkbookInspectionError",
    "inspection_records",
    "inspection_summary",
    "run_workbook_inspection",
    "validate_workbook_inspection_result",
    "validate_workbook_inspection_summary",
]
