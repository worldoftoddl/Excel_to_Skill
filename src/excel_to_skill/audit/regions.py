"""Deterministic workbook-region packets for audit extraction.

The legacy annotator sends only the first N cells of a sheet.  Audit extraction must instead
cover every emitted cell while keeping each model request bounded.  This module partitions the
existing deterministic ``cells.jsonl`` ledger by sheet, row gaps, row span, and cell count.

No semantic decision is made here: a region is only a stable source packet.  Regions never drop
cells and never overlap, which makes coverage measurable and cache keys reproducible.  A small,
separately labelled header/legend context may be carried into a later region, but those cells
never become part of the region's source boundary.
"""
from __future__ import annotations

import json
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_MAX_ROWS = 60
DEFAULT_MAX_CELLS = 250
DEFAULT_ROW_GAP = 3
DEFAULT_MAX_CONTEXT_ROWS = 3
DEFAULT_MAX_CONTEXT_CELLS = 72

_PROMPT_CELL_FIELDS = (
    "cell",
    "row",
    "col",
    "value",
    "formula",
    "cached_value",
    "number_format",
    "merged_range",
    "bold",
    "border",
    "fill",
)

_HEADER_LABELS = {
    "account",
    "assertion",
    "description",
    "legend",
    "procedure",
    "result",
    "status",
    "계정",
    "계정과목",
    "감사절차",
    "감사 절차",
    "결과",
    "경영진 주장",
    "경영진주장",
    "범례",
    "설명",
    "상태",
    "수행결과",
    "위험",
    "절차",
    "주장",
}

_ASSERTION_LEGEND_CODES = {
    "A",
    "ACC",
    "E",
    "EX",
    "R",
    "RO",
    "R&O",
    "C",
    "COMP",
    "O",
    "OCC",
    "CL",
    "CO",
    "CUT",
    "V",
    "VAL",
    "AL",
    "U",
    "P",
    "PD",
    "P&D",
    "D",
}
_ASSERTION_LEGEND_TERMS = {
    "accuracy",
    "existence",
    "rights and obligations",
    "completeness",
    "occurrence",
    "classification",
    "cutoff",
    "cut-off",
    "valuation",
    "understandability",
    "presentation",
    "정확성",
    "실재성",
    "존재성",
    "권리와 의무",
    "완전성",
    "발생성",
    "분류",
    "기간귀속",
    "평가",
    "이해가능성",
    "표시",
    "공시",
}


def _prompt_cells(cells: tuple[dict, ...]) -> list[dict]:
    return [
        {key: cell.get(key) for key in _PROMPT_CELL_FIELDS if cell.get(key) is not None}
        for cell in cells
    ]


@dataclass(frozen=True)
class AuditRegion:
    """A bounded, deterministic slice of one worksheet's emitted cell ledger."""

    region_id: str
    sheet: str
    cell_range: str
    cells: tuple[dict, ...]
    context_cells: tuple[dict, ...] = ()
    context_eligible: bool = False
    context_floor_row: int | None = None

    def prompt_payload(self) -> dict:
        """Return the compact, JSON-safe payload exposed to an extraction model."""
        return {
            "region_id": self.region_id,
            "sheet": self.sheet,
            "range": self.cell_range,
            "cells": _prompt_cells(self.cells),
            "read_only_context": {
                "kind": "header_or_legend",
                "source_eligible": False,
                "context_evidence_eligible": True,
                "context_evidence_role": "label",
                "requires_current_region_source": True,
                "cells": _prompt_cells(self.context_cells),
            },
        }


def _load_meta(pkg: Path) -> dict:
    path = pkg / "meta.json"
    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as e:
        raise ValueError(f"meta.json 없음: {pkg}") from e
    except json.JSONDecodeError as e:
        raise ValueError(f"meta.json 파싱 실패: {e}") from e
    if not isinstance(doc, dict):
        raise ValueError("meta.json은 JSON 객체여야 합니다.")
    return doc


def _load_cells(pkg: Path) -> list[dict]:
    path = pkg / "data" / "cells.jsonl"
    if not path.is_file():
        raise ValueError(f"data/cells.jsonl 없음: {pkg}")
    out: list[dict] = []
    with path.open(encoding="utf-8") as fh:
        for lineno, line in enumerate(fh, 1):
            if not line.strip():
                continue
            try:
                cell = json.loads(line)
            except json.JSONDecodeError as e:
                raise ValueError(f"cells.jsonl {lineno}행 파싱 실패: {e}") from e
            if not isinstance(cell, dict):
                raise ValueError(f"cells.jsonl {lineno}행은 JSON 객체여야 합니다.")
            if not isinstance(cell.get("sheet"), str):
                raise ValueError(f"cells.jsonl {lineno}행 sheet가 문자열이 아닙니다.")
            if not isinstance(cell.get("row"), int) or not isinstance(cell.get("col"), int):
                raise ValueError(f"cells.jsonl {lineno}행 row/col이 정수가 아닙니다.")
            out.append(cell)
    return out


def _range_for(cells: list[dict]) -> str:
    min_row = min(c["row"] for c in cells)
    max_row = max(c["row"] for c in cells)
    min_col = min(c["col"] for c in cells)
    max_col = max(c["col"] for c in cells)
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _content_value(cell: dict):
    """Return observable content for the bounded context heuristic only."""
    for field in ("value", "cached_value", "formula"):
        value = cell.get(field)
        if value is not None and (not isinstance(value, str) or value.strip()):
            return value
    return None


def _overlaps_columns(cell: dict, min_col: int, max_col: int) -> bool:
    cell_min = cell_max = cell["col"]
    merged = cell.get("merged_range")
    if isinstance(merged, str):
        try:
            cell_min, _, cell_max, _ = range_boundaries(merged)
        except ValueError:
            # A malformed optional merge hint must not prevent deterministic partitioning.
            cell_min = cell_max = cell["col"]
    return cell_min <= max_col and cell_max >= min_col


def _looks_like_header_or_legend(cells: list[dict]) -> bool:
    """Conservatively identify a row that can clarify a later continuation region."""
    values = [_content_value(cell) for cell in cells]
    values = [value for value in values if value is not None]
    if not values or any(cell.get("formula") is not None for cell in cells):
        return False

    texts = [value.strip() for value in values if isinstance(value, str) and value.strip()]
    folded = [value.casefold() for value in texts]
    if any(
        value in _HEADER_LABELS
        or value.startswith(("※", "*", "주:", "주：", "주)", "legend"))
        for value in folded
    ):
        return True

    # Unstyled assertion legends are common in audit templates (for example, separate
    # ``A``/``Accuracy`` and ``E``/``Existence`` cells).  Recognize that narrow pattern without
    # treating arbitrary short text rows as headers.
    normalized_codes = {
        value.upper().strip(".:=-()[]{} ")
        for value in texts
    }
    joined = " ".join(folded)
    has_assertion_code = bool(normalized_codes & _ASSERTION_LEGEND_CODES)
    has_assertion_term = any(term in joined for term in _ASSERTION_LEGEND_TERMS)
    if has_assertion_code and has_assertion_term:
        return True

    rendered = [str(value).strip() for value in values]
    short_values = all(len(value) <= 48 and "\n" not in value for value in rendered)
    style_cue = any(
        cell.get("bold") is True
        or cell.get("fill") not in (None, "")
        or cell.get("merged_range") not in (None, "")
        for cell in cells
    )
    if style_cue and texts and short_values:
        return True
    return False


def _context_for_region(
    region: AuditRegion,
    rows: dict[int, list[dict]],
    *,
    max_context_rows: int,
    max_context_cells: int,
) -> tuple[dict, ...]:
    if not region.context_eligible or max_context_rows == 0 or max_context_cells == 0:
        return ()

    min_col, min_row, max_col, _ = range_boundaries(region.cell_range)
    selected_rows: list[list[dict]] = []
    selected_count = 0
    floor = region.context_floor_row if region.context_floor_row is not None else min_row
    for row_no in sorted(
        (row for row in rows if floor <= row < min_row), reverse=True
    ):
        overlapping = sorted(
            (
                cell
                for cell in rows[row_no]
                if _overlaps_columns(cell, min_col, max_col)
            ),
            key=lambda cell: (cell["col"], cell.get("cell", "")),
        )
        if not overlapping or not _looks_like_header_or_legend(overlapping):
            continue
        remaining = max_context_cells - selected_count
        if remaining <= 0:
            break
        selected_rows.append(overlapping[:remaining])
        selected_count += min(len(overlapping), remaining)
        if len(selected_rows) >= max_context_rows or selected_count >= max_context_cells:
            break

    return tuple(cell for row in reversed(selected_rows) for cell in row)


def build_regions(
    pkg: Path | str,
    *,
    max_rows: int = DEFAULT_MAX_ROWS,
    max_cells: int = DEFAULT_MAX_CELLS,
    row_gap: int = DEFAULT_ROW_GAP,
    max_context_rows: int = DEFAULT_MAX_CONTEXT_ROWS,
    max_context_cells: int = DEFAULT_MAX_CONTEXT_CELLS,
    sheet_names: Sequence[str] | None = None,
) -> list[AuditRegion]:
    """Partition every ledger cell into stable, non-overlapping worksheet regions.

    A new region starts when the next populated/significantly-formatted row is separated by more
    than ``row_gap``, the inclusive row span would exceed ``max_rows``, or the cell-count ceiling
    would be crossed.  A single very wide row is split into column chunks so no cell is lost.
    """
    if (
        max_rows < 1
        or max_cells < 1
        or row_gap < 0
        or max_context_rows < 0
        or max_context_cells < 0
    ):
        raise ValueError(
            "max_rows/max_cells는 1 이상, row_gap/max_context_rows/"
            "max_context_cells는 0 이상이어야 합니다."
        )

    pkg = Path(pkg)
    meta = _load_meta(pkg)
    cells = _load_cells(pkg)
    sheet_order = [s.get("name") for s in meta.get("sheets", []) if isinstance(s, dict)]
    if sheet_names is not None:
        if isinstance(sheet_names, (str, bytes)):
            raise ValueError("sheet_names는 문자열 하나가 아니라 시트명 목록이어야 합니다.")
        selected = tuple(sheet_names)
        if not selected or any(not isinstance(name, str) or not name for name in selected):
            raise ValueError("sheet_names는 비어 있지 않은 시트명 목록이어야 합니다.")
        if len(set(selected)) != len(selected):
            raise ValueError("sheet_names에 중복 시트명이 있습니다.")
        known = {name for name in sheet_order if isinstance(name, str)}
        unknown = sorted(set(selected) - known)
        if unknown:
            raise ValueError(f"meta.json에 없는 시트입니다: {unknown}")
        selected_set = set(selected)
        sheet_order = [name for name in sheet_order if name in selected_set]
        cells = [cell for cell in cells if cell["sheet"] in selected_set]
    grouped: dict[str, list[dict]] = {name: [] for name in sheet_order if isinstance(name, str)}
    for cell in cells:
        grouped.setdefault(cell["sheet"], []).append(cell)

    regions: list[AuditRegion] = []
    serial = 0

    def flush(
        sheet: str,
        batch: list[dict],
        *,
        context_eligible: bool,
        context_floor_row: int | None,
    ) -> None:
        nonlocal serial
        if not batch:
            return
        serial += 1
        ordered = sorted(batch, key=lambda c: (c["row"], c["col"], c.get("cell", "")))
        regions.append(
            AuditRegion(
                region_id=f"region:{serial:05d}",
                sheet=sheet,
                cell_range=_range_for(ordered),
                cells=tuple(ordered),
                context_eligible=context_eligible,
                context_floor_row=context_floor_row,
            )
        )

    ordered_sheets = sheet_order + sorted(set(grouped) - set(sheet_order))
    rows_by_sheet: dict[str, dict[int, list[dict]]] = {}
    for sheet in ordered_sheets:
        sheet_cells = sorted(
            grouped.get(sheet, []), key=lambda c: (c["row"], c["col"], c.get("cell", ""))
        )
        rows: dict[int, list[dict]] = {}
        for cell in sheet_cells:
            rows.setdefault(cell["row"], []).append(cell)
        rows_by_sheet[sheet] = rows

        batch: list[dict] = []
        batch_context_eligible = False
        block_start_row: int | None = None
        start_row: int | None = None
        previous_row: int | None = None
        for row_no in sorted(rows):
            row_cells = rows[row_no]
            if len(row_cells) > max_cells:
                flush(
                    sheet,
                    batch,
                    context_eligible=batch_context_eligible,
                    context_floor_row=block_start_row,
                )
                batch = []
                batch_context_eligible = False
                block_start_row = None
                start_row = previous_row = None
                for offset in range(0, len(row_cells), max_cells):
                    flush(
                        sheet,
                        row_cells[offset : offset + max_cells],
                        context_eligible=False,
                        context_floor_row=row_no,
                    )
                continue

            gap_split = bool(batch) and row_no - (previous_row or row_no) > row_gap
            would_split = bool(batch) and (
                gap_split
                or row_no - (start_row or row_no) + 1 > max_rows
                or len(batch) + len(row_cells) > max_cells
            )
            if would_split:
                flush(
                    sheet,
                    batch,
                    context_eligible=batch_context_eligible,
                    context_floor_row=block_start_row,
                )
                batch = []
                # Only a size/span split is known to continue the same table.  A row gap starts
                # an independent block and must not inherit an earlier block's interpretation.
                batch_context_eligible = not gap_split
                if gap_split:
                    block_start_row = row_no
                start_row = None
            if not batch:
                start_row = row_no
                if block_start_row is None:
                    block_start_row = row_no
            batch.extend(row_cells)
            previous_row = row_no
        flush(
            sheet,
            batch,
            context_eligible=batch_context_eligible,
            context_floor_row=block_start_row,
        )

    regions = [
        AuditRegion(
            region_id=region.region_id,
            sheet=region.sheet,
            cell_range=region.cell_range,
            cells=region.cells,
            context_cells=_context_for_region(
                region,
                rows_by_sheet[region.sheet],
                max_context_rows=max_context_rows,
                max_context_cells=max_context_cells,
            ),
            context_eligible=region.context_eligible,
            context_floor_row=region.context_floor_row,
        )
        for region in regions
    ]

    if sum(len(region.cells) for region in regions) != len(cells):
        raise AssertionError("region partition이 원장 셀 수를 보존하지 못했습니다.")
    return regions
