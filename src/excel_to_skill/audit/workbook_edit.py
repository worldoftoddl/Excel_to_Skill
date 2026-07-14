"""Deterministic contracts for approved, bounded Office.js workbook edits.

This module never opens or mutates a workbook.  It turns a small, one-sheet proposal into an
exact diff over an Office.js readback, binds a human approval to that exact preview, materializes
the only apply manifest an executor may consume, and verifies the executor's reread witness.
Every document is strict-schema validated and content-addressed.
"""
from __future__ import annotations

import copy
import hashlib
import json
import math
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache

import jsonschema
from openpyxl.formula import Tokenizer
from openpyxl.formula.tokenizer import TokenizerError
from openpyxl.utils import column_index_from_string, get_column_letter

from ..resources import SCHEMA_DIR
from .model import AuditModelError, canonical_json, json_sha256


PROPOSAL_SCHEMA = "audit_workbook_edit_proposal.v1"
PREVIEW_SCHEMA = "audit_workbook_edit_preview.v1"
APPROVAL_SCHEMA = "audit_workbook_edit_approval.v1"
APPLY_MANIFEST_SCHEMA = "audit_workbook_edit_apply_manifest.v1"
WITNESS_SCHEMA = "audit_workbook_edit_witness.v1"
VERIFICATION_SCHEMA = "audit_workbook_edit_verification.v1"

MAX_CHANGES = 100
MAX_CELL_TEXT = 32_767
MAX_FORMULA_LENGTH = 8_192
MAX_NUMBER_FORMAT_LENGTH = 500
MAX_EXCEL_ROWS = 1_048_576
MAX_EXCEL_COLUMNS = 16_384
MAX_ARTIFACT_PAYLOAD_BYTES = 600_000
MAX_FORMULA_REFERENCE_CELLS = 100_000
MAX_PROPOSAL_REFERENCE_CELLS = 250_000
MAX_SAFE_INTEGER = 9_007_199_254_740_991

_SHA256_RE = re.compile(r"^[0-9a-f]{64}$")
_OPAQUE_ID_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.:-]{0,127}$")
_CELL_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")
_STATIC_RANGE_RE = re.compile(
    r"^\$?[A-Za-z]{1,3}\$?[1-9][0-9]*(?::\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)?$"
)
_FORMULA_KINDS = frozenset(
    {"set_value", "set_formula", "set_number_format", "clear_contents"}
)
_ALLOWED_FUNCTIONS = frozenset(
    {
        "ABS",
        "AND",
        "AVERAGE",
        "COUNT",
        "COUNTA",
        "COUNTBLANK",
        "COUNTIF",
        "COUNTIFS",
        "EXACT",
        "IF",
        "IFERROR",
        "INDEX",
        "ISBLANK",
        "ISERROR",
        "ISNA",
        "ISNUMBER",
        "ISTEXT",
        "LEFT",
        "LEN",
        "LOWER",
        "MATCH",
        "MAX",
        "MID",
        "MIN",
        "NOT",
        "OR",
        "RIGHT",
        "ROUND",
        "ROUNDDOWN",
        "ROUNDUP",
        "SUBTOTAL",
        "SUM",
        "SUMIF",
        "SUMIFS",
        "TEXT",
        "TRIM",
        "UPPER",
        "VALUE",
        "VLOOKUP",
        "XLOOKUP",
    }
)
_DISALLOWED_FUNCTIONS = frozenset({"INDIRECT", "OFFSET"})
_RANGE_REDUCERS = frozenset(
    {
        "AND",
        "AVERAGE",
        "COUNT",
        "COUNTA",
        "COUNTBLANK",
        "MAX",
        "MIN",
        "OR",
        "SUBTOTAL",
        "SUM",
    }
)
_ALLOWED_INFIX = frozenset({"+", "-", "*", "/", "^", "&", "=", "<>", "<=", ">=", "<", ">"})
_SCHEMA_FILES = {
    PROPOSAL_SCHEMA: "audit_workbook_edit_proposal.schema.json",
    PREVIEW_SCHEMA: "audit_workbook_edit_preview.schema.json",
    APPROVAL_SCHEMA: "audit_workbook_edit_approval.schema.json",
    APPLY_MANIFEST_SCHEMA: "audit_workbook_edit_apply_manifest.schema.json",
    WITNESS_SCHEMA: "audit_workbook_edit_witness.schema.json",
    VERIFICATION_SCHEMA: "audit_workbook_edit_verification.schema.json",
}
_EDIT_LIMITATIONS = [
    "proposal-local formula cycle만 검사하며 대상 밖 dependent cell 변화는 검증하지 않습니다.",
    "executor가 보고한 Office readback이며 backend가 workbook 계산을 독립 재수행하지 않습니다.",
    "session 검증은 asset 저장 또는 새 prepared bundle 생성을 의미하지 않습니다.",
]


class WorkbookEditError(RuntimeError):
    """A workbook edit artifact failed a stable contract boundary."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        super().__init__(message)


def _fail(code: str, message: str) -> WorkbookEditError:
    return WorkbookEditError(code, message)


def _text(value: object, *, field: str, maximum: int) -> str:
    if (
        not isinstance(value, str)
        or not value
        or value != value.strip()
        or len(value) > maximum
        or any(ord(character) < 32 or ord(character) == 127 for character in value)
    ):
        raise _fail("INVALID_INPUT", f"{field} 계약이 유효하지 않습니다.")
    return value


def _opaque(value: object, *, field: str) -> str:
    text = _text(value, field=field, maximum=128)
    if _OPAQUE_ID_RE.fullmatch(text) is None:
        raise _fail("INVALID_INPUT", f"{field}는 opaque identifier여야 합니다.")
    return text


def _sha256(value: object, *, field: str) -> str:
    if not isinstance(value, str) or _SHA256_RE.fullmatch(value) is None:
        raise _fail("INVALID_INPUT", f"{field}는 SHA-256이어야 합니다.")
    return value


def _sheet(value: object) -> str:
    text = _text(value, field="sheet", maximum=31)
    if any(character in text for character in "[]:*?/\\"):
        raise _fail("INVALID_INPUT", "sheet 이름이 유효하지 않습니다.")
    return text


def _cell(value: object) -> tuple[str, tuple[int, int]]:
    if not isinstance(value, str):
        raise _fail("INVALID_INPUT", "cell 주소가 유효하지 않습니다.")
    match = _CELL_RE.fullmatch(value)
    if match is None:
        raise _fail("INVALID_INPUT", "single-cell A1 주소만 허용됩니다.")
    column = column_index_from_string(match.group(1).upper())
    row = int(match.group(2))
    if column > MAX_EXCEL_COLUMNS or row > MAX_EXCEL_ROWS:
        raise _fail("INVALID_INPUT", "cell 주소가 Excel grid를 벗어났습니다.")
    return f"{get_column_letter(column)}{row}", (row, column)


def _scalar(value: object, *, field: str, allow_null: bool) -> object:
    if value is None:
        if allow_null:
            return None
        raise _fail("INVALID_INPUT", f"{field}에는 null을 쓸 수 없습니다.")
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        if not -MAX_SAFE_INTEGER <= value <= MAX_SAFE_INTEGER:
            raise _fail("INVALID_INPUT", f"{field} 정수가 JavaScript safe 범위를 벗어났습니다.")
        return value
    if isinstance(value, float):
        if not math.isfinite(value) or abs(value) > MAX_SAFE_INTEGER:
            raise _fail(
                "INVALID_INPUT",
                f"{field} 숫자가 finite JavaScript safe 범위를 벗어났습니다.",
            )
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, str):
        if len(value) > MAX_CELL_TEXT:
            raise _fail("INVALID_INPUT", f"{field} 문자열이 Excel 상한을 초과했습니다.")
        return value
    raise _fail("INVALID_INPUT", f"{field}은 JSON scalar여야 합니다.")


def _bounded_payload(value: object, *, field: str) -> None:
    try:
        size = len(canonical_json(value).encode("utf-8"))
    except AuditModelError as exc:
        raise _fail("INVALID_INPUT", f"{field}은 canonical JSON이어야 합니다.") from exc
    if size > MAX_ARTIFACT_PAYLOAD_BYTES:
        raise _fail("LIMIT_EXCEEDED", f"{field}이 600KB 상한을 초과했습니다.")


def _json_equal(left: object, right: object) -> bool:
    """Compare JSON values without Python's bool/int equality coercion."""

    try:
        return canonical_json(left) == canonical_json(right)
    except AuditModelError as exc:
        raise _fail("CONTRACT_MISMATCH", "JSON 값을 정확히 비교할 수 없습니다.") from exc


def _literal_value(value: object) -> object:
    clean = _scalar(value, field="set_value.value", allow_null=False)
    if isinstance(clean, str):
        if clean == "":
            raise _fail(
                "INVALID_INPUT",
                "빈 문자열 literal은 Office.js에서 blank cell과 구분해 재현할 수 없습니다.",
            )
        candidate = clean.lstrip(" \t\r\n")
        if candidate.startswith(("=", "+", "-", "@")):
            raise _fail(
                "FORMULA_INJECTION_BLOCKED",
                "literal value가 formula로 해석될 수 있어 거부되었습니다.",
            )
    return clean


def _number_format(value: object) -> str:
    if not isinstance(value, str) or not value or len(value) > MAX_NUMBER_FORMAT_LENGTH:
        raise _fail("INVALID_INPUT", "number_format 계약이 유효하지 않습니다.")
    if any(ord(character) < 32 or ord(character) == 127 for character in value):
        raise _fail("INVALID_INPUT", "number_format에 제어 문자를 쓸 수 없습니다.")
    return value


def _iso_datetime(value: object, *, field: str) -> str:
    text = _text(value, field=field, maximum=64)
    candidate = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(candidate)
    except ValueError as exc:
        raise _fail("INVALID_INPUT", f"{field}는 ISO 8601 datetime이어야 합니다.") from exc
    if "T" not in text or parsed.tzinfo is None:
        raise _fail("INVALID_INPUT", f"{field}에는 timezone이 필요합니다.")
    return text


def _unquote_sheet(value: str) -> str:
    if value.startswith("'") or value.endswith("'"):
        if len(value) < 2 or not value.startswith("'") or not value.endswith("'"):
            raise _fail("UNSAFE_FORMULA", "formula sheet qualifier가 유효하지 않습니다.")
        inner = value[1:-1]
        index = 0
        decoded: list[str] = []
        while index < len(inner):
            if inner[index] == "'":
                if index + 1 >= len(inner) or inner[index + 1] != "'":
                    raise _fail("UNSAFE_FORMULA", "formula sheet quoting이 유효하지 않습니다.")
                decoded.append("'")
                index += 2
            else:
                decoded.append(inner[index])
                index += 1
        return "".join(decoded)
    return value


def _static_reference(value: str, *, sheet: str) -> tuple[int, frozenset[str]]:
    if "[" in value or "]" in value or "|" in value:
        raise _fail("UNSAFE_FORMULA", "external/structured formula reference는 허용되지 않습니다.")
    reference = value
    if "!" in reference:
        if reference.count("!") != 1:
            raise _fail("UNSAFE_FORMULA", "formula reference qualifier가 유효하지 않습니다.")
        qualifier, reference = reference.split("!", 1)
        if _unquote_sheet(qualifier).casefold() != sheet.casefold():
            raise _fail("UNSAFE_FORMULA", "formula는 같은 sheet만 참조할 수 있습니다.")
    if _STATIC_RANGE_RE.fullmatch(reference) is None:
        raise _fail("UNSAFE_FORMULA", "static A1 cell/range reference만 허용됩니다.")
    positions: list[tuple[int, int]] = []
    for endpoint in reference.split(":"):
        match = _CELL_RE.fullmatch(endpoint)
        assert match is not None
        column = column_index_from_string(match.group(1).upper())
        row = int(match.group(2))
        if column > MAX_EXCEL_COLUMNS or row > MAX_EXCEL_ROWS:
            raise _fail("UNSAFE_FORMULA", "formula reference가 Excel grid를 벗어났습니다.")
        positions.append((row, column))
    if len(positions) == 1:
        row, column = positions[0]
        return 1, frozenset({f"{get_column_letter(column)}{row}"})
    (first_row, first_column), (last_row, last_column) = positions
    minimum_row, maximum_row = sorted((first_row, last_row))
    minimum_column, maximum_column = sorted((first_column, last_column))
    footprint = (maximum_row - minimum_row + 1) * (maximum_column - minimum_column + 1)
    if footprint > MAX_FORMULA_REFERENCE_CELLS:
        raise _fail(
            "LIMIT_EXCEEDED",
            "formula reference footprint 상한을 초과했습니다.",
        )
    references = frozenset(
        f"{get_column_letter(column)}{row}"
        for row in range(minimum_row, maximum_row + 1)
        for column in range(minimum_column, maximum_column + 1)
    )
    return footprint, references


def _function_consumes_range(function: str, argument: int) -> bool:
    if function in _RANGE_REDUCERS:
        return True
    if function == "COUNTIF":
        return argument == 0
    if function == "COUNTIFS":
        return argument % 2 == 0
    if function == "SUMIF":
        return argument in {0, 2}
    if function == "SUMIFS":
        return argument == 0 or argument % 2 == 1
    if function in {"MATCH", "VLOOKUP"}:
        return argument == 1
    # INDEX and XLOOKUP are intentionally absent: without a full Excel evaluator their return
    # shape cannot be proven scalar for every valid argument combination.
    return False


def _analyze_formula(value: object, *, sheet: str) -> tuple[str, int, frozenset[str]]:
    if (
        not isinstance(value, str)
        or not value.startswith("=")
        or value != value.strip()
        or len(value) > MAX_FORMULA_LENGTH
        or len(value) == 1
    ):
        raise _fail("UNSAFE_FORMULA", "formula는 locale-neutral =A1 형식이어야 합니다.")
    try:
        tokens = Tokenizer(value).items
    except (TokenizerError, IndexError, TypeError, ValueError) as exc:
        raise _fail("UNSAFE_FORMULA", "formula를 안전하게 분석할 수 없습니다.") from exc
    if not tokens:
        raise _fail("UNSAFE_FORMULA", "formula token이 없습니다.")
    function_stack: list[list[object]] = []
    referenced_cells = 0
    static_references: set[str] = set()
    for token in tokens:
        token_type = token.type
        subtype = token.subtype
        token_value = token.value
        if token_type == "FUNC":
            if subtype == "CLOSE" and token_value == ")":
                if not function_stack:
                    raise _fail("UNSAFE_FORMULA", "formula function nesting이 유효하지 않습니다.")
                function_stack.pop()
                continue
            if subtype != "OPEN" or not token_value.endswith("("):
                raise _fail("UNSAFE_FORMULA", "formula function token이 유효하지 않습니다.")
            function = token_value[:-1].upper()
            if function in _DISALLOWED_FUNCTIONS or function not in _ALLOWED_FUNCTIONS:
                raise _fail("UNSAFE_FORMULA", "허용되지 않은 formula function입니다.")
            function_stack.append([function, 0])
            continue
        if token_type == "OPERAND":
            if subtype == "RANGE":
                footprint, references = _static_reference(token_value, sheet=sheet)
                if footprint > 1 and (
                    not function_stack
                    or not _function_consumes_range(
                        str(function_stack[-1][0]),
                        int(function_stack[-1][1]),
                    )
                ):
                    raise _fail(
                        "UNSAFE_FORMULA",
                        "multi-cell range는 scalar consumer 내부에서만 허용됩니다.",
                    )
                referenced_cells += footprint
                static_references.update(references)
                if referenced_cells > MAX_FORMULA_REFERENCE_CELLS:
                    raise _fail(
                        "LIMIT_EXCEEDED",
                        "formula reference footprint 상한을 초과했습니다.",
                    )
            elif subtype not in {"NUMBER", "TEXT", "LOGICAL", "ERROR"}:
                raise _fail("UNSAFE_FORMULA", "formula operand가 허용되지 않습니다.")
            continue
        if token_type == "SEP":
            if subtype != "ARG" or token_value != ",":
                raise _fail("UNSAFE_FORMULA", "locale-neutral comma argument만 허용됩니다.")
            if not function_stack:
                raise _fail("UNSAFE_FORMULA", "formula argument nesting이 유효하지 않습니다.")
            function_stack[-1][1] = int(function_stack[-1][1]) + 1
            continue
        if token_type == "OPERATOR-INFIX":
            if token_value not in _ALLOWED_INFIX:
                raise _fail("UNSAFE_FORMULA", "formula operator가 허용되지 않습니다.")
            continue
        if token_type == "OPERATOR-PREFIX":
            if token_value not in {"+", "-"}:
                raise _fail("UNSAFE_FORMULA", "formula prefix operator가 허용되지 않습니다.")
            continue
        if token_type == "OPERATOR-POSTFIX":
            if token_value != "%":
                raise _fail("UNSAFE_FORMULA", "formula postfix operator가 허용되지 않습니다.")
            continue
        if token_type == "PAREN" and (
            subtype == "OPEN" and token_value == "(" or subtype == "CLOSE" and token_value == ")"
        ):
            continue
        raise _fail("UNSAFE_FORMULA", "formula token이 허용되지 않습니다.")
    if function_stack:
        raise _fail("UNSAFE_FORMULA", "formula function nesting이 유효하지 않습니다.")
    return value, referenced_cells, frozenset(static_references)


@dataclass(frozen=True)
class _FormulaAnalysis:
    footprint: int
    references: frozenset[str]


def _observed_formula(value: object) -> str:
    if (
        not isinstance(value, str)
        or not value.startswith("=")
        or value != value.strip()
        or len(value) > MAX_FORMULA_LENGTH
    ):
        raise _fail("INVALID_INPUT", "observed formula 계약이 유효하지 않습니다.")
    return value


def _scope(sheet: str) -> dict[str, object]:
    return {
        "kind": "sheet",
        "sheet": sheet,
        "id": hashlib.sha256(sheet.encode("utf-8")).hexdigest(),
    }


def _binding(
    *,
    bundle_id: object,
    snapshot_id: object,
    workbook_sha256: object,
    sheet: object,
) -> dict[str, object]:
    clean_sheet = _sheet(sheet)
    return {
        "bundle_id": _opaque(bundle_id, field="bundle_id"),
        "snapshot_id": _sha256(snapshot_id, field="snapshot_id"),
        "workbook_sha256": _sha256(workbook_sha256, field="workbook_sha256"),
        "scope": _scope(clean_sheet),
    }


def _office_binding(
    *,
    session_id: object,
    revision_id: object,
    worksheet_id: object,
    sheet: str,
) -> dict[str, object]:
    return {
        "session_id": _opaque(session_id, field="office.session_id"),
        "revision_id": _opaque(revision_id, field="office.revision_id"),
        "worksheet_id": _opaque(worksheet_id, field="office.worksheet_id"),
        "sheet": sheet,
    }


def _change(
    value: object,
    *,
    binding: Mapping[str, object],
) -> tuple[dict, tuple[int, int], _FormulaAnalysis | None]:
    if not isinstance(value, Mapping):
        raise _fail("INVALID_INPUT", "change는 객체여야 합니다.")
    fields = set(value)
    supplied_ref = value.get("change_ref")
    payload_fields = fields - {"change_ref"}
    kind = value.get("kind")
    if kind not in _FORMULA_KINDS:
        raise _fail("INVALID_INPUT", "지원하지 않는 workbook edit kind입니다.")
    cell, position = _cell(value.get("cell"))
    clean: dict[str, object] = {"cell": cell, "kind": kind}
    formula_analysis: _FormulaAnalysis | None = None
    if kind == "set_value":
        if payload_fields != {"cell", "kind", "value"}:
            raise _fail("INVALID_INPUT", "set_value 필드가 유효하지 않습니다.")
        clean["value"] = _literal_value(value.get("value"))
    elif kind == "set_formula":
        if payload_fields != {"cell", "kind", "formula"}:
            raise _fail("INVALID_INPUT", "set_formula 필드가 유효하지 않습니다.")
        sheet = binding["scope"]["sheet"]
        formula, footprint, references = _analyze_formula(
            value.get("formula"),
            sheet=str(sheet),
        )
        clean["formula"] = formula
        formula_analysis = _FormulaAnalysis(footprint, references)
    elif kind == "set_number_format":
        if payload_fields != {"cell", "kind", "number_format"}:
            raise _fail("INVALID_INPUT", "set_number_format 필드가 유효하지 않습니다.")
        clean["number_format"] = _number_format(value.get("number_format"))
    else:
        if payload_fields != {"cell", "kind"}:
            raise _fail("INVALID_INPUT", "clear_contents 필드가 유효하지 않습니다.")
    expected_ref = "edit-change:" + json_sha256(
        {"binding": binding, "change": clean}
    )
    if supplied_ref is not None and supplied_ref != expected_ref:
        raise _fail("CONTRACT_MISMATCH", "change_ref digest가 payload와 다릅니다.")
    clean["change_ref"] = expected_ref
    return clean, position, formula_analysis


def _changes(value: object, *, binding: Mapping[str, object]) -> list[dict]:
    if (
        not isinstance(value, Sequence)
        or isinstance(value, (str, bytes, bytearray))
        or not 1 <= len(value) <= MAX_CHANGES
    ):
        raise _fail("INVALID_INPUT", f"changes는 1~{MAX_CHANGES}건이어야 합니다.")
    cleaned: list[tuple[tuple[int, int], dict]] = []
    formula_analyses: dict[str, _FormulaAnalysis] = {}
    proposal_footprint = 0
    for raw in value:
        item, position, formula_analysis = _change(raw, binding=binding)
        if formula_analysis is not None:
            proposal_footprint += formula_analysis.footprint
            if proposal_footprint > MAX_PROPOSAL_REFERENCE_CELLS:
                raise _fail(
                    "LIMIT_EXCEEDED",
                    "proposal formula reference footprint 상한을 초과했습니다.",
                )
            formula_analyses[str(item["cell"])] = formula_analysis
        cleaned.append((position, item))
    cells = [item["cell"] for _, item in cleaned]
    if len(cells) != len(set(cells)):
        raise _fail("DUPLICATE_CELL", "한 proposal에서 같은 cell을 두 번 편집할 수 없습니다.")
    cleaned.sort(key=lambda pair: (pair[0][0], pair[0][1], pair[1]["kind"]))
    result = [item for _, item in cleaned]
    edit_cells = {str(item["cell"]) for item in result}
    formula_graph: dict[str, frozenset[str]] = {}
    for item in result:
        if item["kind"] != "set_formula":
            continue
        references = formula_analyses[str(item["cell"])].references
        formula_graph[str(item["cell"])] = frozenset(references.intersection(edit_cells))
    _assert_acyclic_formula_graph(formula_graph)
    _bounded_payload({"changes": result}, field="proposal changes")
    return result


def _assert_acyclic_formula_graph(graph: Mapping[str, frozenset[str]]) -> None:
    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(cell: str) -> None:
        if cell in visiting:
            raise _fail(
                "UNSAFE_FORMULA",
                "proposal-local formula cycle은 허용되지 않습니다.",
            )
        if cell in visited:
            return
        visiting.add(cell)
        for dependency in sorted(graph.get(cell, ())):
            if dependency in graph:
                visit(dependency)
        visiting.remove(cell)
        visited.add(cell)

    for cell in sorted(graph):
        visit(cell)


def _authored(value: object) -> dict[str, object]:
    if not isinstance(value, Mapping):
        raise _fail("INVALID_INPUT", "authored state는 객체여야 합니다.")
    kind = value.get("kind")
    fields = set(value)
    if kind == "blank" and fields == {"kind"}:
        return {"kind": "blank"}
    if kind == "value" and fields == {"kind", "value"}:
        return {
            "kind": "value",
            "value": _scalar(value.get("value"), field="authored.value", allow_null=False),
        }
    if kind == "formula" and fields == {"kind", "formula"}:
        formula = _observed_formula(value.get("formula"))
        return {"kind": "formula", "formula": formula}
    raise _fail("INVALID_INPUT", "authored state kind/필드가 유효하지 않습니다.")


def _before_state(value: object) -> tuple[dict, tuple[int, int]]:
    if not isinstance(value, Mapping) or set(value) != {
        "cell", "authored", "calculated_value", "calculated_type",
        "number_format", "target_constraints"
    }:
        raise _fail("INVALID_INPUT", "before state 필드가 유효하지 않습니다.")
    cell, position = _cell(value.get("cell"))
    authored = _authored(value.get("authored"))
    calculated = _scalar(
        value.get("calculated_value"), field="calculated_value", allow_null=True
    )
    calculated_type = value.get("calculated_type")
    if calculated_type not in {"empty", "string", "number", "boolean", "error"}:
        raise _fail("INVALID_INPUT", "calculated_type이 유효하지 않습니다.")
    type_matches = (
        calculated_type == "empty" and calculated is None
        or calculated_type in {"string", "error"} and isinstance(calculated, str)
        or calculated_type == "boolean" and isinstance(calculated, bool)
        or calculated_type == "number"
        and isinstance(calculated, (int, float))
        and not isinstance(calculated, bool)
    )
    if not type_matches:
        raise _fail("INVALID_INPUT", "calculated_type/value가 일치하지 않습니다.")
    if calculated_type == "error" and not calculated.startswith("#"):
        raise _fail("INVALID_INPUT", "error calculated_value는 Excel error 문자열이어야 합니다.")
    constraints = value.get("target_constraints")
    if not isinstance(constraints, Mapping) or set(constraints) != {
        "merged", "spill", "protected", "table_member"
    }:
        raise _fail("INVALID_INPUT", "target_constraints 필드가 유효하지 않습니다.")
    if (
        not isinstance(constraints.get("merged"), bool)
        or constraints.get("spill") not in {"none", "parent", "child"}
        or not isinstance(constraints.get("protected"), bool)
        or not isinstance(constraints.get("table_member"), bool)
    ):
        raise _fail("INVALID_INPUT", "target_constraints 값이 유효하지 않습니다.")
    if authored["kind"] == "blank" and (
        calculated is not None or calculated_type != "empty"
    ):
        raise _fail("INVALID_INPUT", "blank cell의 calculated_value는 null이어야 합니다.")
    if authored["kind"] == "value" and not _json_equal(calculated, authored["value"]):
        raise _fail("INVALID_INPUT", "value cell의 authored/calculated 값이 다릅니다.")
    if authored["kind"] == "value":
        expected_type = (
            "boolean" if isinstance(authored["value"], bool)
            else "number" if isinstance(authored["value"], (int, float))
            else "string"
        )
        if calculated_type != expected_type:
            raise _fail("INVALID_INPUT", "value cell의 calculated_type이 authored와 다릅니다.")
    return {
        "cell": cell,
        "authored": authored,
        "calculated_value": calculated,
        "calculated_type": calculated_type,
        "number_format": _number_format(value.get("number_format")),
        "target_constraints": {
            "merged": constraints["merged"],
            "spill": constraints["spill"],
            "protected": constraints["protected"],
            "table_member": constraints["table_member"],
        },
    }, position


def _before_states(value: object, *, cells: list[str]) -> list[dict]:
    if (
        not isinstance(value, Sequence)
        or isinstance(value, (str, bytes, bytearray))
        or len(value) != len(cells)
    ):
        raise _fail("INVALID_INPUT", "before state가 proposal cell과 일치하지 않습니다.")
    cleaned = [_before_state(item) for item in value]
    found = [item[0]["cell"] for item in cleaned]
    if len(found) != len(set(found)) or set(found) != set(cells):
        raise _fail("INVALID_INPUT", "before state cell binding이 proposal과 다릅니다.")
    cleaned.sort(key=lambda pair: pair[1])
    return [item for item, _ in cleaned]


def _authored_projection(state: Mapping[str, object]) -> dict[str, object]:
    return {
        "cell": state["cell"],
        "authored": copy.deepcopy(state["authored"]),
        "number_format": state["number_format"],
    }


def _authored_cell_state(value: object) -> tuple[dict[str, object], tuple[int, int]]:
    if not isinstance(value, Mapping) or set(value) != {
        "cell", "authored", "number_format"
    }:
        raise _fail("CONTRACT_MISMATCH", "expected authored state 필드가 유효하지 않습니다.")
    cell, position = _cell(value.get("cell"))
    return {
        "cell": cell,
        "authored": _authored(value.get("authored")),
        "number_format": _number_format(value.get("number_format")),
    }, position


def _validate_exact_diff(
    *,
    binding: Mapping[str, object],
    before: Sequence[Mapping[str, object]],
    expected_after: object,
    diff: object,
) -> tuple[list[dict], list[dict], list[dict]]:
    cells = [item.get("cell") for item in before]
    clean_before = _before_states(before, cells=cells)
    if (
        not isinstance(expected_after, Sequence)
        or isinstance(expected_after, (str, bytes, bytearray))
        or not isinstance(diff, Sequence)
        or isinstance(diff, (str, bytes, bytearray))
        or len(expected_after) != len(clean_before)
        or len(diff) != len(clean_before)
    ):
        raise _fail("CONTRACT_MISMATCH", "expected/diff count가 before와 다릅니다.")
    clean_expected_pairs = [_authored_cell_state(item) for item in expected_after]
    clean_expected_pairs.sort(key=lambda pair: pair[1])
    clean_expected = [item for item, _ in clean_expected_pairs]
    if [item["cell"] for item in clean_expected] != [item["cell"] for item in clean_before]:
        raise _fail("CONTRACT_MISMATCH", "expected cell binding이 before와 다릅니다.")
    by_before = {item["cell"]: item for item in clean_before}
    by_after = {item["cell"]: item for item in clean_expected}
    clean_diff: list[dict] = []
    seen_refs: set[str] = set()
    for raw in diff:
        if not isinstance(raw, Mapping) or set(raw) != {
            "change_ref", "cell", "kind", "before", "after"
        }:
            raise _fail("CONTRACT_MISMATCH", "diff item 필드가 유효하지 않습니다.")
        cell, position = _cell(raw.get("cell"))
        kind = raw.get("kind")
        if kind not in _FORMULA_KINDS or cell not in by_before:
            raise _fail("CONTRACT_MISMATCH", "diff kind/cell binding이 유효하지 않습니다.")
        before_item, _ = _before_state(raw.get("before"))
        after_item, _ = _authored_cell_state(raw.get("after"))
        if (
            not _json_equal(before_item, by_before[cell])
            or not _json_equal(after_item, by_after[cell])
        ):
            raise _fail("CONTRACT_MISMATCH", "diff before/after가 exact state와 다릅니다.")
        if _json_equal(_authored_projection(before_item), after_item):
            raise _fail("NO_OP_EDIT", "exact diff에 no-op edit가 있습니다.")
        if kind == "set_value":
            authored = after_item["authored"]
            if (
                authored["kind"] != "value"
                or not _json_equal(_literal_value(authored["value"]), authored["value"])
                or after_item["number_format"] != before_item["number_format"]
            ):
                raise _fail("CONTRACT_MISMATCH", "set_value exact diff가 유효하지 않습니다.")
            change = {"change_ref": raw.get("change_ref"), "cell": cell, "kind": kind, "value": authored["value"]}
        elif kind == "set_formula":
            authored = after_item["authored"]
            if (
                authored["kind"] != "formula"
                or after_item["number_format"] != before_item["number_format"]
            ):
                raise _fail("CONTRACT_MISMATCH", "set_formula exact diff가 유효하지 않습니다.")
            change = {"change_ref": raw.get("change_ref"), "cell": cell, "kind": kind, "formula": authored["formula"]}
        elif kind == "set_number_format":
            if not _json_equal(after_item["authored"], before_item["authored"]):
                raise _fail("CONTRACT_MISMATCH", "format edit가 cell content를 바꿀 수 없습니다.")
            change = {
                "change_ref": raw.get("change_ref"), "cell": cell, "kind": kind,
                "number_format": after_item["number_format"],
            }
        else:
            if (
                after_item["authored"] != {"kind": "blank"}
                or after_item["number_format"] != before_item["number_format"]
            ):
                raise _fail("CONTRACT_MISMATCH", "clear_contents exact diff가 유효하지 않습니다.")
            change = {"change_ref": raw.get("change_ref"), "cell": cell, "kind": kind}
        canonical_change, _, _ = _change(change, binding=binding)
        if canonical_change["change_ref"] in seen_refs:
            raise _fail("CONTRACT_MISMATCH", "diff change_ref가 중복되었습니다.")
        seen_refs.add(canonical_change["change_ref"])
        clean_diff.append({
            "change_ref": canonical_change["change_ref"],
            "cell": cell,
            "kind": kind,
            "before": before_item,
            "after": after_item,
            "_position": position,
        })
    clean_diff.sort(key=lambda item: item.pop("_position"))
    if [item["cell"] for item in clean_diff] != [item["cell"] for item in clean_before]:
        raise _fail("CONTRACT_MISMATCH", "diff cell set이 before와 다릅니다.")
    return clean_before, clean_expected, clean_diff


def _assert_safe_targets(states: Sequence[Mapping[str, object]]) -> None:
    for state in states:
        constraints = state["target_constraints"]
        if (
            constraints["merged"] is not False
            or constraints["spill"] != "none"
            or constraints["protected"] is not False
            or constraints["table_member"] is not False
        ):
            raise _fail(
                "UNSAFE_TARGET",
                "merged/spill/protected/table target은 v1 edit에서 허용되지 않습니다.",
            )


def _targets_are_safe(states: Sequence[Mapping[str, object]]) -> bool:
    return all(
        state["target_constraints"]["merged"] is False
        and state["target_constraints"]["spill"] == "none"
        and state["target_constraints"]["protected"] is False
        and state["target_constraints"]["table_member"] is False
        for state in states
    )


def _expected(change: Mapping[str, object], before: Mapping[str, object]) -> dict[str, object]:
    expected = _authored_projection(before)
    kind = change["kind"]
    if kind == "set_value":
        expected["authored"] = {"kind": "value", "value": copy.deepcopy(change["value"])}
    elif kind == "set_formula":
        expected["authored"] = {"kind": "formula", "formula": change["formula"]}
    elif kind == "set_number_format":
        expected["number_format"] = change["number_format"]
    elif kind == "clear_contents":
        expected["authored"] = {"kind": "blank"}
    return expected


def _schema_version(document: object, expected: str) -> Mapping[str, object]:
    if not isinstance(document, Mapping) or document.get("schema_version") != expected:
        raise _fail("CONTRACT_MISMATCH", f"{expected} 문서가 아닙니다.")
    return document


@lru_cache(maxsize=6)
def _schema(version: str) -> dict[str, object]:
    try:
        filename = _SCHEMA_FILES[version]
        value = json.loads((SCHEMA_DIR / filename).read_text(encoding="utf-8"))
    except (KeyError, OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise _fail("CONTRACT_MISMATCH", f"{version} schema를 읽을 수 없습니다.") from exc
    if not isinstance(value, dict):
        raise _fail("CONTRACT_MISMATCH", f"{version} schema가 객체가 아닙니다.")
    return value


def _validate_schema(document: object, version: str) -> None:
    try:
        jsonschema.validate(document, _schema(version))
    except (jsonschema.ValidationError, jsonschema.SchemaError) as exc:
        raise _fail("CONTRACT_MISMATCH", f"{version} schema 검증에 실패했습니다.") from exc


def _trust_fields(document: Mapping[str, object]) -> None:
    if (
        document.get("status") != "proposed"
        or document.get("review_status") != "unreviewed"
        or document.get("application_status") != "not_applied"
        or document.get("outside_prepared_bundle") is not True
    ):
        raise _fail("CONTRACT_MISMATCH", "workbook edit trust status가 유효하지 않습니다.")


def _proposal_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": PROPOSAL_SCHEMA,
        "binding": document["binding"],
        "changes": document["changes"],
    }


def _validate_proposal(value: object) -> dict:
    document = copy.deepcopy(dict(_schema_version(value, PROPOSAL_SCHEMA)))
    _validate_schema(document, PROPOSAL_SCHEMA)
    _trust_fields(document)
    binding = _binding(
        bundle_id=document["binding"]["bundle_id"],
        snapshot_id=document["binding"]["snapshot_id"],
        workbook_sha256=document["binding"]["workbook_sha256"],
        sheet=document["binding"]["scope"]["sheet"],
    )
    changes = _changes(document["changes"], binding=binding)
    digest = json_sha256(_proposal_identity({"binding": binding, "changes": changes}))
    if (
        not _json_equal(binding, document["binding"])
        or not _json_equal(changes, document["changes"])
        or document["proposal_sha256"] != digest
        or document["proposal_ref"] != "edit-proposal:" + digest
    ):
        raise _fail("CONTRACT_MISMATCH", "proposal digest/binding 계약이 다릅니다.")
    return document


def create_edit_proposal(
    *,
    bundle_id: str,
    snapshot_id: str,
    workbook_sha256: str,
    sheet: str,
    changes: Sequence[Mapping[str, object]],
) -> dict:
    """Create one deterministic proposed/unreviewed/not-applied one-sheet edit intent."""

    binding = _binding(
        bundle_id=bundle_id,
        snapshot_id=snapshot_id,
        workbook_sha256=workbook_sha256,
        sheet=sheet,
    )
    clean_changes = _changes(changes, binding=binding)
    body = {"binding": binding, "changes": clean_changes}
    digest = json_sha256(_proposal_identity(body))
    document = {
        "schema_version": PROPOSAL_SCHEMA,
        "proposal_ref": "edit-proposal:" + digest,
        "proposal_sha256": digest,
        "status": "proposed",
        "review_status": "unreviewed",
        "application_status": "not_applied",
        "outside_prepared_bundle": True,
        **body,
    }
    return _validate_proposal(document)


def _preview_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": PREVIEW_SCHEMA,
        "proposal_ref": document["proposal_ref"],
        "proposal_sha256": document["proposal_sha256"],
        "binding": document["binding"],
        "office_binding": document["office_binding"],
        "before_sha256": document["before_sha256"],
        "expected_after_sha256": document["expected_after_sha256"],
        "diff_sha256": document["diff_sha256"],
    }


def _validate_preview(value: object, *, proposal: object | None = None) -> dict:
    document = copy.deepcopy(dict(_schema_version(value, PREVIEW_SCHEMA)))
    _validate_schema(document, PREVIEW_SCHEMA)
    _trust_fields(document)
    raw_binding = document["binding"]
    canonical_binding = _binding(
        bundle_id=raw_binding["bundle_id"],
        snapshot_id=raw_binding["snapshot_id"],
        workbook_sha256=raw_binding["workbook_sha256"],
        sheet=raw_binding["scope"]["sheet"],
    )
    raw_office = document["office_binding"]
    canonical_office = _office_binding(
        session_id=raw_office["session_id"],
        revision_id=raw_office["revision_id"],
        worksheet_id=raw_office["worksheet_id"],
        sheet=str(canonical_binding["scope"]["sheet"]),
    )
    if not _json_equal(canonical_binding, raw_binding) or not _json_equal(
        canonical_office, raw_office
    ):
        raise _fail("CONTRACT_MISMATCH", "preview binding이 canonical하지 않습니다.")
    canonical_before, canonical_after, canonical_diff = _validate_exact_diff(
        binding=canonical_binding,
        before=document["before"],
        expected_after=document["expected_after"],
        diff=document["diff"],
    )
    if (
        not _json_equal(canonical_before, document["before"])
        or not _json_equal(canonical_after, document["expected_after"])
        or not _json_equal(canonical_diff, document["diff"])
    ):
        raise _fail("CONTRACT_MISMATCH", "preview exact diff가 canonical하지 않습니다.")
    _assert_safe_targets(canonical_before)
    before_sha = json_sha256(document["before"])
    after_sha = json_sha256(document["expected_after"])
    diff_sha = json_sha256(document["diff"])
    _bounded_payload(
        {
            "before": document["before"],
            "expected_after": document["expected_after"],
            "diff": document["diff"],
        },
        field="preview payload",
    )
    preview_sha = json_sha256(_preview_identity(document))
    if (
        document["before_sha256"] != before_sha
        or document["expected_after_sha256"] != after_sha
        or document["diff_sha256"] != diff_sha
        or document["preview_sha256"] != preview_sha
        or document["preview_ref"] != "edit-preview:" + preview_sha
        or document["office_binding"]["sheet"] != document["binding"]["scope"]["sheet"]
    ):
        raise _fail("CONTRACT_MISMATCH", "preview digest/binding 계약이 다릅니다.")
    if proposal is not None:
        clean_proposal = _validate_proposal(proposal)
        if (
            document["proposal_ref"] != clean_proposal["proposal_ref"]
            or document["proposal_sha256"] != clean_proposal["proposal_sha256"]
            or not _json_equal(document["binding"], clean_proposal["binding"])
        ):
            raise _fail("CONTRACT_MISMATCH", "preview가 proposal과 연결되지 않습니다.")
        expected_cells = [item["cell"] for item in clean_proposal["changes"]]
        before = _before_states(document["before"], cells=expected_cells)
        _assert_safe_targets(before)
        by_cell = {item["cell"]: item for item in before}
        expected_after = [
            _expected(change, by_cell[change["cell"]])
            for change in clean_proposal["changes"]
        ]
        diff = [
            {
                "change_ref": change["change_ref"],
                "cell": change["cell"],
                "kind": change["kind"],
                "before": copy.deepcopy(by_cell[change["cell"]]),
                "after": copy.deepcopy(expected),
            }
            for change, expected in zip(clean_proposal["changes"], expected_after)
        ]
        if (
            not _json_equal(before, document["before"])
            or not _json_equal(expected_after, document["expected_after"])
            or not _json_equal(diff, document["diff"])
        ):
            raise _fail("CONTRACT_MISMATCH", "preview exact diff가 proposal과 다릅니다.")
    return document


def create_edit_preview(
    proposal: Mapping[str, object],
    *,
    office_session_id: str,
    office_revision_id: str,
    worksheet_id: str,
    before: Sequence[Mapping[str, object]],
) -> dict:
    """Bind a proposal to an exact Office session readback and materialize its diff."""

    clean_proposal = _validate_proposal(proposal)
    sheet = clean_proposal["binding"]["scope"]["sheet"]
    office = _office_binding(
        session_id=office_session_id,
        revision_id=office_revision_id,
        worksheet_id=worksheet_id,
        sheet=sheet,
    )
    cells = [item["cell"] for item in clean_proposal["changes"]]
    clean_before = _before_states(before, cells=cells)
    _assert_safe_targets(clean_before)
    by_cell = {item["cell"]: item for item in clean_before}
    expected_after = [
        _expected(change, by_cell[change["cell"]])
        for change in clean_proposal["changes"]
    ]
    if any(
        _json_equal(_authored_projection(by_cell[change["cell"]]), expected)
        for change, expected in zip(clean_proposal["changes"], expected_after)
    ):
        raise _fail("NO_OP_EDIT", "현재 Office state를 바꾸지 않는 edit가 포함되어 있습니다.")
    diff = [
        {
            "change_ref": change["change_ref"],
            "cell": change["cell"],
            "kind": change["kind"],
            "before": copy.deepcopy(by_cell[change["cell"]]),
            "after": copy.deepcopy(expected),
        }
        for change, expected in zip(clean_proposal["changes"], expected_after)
    ]
    document = {
        "schema_version": PREVIEW_SCHEMA,
        "preview_ref": "",
        "preview_sha256": "",
        "proposal_ref": clean_proposal["proposal_ref"],
        "proposal_sha256": clean_proposal["proposal_sha256"],
        "status": "proposed",
        "review_status": "unreviewed",
        "application_status": "not_applied",
        "outside_prepared_bundle": True,
        "binding": copy.deepcopy(clean_proposal["binding"]),
        "office_binding": office,
        "before": clean_before,
        "before_sha256": json_sha256(clean_before),
        "expected_after": expected_after,
        "expected_after_sha256": json_sha256(expected_after),
        "diff": diff,
        "diff_sha256": json_sha256(diff),
    }
    digest = json_sha256(_preview_identity(document))
    document["preview_ref"] = "edit-preview:" + digest
    document["preview_sha256"] = digest
    return _validate_preview(document, proposal=clean_proposal)


def _approval_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": APPROVAL_SCHEMA,
        "proposal_ref": document["proposal_ref"],
        "preview_ref": document["preview_ref"],
        "preview_sha256": document["preview_sha256"],
        "approver_id": document["approver_id"],
        "expires_at": document["expires_at"],
        "binding": document["binding"],
        "office_binding": document["office_binding"],
    }


def _validate_approval(value: object, *, preview: object | None = None) -> dict:
    document = copy.deepcopy(dict(_schema_version(value, APPROVAL_SCHEMA)))
    _validate_schema(document, APPROVAL_SCHEMA)
    _trust_fields(document)
    if document.get("approval_status") != "approved":
        raise _fail("CONTRACT_MISMATCH", "approval status가 유효하지 않습니다.")
    if _iso_datetime(document.get("expires_at"), field="expires_at") != document["expires_at"]:
        raise _fail("CONTRACT_MISMATCH", "approval expiry가 canonical하지 않습니다.")
    digest = json_sha256(_approval_identity(document))
    if (
        document["approval_sha256"] != digest
        or document["approval_ref"] != "edit-approval:" + digest
    ):
        raise _fail("CONTRACT_MISMATCH", "approval digest 계약이 다릅니다.")
    if preview is not None:
        clean_preview = _validate_preview(preview)
        if any(
            document[field] != clean_preview[field]
            for field in (
                "proposal_ref", "preview_ref", "preview_sha256", "binding", "office_binding"
            )
        ):
            raise _fail("CONTRACT_MISMATCH", "approval이 exact preview와 다릅니다.")
    return document


def create_edit_approval(
    preview: Mapping[str, object],
    *,
    approver_id: str,
    expires_at: str,
) -> dict:
    """Create a human authorization bound to one exact preview digest."""

    clean_preview = _validate_preview(preview)
    document = {
        "schema_version": APPROVAL_SCHEMA,
        "approval_ref": "",
        "approval_sha256": "",
        "proposal_ref": clean_preview["proposal_ref"],
        "preview_ref": clean_preview["preview_ref"],
        "preview_sha256": clean_preview["preview_sha256"],
        "approver_id": _opaque(approver_id, field="approver_id"),
        "expires_at": _iso_datetime(expires_at, field="expires_at"),
        "approval_status": "approved",
        "status": "proposed",
        "review_status": "unreviewed",
        "application_status": "not_applied",
        "outside_prepared_bundle": True,
        "binding": copy.deepcopy(clean_preview["binding"]),
        "office_binding": copy.deepcopy(clean_preview["office_binding"]),
    }
    digest = json_sha256(_approval_identity(document))
    document["approval_ref"] = "edit-approval:" + digest
    document["approval_sha256"] = digest
    return _validate_approval(document, preview=clean_preview)


def _manifest_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": APPLY_MANIFEST_SCHEMA,
        "proposal_ref": document["proposal_ref"],
        "preview_ref": document["preview_ref"],
        "approval_ref": document["approval_ref"],
        "approval_expires_at": document["approval_expires_at"],
        "execution_id": document["execution_id"],
        "fencing_token": document["fencing_token"],
        "challenge_nonce": document["challenge_nonce"],
        "binding": document["binding"],
        "office_binding": document["office_binding"],
        "before_sha256": document["before_sha256"],
        "expected_after_sha256": document["expected_after_sha256"],
        "diff_sha256": document["diff_sha256"],
        "limitations": document["limitations"],
    }


def _validate_manifest(
    value: object,
    *,
    preview: object | None = None,
    approval: object | None = None,
) -> dict:
    document = copy.deepcopy(dict(_schema_version(value, APPLY_MANIFEST_SCHEMA)))
    _validate_schema(document, APPLY_MANIFEST_SCHEMA)
    _trust_fields(document)
    if document.get("limitations") != _EDIT_LIMITATIONS:
        raise _fail("CONTRACT_MISMATCH", "apply manifest limitation 계약이 다릅니다.")
    raw_binding = document["binding"]
    canonical_binding = _binding(
        bundle_id=raw_binding["bundle_id"],
        snapshot_id=raw_binding["snapshot_id"],
        workbook_sha256=raw_binding["workbook_sha256"],
        sheet=raw_binding["scope"]["sheet"],
    )
    raw_office = document["office_binding"]
    canonical_office = _office_binding(
        session_id=raw_office["session_id"],
        revision_id=raw_office["revision_id"],
        worksheet_id=raw_office["worksheet_id"],
        sheet=str(canonical_binding["scope"]["sheet"]),
    )
    if not _json_equal(canonical_binding, raw_binding) or not _json_equal(
        canonical_office, raw_office
    ):
        raise _fail("CONTRACT_MISMATCH", "apply manifest binding이 canonical하지 않습니다.")
    if (
        _opaque(document["execution_id"], field="execution_id") != document["execution_id"]
        or _opaque(document["challenge_nonce"], field="challenge_nonce")
        != document["challenge_nonce"]
        or not isinstance(document["fencing_token"], int)
        or isinstance(document["fencing_token"], bool)
        or document["fencing_token"] < 1
        or document["fencing_token"] > MAX_SAFE_INTEGER
        or _iso_datetime(document["approval_expires_at"], field="approval_expires_at")
        != document["approval_expires_at"]
    ):
        raise _fail("CONTRACT_MISMATCH", "apply manifest execution binding이 유효하지 않습니다.")
    canonical_before, canonical_after, canonical_diff = _validate_exact_diff(
        binding=canonical_binding,
        before=document["before"],
        expected_after=document["expected_after"],
        diff=document["diff"],
    )
    if (
        not _json_equal(canonical_before, document["before"])
        or not _json_equal(canonical_after, document["expected_after"])
        or not _json_equal(canonical_diff, document["diff"])
    ):
        raise _fail("CONTRACT_MISMATCH", "apply manifest exact diff가 canonical하지 않습니다.")
    _assert_safe_targets(canonical_before)
    if (
        document["before_sha256"] != json_sha256(document["before"])
        or document["expected_after_sha256"] != json_sha256(document["expected_after"])
        or document["diff_sha256"] != json_sha256(document["diff"])
    ):
        raise _fail("CONTRACT_MISMATCH", "apply manifest payload digest가 다릅니다.")
    digest = json_sha256(_manifest_identity(document))
    if (
        document["manifest_sha256"] != digest
        or document["manifest_ref"] != "edit-manifest:" + digest
    ):
        raise _fail("CONTRACT_MISMATCH", "apply manifest identity가 다릅니다.")
    if preview is not None:
        clean_preview = _validate_preview(preview)
        pairs = {
            "proposal_ref": "proposal_ref",
            "preview_ref": "preview_ref",
            "binding": "binding",
            "office_binding": "office_binding",
            "before": "before",
            "before_sha256": "before_sha256",
            "expected_after": "expected_after",
            "expected_after_sha256": "expected_after_sha256",
            "diff": "diff",
            "diff_sha256": "diff_sha256",
        }
        if any(
            not _json_equal(document[target], clean_preview[source])
            for target, source in pairs.items()
        ):
            raise _fail("CONTRACT_MISMATCH", "apply manifest가 preview와 다릅니다.")
    if approval is not None:
        clean_approval = _validate_approval(approval, preview=preview)
        if (
            document["approval_ref"] != clean_approval["approval_ref"]
            or document["approval_expires_at"] != clean_approval["expires_at"]
        ):
            raise _fail("CONTRACT_MISMATCH", "apply manifest approval이 다릅니다.")
    return document


def create_apply_manifest(
    preview: Mapping[str, object],
    approval: Mapping[str, object],
    *,
    execution_id: str,
    fencing_token: int,
    challenge_nonce: str,
) -> dict:
    """Materialize the immutable, exact command an Office.js executor may apply."""

    clean_preview = _validate_preview(preview)
    clean_approval = _validate_approval(approval, preview=clean_preview)
    if (
        not isinstance(fencing_token, int)
        or isinstance(fencing_token, bool)
        or fencing_token < 1
        or fencing_token > MAX_SAFE_INTEGER
    ):
        raise _fail("INVALID_INPUT", "fencing_token은 JavaScript safe 양의 정수여야 합니다.")
    document = {
        "schema_version": APPLY_MANIFEST_SCHEMA,
        "manifest_ref": "",
        "manifest_sha256": "",
        "proposal_ref": clean_preview["proposal_ref"],
        "preview_ref": clean_preview["preview_ref"],
        "approval_ref": clean_approval["approval_ref"],
        "approval_expires_at": clean_approval["expires_at"],
        "execution_id": _opaque(execution_id, field="execution_id"),
        "fencing_token": fencing_token,
        "challenge_nonce": _opaque(challenge_nonce, field="challenge_nonce"),
        "status": "proposed",
        "review_status": "unreviewed",
        "application_status": "not_applied",
        "outside_prepared_bundle": True,
        "binding": copy.deepcopy(clean_preview["binding"]),
        "office_binding": copy.deepcopy(clean_preview["office_binding"]),
        "before": copy.deepcopy(clean_preview["before"]),
        "before_sha256": clean_preview["before_sha256"],
        "expected_after": copy.deepcopy(clean_preview["expected_after"]),
        "expected_after_sha256": clean_preview["expected_after_sha256"],
        "diff": copy.deepcopy(clean_preview["diff"]),
        "diff_sha256": clean_preview["diff_sha256"],
        "limitations": list(_EDIT_LIMITATIONS),
    }
    digest = json_sha256(_manifest_identity(document))
    document["manifest_ref"] = "edit-manifest:" + digest
    document["manifest_sha256"] = digest
    return _validate_manifest(
        document,
        preview=clean_preview,
        approval=clean_approval,
    )


def _witness_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": WITNESS_SCHEMA,
        "manifest_ref": document["manifest_ref"],
        "manifest_sha256": document["manifest_sha256"],
        "execution_id": document["execution_id"],
        "fencing_token": document["fencing_token"],
        "challenge_nonce": document["challenge_nonce"],
        "office_binding": document["office_binding"],
        "executor_id": document["executor_id"],
        "outcome": document["outcome"],
        "write_started": document["write_started"],
        "recalculation": document["recalculation"],
        "observed_before_sha256": document["observed_before_sha256"],
        "actual_after_sha256": document["actual_after_sha256"],
    }


def _validate_witness(value: object, *, manifest: object | None = None) -> dict:
    document = copy.deepcopy(dict(_schema_version(value, WITNESS_SCHEMA)))
    _validate_schema(document, WITNESS_SCHEMA)
    before_sha = json_sha256(document["observed_before"])
    after = document["actual_after"]
    _bounded_payload(
        {"observed_before": document["observed_before"], "actual_after": after},
        field="execution witness payload",
    )
    after_sha = None if after is None else json_sha256(after)
    if (
        document["observed_before_sha256"] != before_sha
        or document["actual_after_sha256"] != after_sha
    ):
        raise _fail("CONTRACT_MISMATCH", "execution witness payload digest가 다릅니다.")
    digest = json_sha256(_witness_identity(document))
    if (
        document["witness_sha256"] != digest
        or document["witness_ref"] != "edit-witness:" + digest
    ):
        raise _fail("CONTRACT_MISMATCH", "execution witness identity가 다릅니다.")
    outcome = document["outcome"]
    if outcome == "stale_precondition":
        if document["write_started"] is not False or after is not None:
            raise _fail("CONTRACT_MISMATCH", "stale witness가 write를 주장할 수 없습니다.")
    elif outcome == "applied":
        if document["write_started"] is not True or after is None:
            raise _fail("CONTRACT_MISMATCH", "applied witness에는 reread state가 필요합니다.")
    elif outcome == "indeterminate":
        if document["write_started"] is not True:
            raise _fail("CONTRACT_MISMATCH", "indeterminate witness는 write 시작 뒤에만 허용됩니다.")
    if manifest is not None:
        clean_manifest = _validate_manifest(manifest)
        if (
            document["manifest_ref"] != clean_manifest["manifest_ref"]
            or document["manifest_sha256"] != clean_manifest["manifest_sha256"]
            or document["execution_id"] != clean_manifest["execution_id"]
            or document["fencing_token"] != clean_manifest["fencing_token"]
            or document["challenge_nonce"] != clean_manifest["challenge_nonce"]
            or not _json_equal(document["office_binding"], clean_manifest["office_binding"])
        ):
            raise _fail("CONTRACT_MISMATCH", "execution witness가 manifest와 다릅니다.")
        cells = [item["cell"] for item in clean_manifest["before"]]
        observed = _before_states(document["observed_before"], cells=cells)
        if not _json_equal(observed, document["observed_before"]):
            raise _fail("CONTRACT_MISMATCH", "execution before state가 canonical하지 않습니다.")
        if after is not None:
            actual = _before_states(after, cells=cells)
            if not _json_equal(actual, after):
                raise _fail("CONTRACT_MISMATCH", "execution after state가 canonical하지 않습니다.")
        if (
            outcome == "applied"
            and any(item["kind"] == "set_formula" for item in clean_manifest["diff"])
            and document["recalculation"] != "recalculate"
        ):
            raise _fail(
                "CONTRACT_MISMATCH",
                "set_formula applied witness에는 recalculation이 필요합니다.",
            )
    return document


def create_execution_witness(
    manifest: Mapping[str, object],
    *,
    executor_id: str,
    outcome: str,
    observed_before: Sequence[Mapping[str, object]],
    actual_after: Sequence[Mapping[str, object]] | None,
    recalculation: str = "recalculate",
) -> dict:
    """Create the bounded Office.js reread witness consumed by verification."""

    clean_manifest = _validate_manifest(manifest)
    if outcome not in {"applied", "stale_precondition", "indeterminate"}:
        raise _fail("INVALID_INPUT", "execution outcome이 유효하지 않습니다.")
    if recalculation not in {"none", "recalculate"}:
        raise _fail("INVALID_INPUT", "recalculation mode가 유효하지 않습니다.")
    cells = [item["cell"] for item in clean_manifest["before"]]
    clean_before = _before_states(observed_before, cells=cells)
    clean_after = None if actual_after is None else _before_states(actual_after, cells=cells)
    document = {
        "schema_version": WITNESS_SCHEMA,
        "witness_ref": "",
        "witness_sha256": "",
        "manifest_ref": clean_manifest["manifest_ref"],
        "manifest_sha256": clean_manifest["manifest_sha256"],
        "execution_id": clean_manifest["execution_id"],
        "fencing_token": clean_manifest["fencing_token"],
        "challenge_nonce": clean_manifest["challenge_nonce"],
        "office_binding": copy.deepcopy(clean_manifest["office_binding"]),
        "executor_id": _opaque(executor_id, field="executor_id"),
        "outcome": outcome,
        "write_started": outcome != "stale_precondition",
        "recalculation": recalculation,
        "observed_before": clean_before,
        "observed_before_sha256": json_sha256(clean_before),
        "actual_after": clean_after,
        "actual_after_sha256": None if clean_after is None else json_sha256(clean_after),
    }
    digest = json_sha256(_witness_identity(document))
    document["witness_ref"] = "edit-witness:" + digest
    document["witness_sha256"] = digest
    return _validate_witness(document, manifest=clean_manifest)


def _verification_identity(document: Mapping[str, object]) -> dict[str, object]:
    return {
        "schema_version": VERIFICATION_SCHEMA,
        "manifest_ref": document["manifest_ref"],
        "witness_ref": document["witness_ref"],
        "status": document["status"],
        "application_status": document["application_status"],
        "asset_persisted": document["asset_persisted"],
        "new_snapshot_required": document["new_snapshot_required"],
        "before_matches": document["before_matches"],
        "expected_after_matches": document["expected_after_matches"],
        "actual_after_sha256": document["actual_after_sha256"],
        "limitations": document["limitations"],
    }


def verify_execution_witness(
    manifest: Mapping[str, object],
    witness: Mapping[str, object],
) -> dict:
    """Verify precondition and authored after-state without treating calculated values as intent."""

    clean_manifest = _validate_manifest(manifest)
    clean_witness = _validate_witness(witness, manifest=clean_manifest)
    before_matches = (
        clean_witness["observed_before_sha256"] == clean_manifest["before_sha256"]
    )
    after_matches: bool | None = None
    if clean_witness["actual_after"] is not None:
        formula_targets = {
            item["cell"]
            for item in clean_manifest["diff"]
            if item["kind"] == "set_formula"
        }
        actual_by_cell = {
            item["cell"]: item for item in clean_witness["actual_after"]
        }
        formula_results_observed = all(
            actual_by_cell[cell]["calculated_type"] in {"string", "number", "boolean"}
            and actual_by_cell[cell]["calculated_value"] not in {None, ""}
            for cell in formula_targets
        )
        actual_authored = [
            _authored_projection(item) for item in clean_witness["actual_after"]
        ]
        after_matches = (
            _json_equal(actual_authored, clean_manifest["expected_after"])
            and _targets_are_safe(clean_witness["actual_after"])
            and formula_results_observed
        )
    outcome = clean_witness["outcome"]
    if outcome == "stale_precondition":
        if before_matches:
            raise _fail(
                "CONTRACT_MISMATCH",
                "stale_precondition witness가 동일 before digest를 보고했습니다.",
            )
        status = "stale_precondition"
        application_status = "not_applied"
    elif outcome == "indeterminate":
        status = "indeterminate"
        application_status = "indeterminate"
    elif before_matches and after_matches is True:
        status = "session_verified"
        application_status = "applied_session_verified"
    else:
        status = "verification_failed"
        application_status = "application_failed"
    new_snapshot_required = application_status != "not_applied"
    document = {
        "schema_version": VERIFICATION_SCHEMA,
        "verification_ref": "",
        "verification_sha256": "",
        "manifest_ref": clean_manifest["manifest_ref"],
        "witness_ref": clean_witness["witness_ref"],
        "status": status,
        "review_status": "unreviewed",
        "application_status": application_status,
        "asset_persisted": False,
        "new_snapshot_required": new_snapshot_required,
        "outside_prepared_bundle": True,
        "before_matches": before_matches,
        "expected_after_matches": after_matches,
        "actual_after_sha256": clean_witness["actual_after_sha256"],
        "limitations": list(_EDIT_LIMITATIONS),
    }
    digest = json_sha256(_verification_identity(document))
    document["verification_ref"] = "edit-verification:" + digest
    document["verification_sha256"] = digest
    _validate_schema(document, VERIFICATION_SCHEMA)
    if json_sha256(_verification_identity(document)) != digest:
        raise _fail("CONTRACT_MISMATCH", "verification digest 계약이 다릅니다.")
    return document


__all__ = [
    "APPLY_MANIFEST_SCHEMA",
    "APPROVAL_SCHEMA",
    "MAX_ARTIFACT_PAYLOAD_BYTES",
    "MAX_CHANGES",
    "MAX_FORMULA_REFERENCE_CELLS",
    "MAX_PROPOSAL_REFERENCE_CELLS",
    "MAX_SAFE_INTEGER",
    "PREVIEW_SCHEMA",
    "PROPOSAL_SCHEMA",
    "VERIFICATION_SCHEMA",
    "WITNESS_SCHEMA",
    "WorkbookEditError",
    "create_apply_manifest",
    "create_edit_approval",
    "create_edit_preview",
    "create_edit_proposal",
    "create_execution_witness",
    "verify_execution_witness",
]
