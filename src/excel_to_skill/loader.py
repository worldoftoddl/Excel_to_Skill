"""§5 로더 사다리 — 파일을 열어 원시 핸들을 반환한다.

1차  openpyxl 일반 로드 (이중: data_only=False / True)
2차  1차 예외 시 read_only=True 폴백 (동일 이중 로드)
3차  read_only에서 병합 정보 미취득 시 zipfile로 <mergeCells> 직파싱
4차  .xls → xlrd (formatting_info=True)

형식 분기는 EXTENSION_FORMATS 매핑으로 확장한다. (후속 단계에서 docx 로더 추가 예정)
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
import xlrd

EXTENSION_FORMATS = {".xlsx": "xlsx", ".xls": "xls"}

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


class UnsupportedFormatError(ValueError):
    """지원하지 않는 입력 형식. 힌트 메시지를 포함한다."""

    def __init__(self, path: Path):
        self.path = path
        super().__init__(
            f"지원하지 않는 형식입니다: {path.name} ({path.suffix or '확장자 없음'}). "
            "지원 형식: xlsx, xls. 변환 후 재시도하십시오."
        )


def detect_format(path: Path) -> str:
    fmt = EXTENSION_FORMATS.get(path.suffix.lower())
    if fmt is None:
        raise UnsupportedFormatError(path)
    return fmt


def open_xlsx_pair(path: Path):
    """이중 로드. 반환: (수식용 wb, 캐시값용 wb, 모드 'normal'|'read_only').

    §2 함정 1: 일반 로드가 일부 파일에서 예외를 던지므로 read_only 폴백은 필수.
    """
    wb_f = None
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False)
        wb_v = openpyxl.load_workbook(path, data_only=True)
        return wb_f, wb_v, "normal"
    except Exception:
        if wb_f is not None:
            wb_f.close()
    wb_f = openpyxl.load_workbook(path, read_only=True, data_only=False)
    wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return wb_f, wb_v, "read_only"


def open_xls(path: Path) -> xlrd.book.Book:
    return xlrd.open_workbook(str(path), formatting_info=True)


def merges_from_xml(path: Path) -> dict[str, list[str]]:
    """3차 안전망(§2 함정 2): sheet XML의 <mergeCells>를 직접 파싱한다.

    반환: 시트명 → 병합 범위 문자열 목록. 차트시트 등 mergeCells가 없는
    대상은 빈 목록.
    """
    result: dict[str, list[str]] = {}
    with zipfile.ZipFile(path) as zf:
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.get("Id"): rel.get("Target")
            for rel in rels_root.iter(f"{{{_NS_PKG_REL}}}Relationship")
        }
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        for sheet_el in wb_root.iter(f"{{{_NS_MAIN}}}sheet"):
            name = sheet_el.get("name")
            rid = sheet_el.get(f"{{{_NS_REL}}}id")
            target = rid_to_target.get(rid)
            if not target:
                result[name] = []
                continue
            if target.startswith("/"):
                target = target.lstrip("/")
            elif not target.startswith("xl/"):
                target = "xl/" + target
            try:
                sheet_root = ET.fromstring(zf.read(target))
            except KeyError:
                result[name] = []
                continue
            result[name] = [
                mc.get("ref")
                for mc in sheet_root.iter(f"{{{_NS_MAIN}}}mergeCell")
                if mc.get("ref")
            ]
    return result
